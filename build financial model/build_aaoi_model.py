#!/usr/bin/env python3
"""
AAOI (Applied Optoelectronics) model — same framework as MU / NBIS / SHOP / APP / SNDK / AXTI.

DATA CAVEAT: built from filings/aggregators, not the user's moomoo screenshots. Income
statement and guidance are exact; balance-sheet / cash-flow / TTM-ratio layers are absent.

TIMING: Q2 2026 reports TODAY, 2026-08-06, after the close (call 4:30pm ET).

THE CENTRAL FACT THIS MODEL IS BUILT AROUND: revenue has gone from $40.67M (Q1 2024) to
$151.14M (Q1 2026) — a 3.7x — and the company has posted an OPERATING LOSS in every single
one of those ten quarters. Gross margin has sat in a 28-31% band for six straight quarters
while revenue grew 51%. The bull case is entirely a bet that the 800G/1.6T mix shift finally
breaks gross margin out of that band. Everything else is secondary.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2", "2025/Q1",
        "2024/Q4", "2024/Q3", "2024/Q2", "2024/Q1", "2023/Q4"]
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Revenue", row(151.14*M, 134.27*M, 118.63*M, 102.95*M, 99.86*M,
                    100.27*M, 65.15*M, 43.27*M, 40.67*M, 60.45*M)),
    ("Gross Profit", row(43.92*M, 41.94*M, 33.26*M, 31.16*M, 30.54*M,
                         28.73*M, 15.92*M, 9.56*M, 7.59*M, 21.60*M)),
    ("Operating Income", row(-12.99*M, -11.50*M, -18.19*M, -15.98*M, -8.94*M,
                             -6.47*M, -16.55*M, -26.24*M, -21.65*M, -4.53*M)),
    ("Net Income", row(-14.28*M, -2.02*M, -17.94*M, -9.10*M, -9.17*M,
                       -119.69*M, -17.76*M, -26.12*M, -23.17*M, -13.86*M)),
    ("Diluted EPS", row(-0.19, -0.03, -0.28, -0.16, -0.18,
                        -2.60, -0.42, -0.66, -0.60, -0.38)),
]

# ============================================================== GUIDANCE =====
GUIDE = [
    ("Q2 2026 report", "2026-08-06 AFTER THE CLOSE; call 4:30pm ET", "company"),
    ("Q2 2026 revenue guidance", "$180M - $198M", "company"),
    ("Q2 2026 non-GAAP EPS guidance", "-$0.03 to +$0.03", "company — i.e. breakeven"),
    ("Q2 2026 consensus revenue", "$191.13M (+85.7% y/y)", "Zacks — mid-to-upper half of guidance"),
    ("Q2 2026 consensus EPS", "+$0.03 — the TOP of guidance", "Zacks, unchanged 30 days"),
    ("800G revenue, Q1 2026", "$4.6M — only 3.0% of Q1 revenue", "company"),
    ("800G units, Q2 2026", "expected ~4x the Q1 unit shipments", "company"),
    ("FY2025 revenue", "$455.72M (+82.8% y/y)", ""),
    ("Analyst mean price target", "$150.30 (+14.6%)", "stockanalysis"),
    ("Analyst consensus rating", "Buy", ""),
]

# ========================================================== MARKET DATA ======
PX_NOW = 131.14
SHARES = 80.24e6
MKT = [
    ("Last price (2026-08-06)", 131.14, "USD, +2.01%"),
    ("Market cap", 10.52*B, "USD"),
    ("Shares outstanding", 80.24e6, "shares"),
    ("52-week high", 233.67, "USD"),
    ("52-week low", 18.50, "USD"),
    ("P/E (TTM)", None, "n/a — unprofitable"),
    ("Forward P/E", 72.07, "x -> implies street forward EPS of ~$1.82"),
    ("TTM revenue", 507.00*M, "USD"),
    ("TTM EPS", -0.66, "USD"),
    ("Analyst mean target", 150.30, "USD (+14.6%)"),
]

# ============================================ Q2-2026 FORECAST ENGINE ========
Q1_REV = 151.14*M
Q1_GM = 43.92/151.14
Q1_OPEX = (43.92 + 12.99) * M          # GP less operating income = $56.91M GAAP opex
GUIDE_LO, GUIDE_HI = 180.0*M, 198.0*M
ST_REV, ST_EPS = 191.13*M, 0.03
# Non-GAAP opex runs below GAAP by roughly the SBC line; calibrated so that the guidance
# midpoint ($189M) with a ~31% gross margin lands on the guided breakeven EPS.
NONGAAP_OPEX_Q1 = 48.0*M
INTEREST = 4.0*M

SCEN = {
    "Bear": dict(rev=180.0*M, gm=0.295, opex=50.0*M, sh=81.0e6,
                 note="guide low; gross margin stays stuck in the 28-31% band"),
    "Base": dict(rev=195.0*M, gm=0.315, opex=51.0*M, sh=81.5e6,
                 note="upper half of guide; margin nudges up on early 800G mix"),
    "Bull": dict(rev=205.0*M, gm=0.335, opex=52.0*M, sh=82.0e6,
                 note="beat the guide; 800G mix finally moves gross margin"),
}


def forecast_q2(p):
    gp = p["rev"] * p["gm"]
    op = gp - p["opex"]
    net = op - INTEREST
    return dict(revenue=p["rev"], qoq=p["rev"]/Q1_REV-1, yoy=p["rev"]/(102.95*M)-1,
                gross_profit=gp, gross_margin=p["gm"], opex=p["opex"],
                op_income=op, op_margin=op/p["rev"], net=net, eps=net/p["sh"],
                vs_guide_mid=p["rev"]/(189.0*M)-1)


Q2 = {k: forecast_q2(v) for k, v in SCEN.items()}

# ================================= FY2026-2027 MODEL + VALUATION ============
FY25_REV = 455.72*M
FY = {
    # The whole valuation hinges on FY2027 gross margin. AAOI has run 28-31% for six
    # quarters while revenue grew 51%; management has targeted "high 30s" for years.
    "Bear": dict(q3=215*M, q4=245*M, rev27=1050*M, gm27=0.300, opex27=230*M,
                 pe=20.0, sh=85e6, note="800G ramps but margin never breaks the band"),
    "Base": dict(q3=234*M, q4=274*M, rev27=1400*M, gm27=0.330, opex27=250*M,
                 pe=32.0, sh=85e6, note="mix shift lifts GM to 33%; real operating leverage"),
    "Bull": dict(q3=250*M, q4=300*M, rev27=1750*M, gm27=0.370, opex27=270*M,
                 pe=42.0, sh=86e6, note="800G/1.6T at scale, US-supply-chain share gains"),
}
PROB = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}


def valuation(k):
    p = FY[k]
    rev26 = Q1_REV + Q2[k]["revenue"] + p["q3"] + p["q4"]
    op27 = p["rev27"] * p["gm27"] - p["opex27"]
    net27 = op27 - 25*M                      # interest on the convertible structure
    eps27 = net27 / p["sh"]
    target = eps27 * p["pe"]
    return dict(rev26=rev26, growth26=rev26/FY25_REV-1, gp27=p["rev27"]*p["gm27"],
                op27=op27, op_margin27=op27/p["rev27"], net27=net27, eps27=eps27,
                target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("AAOI — Applied Optoelectronics model", ""),
    ("Built", "2026-08-06, price $131.14"),
    ("DATA CAVEAT", "Built from filings/aggregators, NOT the user's moomoo screenshots. Income statement and guidance are exact; the balance-sheet, cash-flow and TTM-ratio sheets the other workbooks carry are absent. Send the moomoo Financials tabs to fill them in."),
    ("TIMING", "Q2 2026 reports TODAY, 2026-08-06, after the close. Call 4:30pm ET."),
    ("Latest REPORTED quarter", "Q1 2026: revenue $151.14M (+51.4% y/y), gross margin 29.06%, OPERATING LOSS -$12.99M, EPS -$0.19"),
    ("THE CENTRAL FACT", "Revenue went from $40.67M (Q1 2024) to $151.14M (Q1 2026) — 3.7x — and the company posted an OPERATING LOSS in all ten of those quarters. Gross margin has sat in a 28-31% band for six straight quarters while revenue grew 51%. Scale is not producing leverage."),
    ("What the bull case actually is", "A bet that the 800G/1.6T mix shift breaks gross margin out of the 28-31% band. 800G was only $4.6M of Q1's $151M (3.0%). Everything in the valuation is downstream of that single variable."),
    ("Contrast worth holding in mind", "AXTI flipped to a $10.4M operating profit on $47.6M of revenue with a 45% gross margin. AAOI cannot reach breakeven on $151M with a 29% gross margin. Substrate bottleneck vs. transceiver assembly — different positions in the same AI-optics supply chain."),
    ("Units", "USD; statement sheet displays millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("IS_Quarterly")
ws.cell(row=1, column=1, value="Reported quarterly income statement (calendar quarters).").font = Font(italic=True, size=9, color="7F7F7F")
for j, lab in enumerate(["USD, millions"] + COLS, start=1):
    c = ws.cell(row=2, column=j, value=lab); c.fill = HDR if j == 1 else SUB; c.font = WHITE
    c.alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 40
for j in range(2, n+2):
    ws.column_dimensions[get_column_letter(j)].width = 13
ws.freeze_panes = ws.cell(row=3, column=2)
r = 3
for name, vals in IS:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.number_format = '0.00' if "EPS" in name else NUMFMT
        c.border = THIN
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
rev = dict(IS)["Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Income"]
for name, vals, fmt in [
        ("Revenue q/q %", [rev[i]/rev[i+1]-1 if i+1 < n else None for i in range(n)], PCT),
        ("Revenue y/y %", [rev[i]/rev[i+4]-1 if i+4 < n else None for i in range(n)], PCT),
        ("GROSS MARGIN % <- the whole thesis", [gp[i]/rev[i] for i in range(n)], PCT),
        ("Operating margin %", [op[i]/rev[i] for i in range(n)], PCT),
        ("Implied opex (GP - op income)", [gp[i]-op[i] for i in range(n)], NUMFMT),
        ("Opex as % of revenue", [(gp[i]-op[i])/rev[i] for i in range(n)], PCT),
        ("Opex q/q %", [(gp[i]-op[i])/(gp[i+1]-op[i+1])-1 if i+1 < n else None for i in range(n)], PCT)]:
    c0 = ws.cell(row=r, column=1, value=name)
    if "thesis" in name:
        c0.font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = fmt; c.border = THIN
        if "thesis" in name:
            c.fill = WARN
    r += 1

ws = wb.create_sheet("Guidance_Market")
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 60
r = 1
ws.cell(row=r, column=1, value="COMPANY GUIDANCE & STREET").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for a, b, c_ in GUIDE:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c_).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
ws.cell(row=r, column=1, value="MARKET DATA").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.00' if (val is not None and abs(val) < 10000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1

ws = wb.create_sheet("Forecast_Q2_2026")
ws.column_dimensions["A"].width = 40
for col in "BCDEF":
    ws.column_dimensions[col].width = 15
ws.cell(row=1, column=1, value=f"Q1 2026 base: revenue ${Q1_REV/M:.2f}M, gross margin {Q1_GM:.2%}, GAAP opex ${Q1_OPEX/M:.2f}M. EPS line below is non-GAAP (opex net of SBC).").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value="Guidance $180-198M / -$0.03 to +$0.03 EPS. Consensus $191.13M / +$0.03 sits at the TOP of the EPS guide.").font = Font(italic=True, size=9, color="C00000")
r = 4
ws.cell(row=r, column=1, value="Q2 2026 (reports today AMC)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
c = ws.cell(row=r, column=6, value="Guide (hi)"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, ST_REV, GUIDE_HI),
          ("Revenue q/q", "qoq", PCT, ST_REV/Q1_REV-1, GUIDE_HI/Q1_REV-1),
          ("Revenue y/y", "yoy", PCT, 0.857, None),
          ("vs guidance midpoint $189M", "vs_guide_mid", '+0.0%;-0.0%', None, None),
          ("Gross profit", "gross_profit", NUMFMT, None, None),
          ("GROSS MARGIN", "gross_margin", PCT, None, None),
          ("Operating expense (non-GAAP)", "opex", NUMFMT, None, None),
          ("Operating income (non-GAAP)", "op_income", NUMFMT, None, None),
          ("Operating margin", "op_margin", PCT, None, None),
          ("Net income (non-GAAP)", "net", NUMFMT, None, None),
          ("Non-GAAP EPS", "eps", '0.00', ST_EPS, 0.03)]
for label, key, fmt, street, guide in fields:
    c0 = ws.cell(row=r, column=1, value=label)
    if "GROSS MARGIN" in label:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k][key]); c.number_format = fmt; c.border = THIN
        if "GROSS MARGIN" in label:
            c.fill = WARN
            c.font = BOLD
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    if guide is not None:
        c = ws.cell(row=r, column=6, value=guide); c.number_format = fmt
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=SCEN[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 48
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="Every scenario differs almost entirely by ONE assumption: FY2027 gross margin. Opex and revenue move far less between cases.").font = Font(italic=True, size=9, color="C00000")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("FY2025 actual revenue", lambda k: FY25_REV, NUMFMT, False),
        ("Q3 2026E revenue", lambda k: FY[k]["q3"], NUMFMT, False),
        ("Q4 2026E revenue", lambda k: FY[k]["q4"], NUMFMT, False),
        ("FY2026E revenue", lambda k: VAL[k]["rev26"], NUMFMT, False),
        ("FY2026E growth", lambda k: VAL[k]["growth26"], PCT, False),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT, False),
        ("FY2027E GROSS MARGIN", lambda k: FY[k]["gm27"], PCT, True),
        ("FY2027E gross profit", lambda k: VAL[k]["gp27"], NUMFMT, False),
        ("FY2027E opex", lambda k: FY[k]["opex27"], NUMFMT, False),
        ("FY2027E operating income", lambda k: VAL[k]["op27"], NUMFMT, False),
        ("FY2027E operating margin", lambda k: VAL[k]["op_margin27"], PCT, False),
        ("FY2027E net income", lambda k: VAL[k]["net27"], NUMFMT, False),
        ("FY2027E EPS", lambda k: VAL[k]["eps27"], '$0.00', False),
        ("P/E applied", lambda k: FY[k]["pe"], '0.0"x"', False),
        ("12-month target", lambda k: VAL[k]["target"], '$#,##0.00', True),
        ("Upside vs $131.14", lambda k: VAL[k]["upside"], '+0.0%;-0.0%', True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = WARN if "MARGIN" in label else SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")
r += 2
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
mc = 10.52*B
for label, val, fmt in [
        ("Market cap / TTM revenue ($507M)", mc/(507*M), '0.0"x"'),
        ("Market cap / FY2026E revenue (base)", mc/VAL["Base"]["rev26"], '0.0"x"'),
        ("Market cap / FY2027E revenue (base)", mc/FY["Base"]["rev27"], '0.0"x"'),
        ("P/E on FY2027E base EPS", PX_NOW/VAL["Base"]["eps27"], '0.0"x"'),
        ("Street mean PT $150.30 implies FY27 P/E of", 150.30/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk high $233.67 implies FY27 P/E of", 233.67/VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: what FY2027 gross margin is worth (revenue $1.4B, opex $250M, 32x P/E)").font = BOLD
r += 1
ws.cell(row=r, column=1, value="FY27 gross margin").font = BOLD
for j, h in enumerate(["Gross profit", "Op income", "Net income", "EPS", "Implied price"], start=2):
    c = ws.cell(row=r, column=j, value=h); c.fill = SUB; c.font = WHITE
r += 1
for gm_ in (0.29, 0.31, 0.33, 0.35, 0.37, 0.40):
    gp_ = 1400*M*gm_
    op_ = gp_ - 250*M
    net_ = op_ - 25*M
    eps_ = net_/85e6
    px_ = eps_*32
    c = ws.cell(row=r, column=1, value=gm_); c.number_format = PCT; c.fill = SUB; c.font = WHITE
    for j, (v, f) in enumerate([(gp_, NUMFMT), (op_, NUMFMT), (net_, NUMFMT),
                                (eps_, '$0.00'), (px_, '$#,##0.00')], start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = f
        if j == 6 and abs(px_ - PX_NOW) < 20:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/AAOI_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print(f"\nQ1 2026 actual: rev ${Q1_REV/M:.2f}M (+51.4% y/y), GM {Q1_GM:.2%}, OPERATING LOSS -$12.99M")
print("Ten consecutive operating-loss quarters while revenue grew 3.7x:")
for i in range(n):
    print(f"   {COLS[i]}: rev ${rev[i]/M:7.2f}M  GM {gp[i]/rev[i]:6.2%}  op {op[i]/M:+8.2f}M")
print(f"\nQ2 guidance: ${GUIDE_LO/M:.0f}-{GUIDE_HI/M:.0f}M / -$0.03 to +$0.03 | consensus ${ST_REV/M:.2f}M / +${ST_EPS}")
print("\n== Q2 2026 forecast (reports today AMC) ==")
print(f"{'':32s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}")
for label, key, _, street, guide in fields:
    vals = "".join(f"{Q2[k][key]/M:>12,.1f}" if abs(Q2[k][key]) > 100 else f"{Q2[k][key]:>12.3f}"
                   for k in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.1f}" if abs(street) > 100 else f"{street:>12.3f}")
    print(f"{label:32s}{vals}{s}")
print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY26 ${v['rev26']/M:6.0f}M (+{v['growth26']:.0%}) | FY27 ${FY[k]['rev27']/M:6.0f}M GM {FY[k]['gm27']:.0%} "
          f"-> op ${v['op27']/M:6.0f}M EPS ${v['eps27']:5.2f} x{FY[k]['pe']:4.0f} -> ${v['target']:7.2f} ({v['upside']:+.0%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW})")
print(f"Street $150.30 implies {150.30/VAL['Base']['eps27']:.0f}x FY27E base EPS | today = {PX_NOW/VAL['Base']['eps27']:.0f}x")
print(f"Mkt cap/TTM rev {mc/(507*M):.1f}x | /FY26E {mc/VAL['Base']['rev26']:.1f}x | /FY27E {mc/FY['Base']['rev27']:.1f}x")
