#!/usr/bin/env python3
"""
APP (AppLovin Corporation) financial model — same framework as MU / NBIS / SHOP.

Data entry = 100% from the user's moomoo terminal screenshots (2026-08-05, 12:06 ET):
  1) Valuation > Analyst Ratings (consensus, target price, 14 institution rows)
  2) Financials > Key Indicators   (Quarterly TTM)
  3) Financials > Cash Flow        (Quarterly, All, USD)
  4) Financials > Income Statement (Quarterly, All, USD)
  5) Financials > Balance Sheet    (Quarterly, All, USD)
  + quote panel and Trade Overview money flow

Company guidance + street consensus from the web sit on their own sheet.

NEXT EVENT: Q2 2026 results, 2026-08-05 AFTER THE CLOSE (today).

LEGACY WARNING: AppLovin divested its Apps (mobile games) business, completing in 2025.
moomoo columns 2024/Q1 and older mix the Apps segment and restatements — 2023/Q4 even
shows revenue of -$488.06M, a restatement artifact. Those columns are entered faithfully
but are NOT comparable to today's pure advertising-platform business.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9
K = 1e3

COLS = ["2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2", "2025/Q1",
        "2024/Q4", "2024/Q3", "2024/Q2", "2024/Q1",
        "2023/Q4", "2023/Q3", "2023/Q2", "2023/Q1",
        "2022/Q4", "2022/Q3"]
LEGACY_FROM_IDX = 8          # 2024/Q1 and older = Apps-era / restated
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"too many values: {len(v)}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Total Revenue as Reported", row(1.84*B, 1.66*B, 1.41*B, 1.26*B, 1.16*B,
                                      999.49*M, 835.19*M, 711.02*M, 1.06*B,
                                      -488.06*M, 864.26*M, 750.17*M, 715.41*M, 702.31*M, 713.10*M)),
    ("Total Operating Revenue", row(1.84*B, 1.66*B, 1.41*B, 1.26*B, 1.16*B,
                                    999.49*M, 835.19*M, 711.02*M, 1.06*B,
                                    -488.06*M, 864.26*M, 750.17*M, 715.41*M, 702.31*M, 713.10*M)),
    ("Cost of Revenue", row(203.63*M, 183.53*M, 174.86*M, 155.08*M, 151.68*M,
                            153.39*M, 120.92*M, 121.76*M, 294.15*M,
                            -428.97*M, 265.05*M, 258.58*M, 261.96*M, 369.37*M, 300.99*M)),
    ("Gross Profit", row(1.64*B, 1.47*B, 1.23*B, 1.10*B, 1.01*B,
                         846.09*M, 714.27*M, 589.26*M, 763.97*M,
                         -59.09*M, 599.21*M, 491.59*M, 453.45*M, 332.94*M, 412.11*M)),
    ("Operating Expense", row(198.88*M, 199.17*M, 151.18*M, 146.00*M, 167.31*M,
                              217.76*M, 179.97*M, 204.83*M, 424.41*M,
                              -452.81*M, 412.89*M, 360.26*M, 392.40*M, 355.37*M, 362.84*M)),
    ("  Selling and Admin Expenses", row(104.78*M, 116.95*M, 107.33*M, 101.96*M, 110.91*M,
                                         111.90*M, 99.19*M, 105.71*M, 269.09*M,
                                         -345.03*M, 253.60*M, 222.84*M, 247.55*M, 237.18*M, 240.79*M)),
    ("    – Selling & Marketing Expense", row(60.75*M, 48.78*M, 48.58*M, 46.92*M, 59.38*M,
                                              62.00*M, 62.98*M, 66.97*M, 226.69*M,
                                              -379.73*M, 212.35*M, 192.43*M, 202.98*M, 200.54*M, 196.79*M)),
    ("    – General & Admin Expense", row(44.03*M, 68.18*M, 58.76*M, 55.05*M, 51.52*M,
                                          49.89*M, 36.21*M, 38.75*M, 42.40*M,
                                          34.70*M, 41.25*M, 30.41*M, 44.57*M, 36.64*M, 44.00*M)),
    ("  Research & Development", row(94.10*M, 82.22*M, 43.85*M, 44.03*M, 56.41*M,
                                     105.86*M, 80.78*M, 99.12*M, 155.32*M,
                                     -107.78*M, 159.29*M, 137.42*M, 144.85*M, 118.19*M, 122.06*M)),
    ("Operating Profit", row(1.44*B, 1.28*B, 1.08*B, 957.68*M, 839.98*M,
                             628.33*M, 534.30*M, 384.42*M, 339.56*M,
                             393.72*M, 186.32*M, 131.33*M, 61.05*M, -22.43*M, 49.27*M)),
    ("Net Non-Operating Interest Income (Expense)", row(-51.16*M, -51.29*M, -51.43*M, -51.41*M, -52.89*M,
                                                        -93.93*M, -74.94*M, -74.42*M, -74.18*M,
                                                        -96.49*M, -77.09*M, -50.99*M, -74.51*M, -58.22*M, -47.66*M)),
    ("  Non-Operating Interest Income", row(None, None, None, None, None,
                                            None, None, None, 7.95*M,
                                            None, 1.49*M, None, None, None, 969.00*K)),
    ("  Non-Operating Interest Expense", row(51.16*M, 51.29*M, 51.43*M, 51.41*M, 52.89*M,
                                             93.93*M, 74.94*M, 74.42*M, 74.18*M,
                                             69.43*M, 78.58*M, 50.99*M, 74.51*M, 54.72*M, 48.63*M)),
    ("Other Income (Expense)", row(42.63*M, 29.40*M, -6.63*M, -22.27*M, 7.51*M,
                                   323.00*K, 8.37*M, 7.87*M, 2.57*M,
                                   None, None, 15.46*M, 10.11*M)),
    ("  Special Income (Charges)", row(None, None, None, None, -188.94*M)),
    ("    – Less: Impairment of Capital Assets", row(None, None, None, None, 188.94*M)),
    ("  Other Non-Operating Income (Expenses)", row(42.63*M, 29.40*M, -6.63*M, -22.27*M, 7.51*M,
                                                    323.00*K, 8.37*M, 7.87*M, 2.57*M,
                                                    None, None, 15.46*M, 10.11*M)),
    ("Pretax Profit", row(1.43*B, 1.25*B, 1.02*B, 884.00*M, 794.61*M,
                          534.73*M, 467.73*M, 317.88*M, 267.95*M,
                          299.93*M, 109.23*M, 95.80*M, -3.35*M, -66.17*M, 1.61*M)),
]

# ========================================================= BALANCE SHEET =====
BS = [
    ("Total Assets", row(7.71*B, 7.26*B, 6.34*B, 5.96*B, 5.71*B,
                         5.87*B, 5.44*B, 5.27*B, 5.26*B,
                         5.36*B, 5.01*B, 5.48*B, 5.92*B, 5.85*B, 5.81*B)),
    ("Total Current Assets", row(4.85*B, 4.43*B, 3.49*B, 2.99*B, 2.37*B,
                                 2.31*B, 1.88*B, 1.66*B, 1.61*B,
                                 1.62*B, 1.30*B, 1.70*B, 2.05*B, 1.94*B, 1.82*B)),
    ("  Cash & Equivalents + Short-Term Investments", row(2.76*B, 2.49*B, 1.67*B, 1.19*B, 551.02*M,
                                                          697.03*M, 567.60*M, 460.45*M, 436.34*M,
                                                          502.15*M, 332.49*M, 876.23*M, 1.25*B, 1.08*B, 943.51*M)),
    ("    – Cash and Cash Equivalents", row(2.76*B, 2.49*B, 1.67*B, 1.19*B, 551.02*M,
                                            697.03*M, 567.60*M, 460.45*M, 436.34*M,
                                            502.15*M, 332.49*M, 876.23*M, 1.25*B, 1.08*B, 943.51*M)),
    ("  Receivables / Accounts Receivable", row(1.96*B, 1.82*B, 1.60*B, 1.58*B, 1.58*B,
                                                1.28*B, 1.19*B, 1.07*B, 1.04*B,
                                                953.81*M, 849.14*M, 669.79*M, 637.61*M, 702.81*M, 665.46*M)),
    ("  Restricted Cash", row(*[None]*14, 0.0)),
    ("  Other Current Assets", row(130.88*M, 124.33*M, 216.71*M, 218.40*M, 238.50*M,
                                   140.47*M, 120.39*M, 125.48*M, 136.18*M,
                                   160.20*M, 119.16*M, 150.26*M, 164.86*M, 155.79*M, 215.28*M)),
    ("Total Non-Current Assets", row(2.86*B, 2.83*B, 2.86*B, 2.97*B, 3.34*B,
                                     3.56*B, 3.57*B, 3.61*B, 3.65*B,
                                     3.74*B, 3.70*B, 3.79*B, 3.87*B, 3.91*B, 3.98*B)),
    ("  Net PPE", row(114.82*M, 147.90*M, 130.82*M, 129.60*M, 161.66*M,
                      196.44*M, 177.05*M, 169.21*M, 172.99*M,
                      221.54*M, 155.15*M, 172.78*M, 178.55*M, 138.92*M, 130.82*M)),
    ("    – Gross PPE", row(114.82*M, 265.47*M, 130.82*M, 129.60*M, 161.66*M,
                            287.09*M, 177.05*M, 169.21*M, 172.99*M,
                            293.55*M, 155.15*M, 172.78*M, 178.55*M, 187.78*M, 130.82*M)),
    ("    – Accumulated Depreciation", row(None, -117.57*M, None, None, None,
                                           -90.65*M, None, None, None,
                                           -72.01*M, None, None, None, -48.86*M)),
    ("  Investments and Advances", row(288.67*M, 287.67*M, None, None, 0.0)),
    ("    – Long Term Equity Investment", row(288.67*M, 287.67*M, None, None, 0.0)),
    ("  Goodwill and Other Intangible Assets", row(1.89*B, 1.94*B, 1.96*B, 1.99*B, 2.49*B,
                                                   1.93*B, 2.88*B, 2.92*B, 3.03*B,
                                                   3.14*B, 3.20*B, 3.32*B, 3.41*B, 3.50*B, 3.65*B)),
    ("    – Goodwill", row(1.52*B, 1.54*B, 1.54*B, 1.54*B, 1.64*B,
                           1.46*B, 1.85*B, 1.82*B, 1.83*B,
                           1.84*B, 1.81*B, 1.83*B, 1.83*B, 1.83*B, 1.76*B)),
    ("    – Other Intangible Assets", row(369.00*M, 396.71*M, 421.87*M, 448.18*M, 855.05*M,
                                          472.85*M, 1.02*B, 1.10*B, 1.20*B,
                                          1.29*B, 1.39*B, 1.49*B, 1.57*B, 1.68*B, 1.89*B)),
    ("  Other Non-Current Assets", row(564.60*M, 456.55*M, 761.90*M, 849.73*M, 682.87*M,
                                       1.43*B, 514.07*M, 516.14*M, 456.32*M,
                                       386.00*M, 349.12*M, 288.37*M, 280.68*M, 268.43*M, 198.00*M)),
    ("Total Liabilities", row(5.34*B, 5.12*B, 4.87*B, 4.79*B, 5.13*B,
                              4.78*B, 4.50*B, 4.45*B, 4.50*B,
                              4.10*B, 3.91*B, 3.96*B, 4.02*B, 3.95*B, 3.92*B)),
    ("Total Current Liabilities", row(1.49*B, 1.33*B, 1.07*B, 1.09*B, 1.41*B,
                                      1.06*B, 779.54*M, 729.59*M, 762.52*M,
                                      944.12*M, 805.47*M, 591.94*M, 619.57*M, 578.96*M, 543.63*M)),
]

# ============================================================= CASH FLOW =====
CF = [
    ("Operating Cash Flow", row(1.29*B, 1.31*B, 1.05*B, 772.23*M, 831.71*M,
                                701.00*M, 550.70*M, 454.53*M, 392.78*M,
                                343.99*M, 199.07*M, 229.79*M, 288.66*M, 163.20*M, 174.48*M)),
    ("  Net cash flow from continuing operations", row(1.29*B, 1.31*B, 1.05*B, 772.23*M, 831.71*M,
                                                       701.00*M, 550.70*M, 454.53*M, 392.78*M,
                                                       343.99*M, 199.07*M, 229.79*M, 288.66*M, 163.20*M, 174.48*M)),
    ("    – Net Income from Continuing Operations", row(1.21*B, 1.10*B, 835.55*M, 819.53*M, 576.42*M,
                                                        599.20*M, 434.42*M, 309.97*M, 236.18*M,
                                                        172.23*M, 108.64*M, 80.36*M, -4.52*M, -79.51*M, 23.66*M)),
    ("    – Depreciation & Depletion & Amortization", row(33.67*M, 32.74*M, 35.10*M, 47.05*M, 79.89*M,
                                                          127.84*M, 99.64*M, 108.54*M, 112.67*M,
                                                          119.11*M, 121.80*M, 119.89*M, 128.21*M, 101.58*M, 163.83*M)),
    ("    – Other Non-Cash Items", row(-16.48*M, -189.55*M, 72.41*M, 130.56*M, 8.09*M,
                                       -270.41*M, 88.49*M, 195.74*M, 8.54*M,
                                       7.24*M, 9.37*M, 3.32*M, 9.14*M, 6.03*M, 7.69*M)),
    ("    – Change in Working Capital", row(-14.78*M, 107.87*M, 110.36*M, -57.41*M, -82.90*M,
                                            -153.37*M, -71.84*M, -64.47*M, -59.86*M,
                                            -74.94*M, -151.58*M, -55.03*M, 72.87*M, -40.46*M, -62.84*M)),
    ("        Change in Receivables", row(-138.10*M, -215.47*M, -35.20*M, -124.17*M, -167.38*M,
                                          -229.50*M, -112.35*M, -40.35*M, -84.84*M,
                                          -114.48*M, -180.07*M, -32.34*M, 65.61*M, -35.48*M, 24.07*M)),
    ("        Change in Prepaid Assets", row(-9.83*M, 50.18*M, 63.78*M, 51.18*M, -51.86*M,
                                             -218.53*M, 7.04*M, -652.00*K, 26.81*M,
                                             -104.48*M, -24.36*M, 12.42*M, -17.55*M, 66.52*M, 41.25*M)),
    ("        Change in Payables and Accrued Expense", row(133.15*M, 288.38*M, 81.78*M, -5.79*M, 136.34*M,
                                                           308.77*M, 33.47*M, -23.47*M, -1.84*M,
                                                           162.64*M, 52.85*M, -35.11*M, 24.80*M, 7.59*M, -32.09*M)),
    ("        Change in Other Current Assets", row(*[None]*13, -72.73*M, -7.65*M)),
    ("        Change in Other Current Liabilities", row(*[None]*13, -3.55*M, -3.30*M)),
    ("        Change In Other Working Capital", row(*[None]*13, -2.81*M, -2.62*M)),
    ("Net cash flow from investing", row(-5.25*M, -828.00*K, -19.63*M, 401.55*M, -22.66*M,
                                         -367.00*K, -6.40*M, -68.36*M, -31.64*M,
                                         -6.80*M, -15.83*M, -42.22*M, -12.98*M, 22.29*M, -42.01*M)),
    ("  Net Cash Flow from Continuing Investing", row(-5.25*M, -828.00*K, -19.63*M, 401.55*M, -22.66*M,
                                                      -367.00*K, -6.40*M, -68.36*M, -31.64*M,
                                                      -6.80*M, -15.83*M, -42.22*M, -12.98*M, 22.29*M, -42.01*M)),
    ("    – Capital Expenditure", row(*[None]*14, -1.94*M)),
    ("    – Net PPE Purchase and Sale", row(*[None]*13, -41.00*K, -221.00*K)),
    ("    – Net Intangibles Purchase and Sale", row(None, None, None, None, None,
                                                    None, -3.20*M, None, -12.08*M, -13.46*M)),
    ("    – Net Business Purchase and Sale", row(None, 0.0, -17.41*M, None, 0.0,
                                                 0.0, None, None, None,
                                                 None, None, None, None, -10.08*M, -41.35*M)),
    ("    – Net Investment Purchase and Sale", row(0.0, 0.0, -1.50*M, 0.0, -18.68*M,
                                                   0.0, -650.00*K, -48.00*M, -28.33*M,
                                                   -1.00*M, -100.00*K, 0.0, -16.83*M, -9.80*M, 0.0)),
]

# ======================================================= KEY INDICATORS ======
KI = [
    ("— Profitability (TTM) —", row()),
    ("Gross Margin %", row(88.37, 87.86, 79.69, 78.61, 77.72, 75.22, 73.89, 71.80, 69.90,
                           67.74, 61.91, 58.66, 57.47, 55.41, 60.39)),
    ("Operating Margin %", row(77.09, 75.75, 54.97, 50.94, 46.47, 39.78, 35.78, 29.99, 25.57,
                               19.74, 11.75, 7.61, 4.86, -1.70, 1.14)),
    ("EBIT Margin %", row(77.79, 75.90, 54.68, 50.88, 43.30, 40.22, 35.79, 29.84, 25.57,
                          19.99, 13.00, 8.91, 5.64, -1.18, 1.28)),
    ("Net Margin %", row(64.28, 60.82, 44.86, 42.32, 37.36, 33.49, 26.80, 20.83, 16.39,
                         10.81, 3.43, 0.69, -2.82, -6.84, -2.92)),
    ("EBITDA Margin %", row(80.20, 79.46, 59.28, 57.06, 51.40, 49.75, 46.05, 41.52, 38.63,
                            34.88, 28.55, 24.43, 24.43, 18.24, 20.57)),
    ("Tax Rate %", row(14.69, 13.15, 8.79, 4.93, 0.33, None, 7.02, 6.16, 8.35,
                       None, 6.27, 22.54, 28.32)),
    ("Interest Coverage (x)", row(23.36, 20.09, 13.76, 10.65, 7.49, 5.95, 5.19, 3.95, 3.37,
                                  2.38, 1.52, 1.12, 0.76, -0.19, 0.25)),
    ("R&D Expense Ratio %", row(4.29, 4.13, 8.15, 9.60, 11.81, 13.56, 14.45, 15.91, 16.63,
                                18.04, 18.46, 18.14, 18.10, 18.02, 17.50)),
    ("Sales Expense Ratio %", row(3.33, 3.72, 12.89, 14.42, 15.69, 18.03, 19.99, 21.84, 23.57,
                                  25.30, 26.66, 27.52, 28.63, 32.64, 35.51)),
    ("Administrative Expense Rate %", row(3.67, 4.26, 3.67, 3.64, 3.76, 3.85, 3.67, 4.06, 4.15,
                                          4.65, 5.04, 5.40, 5.88, 6.45, 6.24)),
    ("— Solvency —", row()),
    ("Long-Term Debt to Equity %", row(148.68, 165.40, 238.27, 300.82, 609.98, 324.83, 370.33, 427.35, 459.07,
                                       234.72, 270.05, 212.08, 169.94, 169.90, 172.08)),
    ("Total Assets to Common Equity %", row(326.12, 340.08, 430.35, 510.61, 991.74, 538.55, 580.09, 646.69, 692.25,
                                            426.58, 456.77, 361.28, 311.59, 307.35, 308.49)),
    ("Equity Ratio %", row(30.66, 29.40, 23.24, 19.58, 10.08, 18.57, 17.24, 15.46, 14.45,
                           23.44, 21.89, 27.68, 32.09, 32.54, 32.42)),
    ("Debt to Asset Ratio %", row(148.68, 166.06, 238.27, 300.82, 644.74, 326.17, 374.12, 431.71, 463.75,
                                  252.91, 290.93, 215.20, 172.44, 172.40, 174.57)),
    ("Current Ratio (x)", row(3.24, 3.32, 3.25, 2.74, 1.68, 2.19, 2.41, 2.28, 2.11,
                              1.71, 1.61, 2.87, 3.31, 3.35, 3.36)),
    ("Quick Ratio (x)", row(3.16, 3.23, 3.05, 2.54, 1.51, 1.87, 2.25, 2.10, 1.94,
                            1.54, 1.47, 2.61, 3.04, 3.08, 2.96)),
    ("— Operating Capacity (TTM) —", row()),
    ("Receivable Turnover (x)", row(3.49, 3.39, 4.52, 4.32, 3.93, 3.98, 4.21, 4.54, 4.33,
                                    3.96, 4.00, 4.23, 4.40, 4.63, 5.39)),
    ("Account Payable Turnover (x)", row(1.11, 1.02, 2.71, 2.61, 2.32, 2.50, 3.16, 3.44, 3.24,
                                         3.28, 4.26, 4.38, 3.79, 4.73, 4.90)),
]

# ================================================ ANALYST RATINGS (screenshot 1)
CONSENSUS_RATING = dict(rating="Strong Buy", analysts=21, updated="2026-08-03",
                        buy=95.24, hold=4.76, sell=0.00,
                        pt_high=798.00, pt_avg=668.84, pt_low=515.00, current=423.98)
RATINGS = [
    ("Citi", "Buy", 710, 710, "Maintained", "2026-08-03"),
    ("UBS", "Buy", 750, 798, "Upgrade", "2026-08-03"),
    ("BofA Securities", "Buy", 705, 705, "Maintained", "2026-07-30"),
    ("RBC Capital", "Buy", 700, 700, "Maintained", "2026-07-29"),
    ("Wedbush", "Buy", 640, 640, "Maintained", "2026-07-29"),
    ("Jefferies", "Buy", 700, 700, "Maintained", "2026-07-22"),
    ("Piper Sandler", "Buy", 665, 665, "Maintained", "2026-07-07"),
    ("Wells Fargo", "Buy", 571, 575, "Upgrade", "2026-07-07"),
    ("BTIG", "Buy", 640, 640, "Maintained", "2026-07-01"),
    ("Raymond James", "Buy", None, 640, "New", "2026-06-29"),
    ("Scotiabank", "Buy", 775, 775, "Maintained", "2026-06-26"),
    ("Morgan Stanley", "Buy", 720, 720, "Maintained", "2026-06-11"),
    ("Edgewater Research", "Buy", None, None, "New", "2026-06-05"),
    ("Needham", "Buy", 700, 700, "Maintained", "2026-05-28"),
]

# ========================================================== MARKET DATA ======
PX_NOW = 423.98
SHARES = 335.94e6
MKT = [
    ("Last price (2026-08-05, 12:06 ET)", 423.980, "USD, +1.02%"),
    ("Previous close", 419.700, "USD"),
    ("Open", 433.300, "USD"), ("Day High", 437.610, "USD"), ("Day Low", 422.330, "USD"),
    ("Day Range %", 3.64, "%"),
    ("Volume", 3.27e6, "shares"), ("Turnover", 1.41*B, "USD"),
    ("Average Price", 431.452, "USD"), ("Turnover Rate %", 1.28, "%"),
    ("Vol Ratio", 1.02, "x"), ("Bid/Ask %", -41.84, "% (ask-heavy)"),
    ("Market Cap", 142.43*B, "USD"), ("Float Market Cap", 108.42*B, "USD"),
    ("Total Shares", 335.94e6, "shares"), ("Free Float", 255.71e6, "shares"),
    ("P/E (TTM)", 36.42, "x"), ("P/E (LFY)", 43.49, "x"), ("P/B", 60.267, "x"),
    ("52-week High", 745.610, "USD"), ("52-week Low", 359.000, "USD"),
    ("Historical High", 745.610, "USD"), ("Historical Low", 9.140, "USD"),
    ("Dividend TTM", None, "none"),
    ("Next earnings", None, "2026-08-05 AFTER CLOSE (today)"),
    ("Money flow — NET OUTFLOW", -21.66*M, "inflow 108.02M / outflow 129.68M"),
    ("  Extra Large net", (13.91-21.75)*M, "in 13.91M / out 21.75M"),
    ("  Large net", (30.01-31.39)*M, "in 30.01M / out 31.39M"),
    ("  Medium net", (20.75-26.42)*M, "in 20.75M / out 26.42M"),
    ("  Small net", (43.36-50.11)*M, "in 43.36M / out 50.11M"),
]

# ========================================== GUIDANCE / STREET (non-screenshot) =
STREET = [
    ("Q2 2026 report", "2026-08-05, after the close", "company"),
    ("Q2 2026 revenue GUIDANCE", "$1.915B - $1.945B (+52% to +55% y/y)", "company, given at Q1"),
    ("Q2 2026 adj. EBITDA GUIDANCE", "$1.615B - $1.645B (84%-85% margin)", "company"),
    ("Q2 2026 consensus revenue", "$1.94B (+54% y/y) — sits at the TOP of guidance", "street"),
    ("Q2 2026 consensus EPS", "$3.72 - $3.75 (+64.6% y/y)", "street"),
    ("Q1 2026 actual", "revenue $1.842B (+59%), adj EBITDA $1.56B (record 85% margin), FCF $1.29B", "company"),
    ("Q1 2026 buyback", "2.23M shares for $1.0B; $2.3B remaining authorization", "company"),
    ("Beat streak", "beat EPS consensus in each of the last 4 quarters", "street"),
    ("Analyst mean PT (moomoo)", "$668.84 (21 analysts, Strong Buy, 95.24% Buy)", "screenshot"),
    ("Analyst PT range", "$515 low / $798 high", "screenshot"),
    ("Why the stock de-rated in 2026", "-24% in H1 2026: Meta expanding into untracked iOS ad traffic; Google 'Project Genie' AI game-creation seen as a threat to the mobile-game ad ecosystem; high-multiple AI ad-tech sold hardest on macro risk-off", "web"),
]

# ============================================ Q2-2026 FORECAST ENGINE ========
# Calibrations off Q1 2026 actuals:
#   adj EBITDA 1.56B = operating profit 1.44B + D&A 33.67M + SBC add-back X  ->  X = 86M
SBC_Q1 = 1.56*B - (1.44*B + 33.67*M)
Q1_26_REV = 1.842*B          # exact, from stockanalysis/company (moomoo rounds to 1.84B)
Q2_25_REV = 1.259*B
GUIDE_LO, GUIDE_HI = 1.915*B, 1.945*B
ST_REV, ST_EPS = 1.94*B, 3.72

# seasonal Q2/Q1 step: only 2025 is a clean comp (2024/Q1 and older include the Apps business)
QOQ_2025 = 1.259/1.159

SCEN = {
    #            revenue, gross margin, S&M, G&A, R&D, D&A, net interest, other inc, tax, SBC
    "Bear": dict(rev=1.940*B, gm=0.888, sm=63*M, ga=48*M, rnd=99*M, da=34*M,
                 nint=-51*M, other=20*M, tax=0.150, sbc=88*M, shares=335.0e6),
    "Base": dict(rev=1.985*B, gm=0.890, sm=62*M, ga=46*M, rnd=97*M, da=33*M,
                 nint=-51*M, other=30*M, tax=0.147, sbc=88*M, shares=334.0e6),
    "Bull": dict(rev=2.035*B, gm=0.893, sm=61*M, ga=45*M, rnd=95*M, da=33*M,
                 nint=-51*M, other=40*M, tax=0.145, sbc=90*M, shares=333.5e6),
}


def forecast_q2(p):
    gp = p["rev"] * p["gm"]
    opex = p["sm"] + p["ga"] + p["rnd"]
    op = gp - opex
    adj_ebitda = op + p["da"] + p["sbc"]
    pretax = op + p["nint"] + p["other"]
    net = pretax * (1 - p["tax"])
    return dict(revenue=p["rev"], yoy=p["rev"]/Q2_25_REV-1, qoq=p["rev"]/Q1_26_REV-1,
                gross_profit=gp, gross_margin=p["gm"], opex=opex,
                op_profit=op, op_margin=op/p["rev"],
                adj_ebitda=adj_ebitda, adj_ebitda_margin=adj_ebitda/p["rev"],
                pretax=pretax, net=net, eps=net/p["shares"], fcf=p["rev"]*0.70)


Q2 = {k: forecast_q2(v) for k, v in SCEN.items()}

# ================================= FY2026-2028 MODEL + VALUATION ============
FY25 = (1.159 + 1.259 + 1.405 + 1.658) * B     # 5.481B actual
NET_DEBT = 7.71*B*0.3066*1.4868 - 2.76*B       # LT debt (equity x LTD/E) less cash
EV_NOW = 142.43*B + NET_DEBT

FY = {
    "Bear": dict(rev26=8.20*B, rev27=10.2*B, opm27=0.760, ebitda_m=0.840,
                 pe=18.0, ev_ebitda=14.0, shares27=325e6),
    "Base": dict(rev26=8.45*B, rev27=11.4*B, opm27=0.785, ebitda_m=0.850,
                 pe=26.0, ev_ebitda=20.0, shares27=320e6),
    "Bull": dict(rev26=8.70*B, rev27=13.0*B, opm27=0.800, ebitda_m=0.860,
                 pe=33.0, ev_ebitda=25.0, shares27=316e6),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}


def valuation(p):
    op27 = p["rev27"] * p["opm27"]
    ebitda27 = p["rev27"] * p["ebitda_m"]
    pretax27 = op27 - 0.20*B
    eps27 = pretax27 * (1 - 0.16) / p["shares27"]
    px_pe = eps27 * p["pe"]
    px_ebitda = (ebitda27 * p["ev_ebitda"] - NET_DEBT) / p["shares27"]
    target = (px_pe + px_ebitda) / 2
    return dict(op27=op27, ebitda27=ebitda27, eps27=eps27, px_pe=px_pe,
                px_ebitda=px_ebitda, target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(v) for k, v in FY.items()}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
LEG = PatternFill("solid", fgColor="4A4A4A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'


def style_header(ws, r, labels, width0=46, width=13):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR if j == 1 else (LEG if j-2 >= LEGACY_FROM_IDX else SUB)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = width0
    for j in range(2, len(labels)+1):
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=r+1, column=2)


def write_stmt(wb, title, data, note):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=9, color="7F7F7F")
    style_header(ws, 2, ["USD, millions"] + COLS)
    r = 3
    for name, vals in data:
        c = ws.cell(row=r, column=1, value=name)
        if not name.startswith(" "):
            c.font = BOLD
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = NUMFMT
            cell.border = THIN
        r += 1
    return ws


wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("APP — AppLovin Corporation model", ""),
    ("Built", "2026-08-05, 12:06 ET, price $423.98"),
    ("Primary data source", "moomoo terminal screenshots supplied by the user: Analyst Ratings / Key Indicators / Cash Flow / Income Statement / Balance Sheet / quote panel + money flow"),
    ("Secondary layer", "company guidance and street consensus, isolated on 'Guidance_Street'"),
    ("Latest REPORTED quarter", "2026/Q1: revenue $1.842B (+59% y/y), operating profit $1.44B (78.2% margin), adj EBITDA $1.56B (85%), FCF $1.29B"),
    ("NEXT REPORT", "2026-08-05 AFTER THE CLOSE — the quarter 'Forecast_Q2_2026' projects"),
    ("LEGACY WARNING", "AppLovin divested its Apps (mobile games) segment, completing in 2025. Columns 2024/Q1 and older (grey headers) mix the Apps business and restatements — 2023/Q4 shows revenue of -$488.06M, a restatement artifact. Not comparable to the pure ad-platform business."),
    ("Where the margin step change comes from", "TTM gross margin jumps 79.69% -> 87.86% and operating margin 54.97% -> 75.75% between 2025/Q3 and 2025/Q4. That is the divestiture washing out of the TTM window, not an operational event."),
    ("Units", "USD; number format displays millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

write_stmt(wb, "IS_Quarterly", IS, "moomoo > Financials > Income Statement > Quarterly · All · USD.")
write_stmt(wb, "BS_Quarterly", BS, "moomoo > Financials > Balance Sheet > Quarterly · All · USD.")
write_stmt(wb, "CF_Quarterly", CF, "moomoo > Financials > Cash Flow > Quarterly · All · USD.")

ws = wb.create_sheet("Key_Indicators_TTM")
ws.cell(row=1, column=1, value="moomoo > Financials > Key Indicators. TTM basis.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["TTM indicator"] + COLS)
r = 3
for name, vals in KI:
    c = ws.cell(row=r, column=1, value=name)
    if name.startswith("—"):
        c.font = BOLD
        for j in range(1, n+2):
            ws.cell(row=r, column=j).fill = SEC
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = '0.00'
        cell.border = THIN
    r += 1

ws = wb.create_sheet("Derived_Quarterly")
ws.cell(row=1, column=1, value="Computed from the entered statements. Ignore columns 2024/Q1 and older (Apps-era).").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["Derived metric"] + COLS)
rev = dict(IS)["Total Revenue as Reported"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Profit"]
opx = dict(IS)["Operating Expense"]
ocf = dict(CF)["Operating Cash Flow"]
da = dict(CF)["    – Depreciation & Depletion & Amortization"]
ni = dict(CF)["    – Net Income from Continuing Operations"]
ar = dict(BS)["  Receivables / Accounts Receivable"]
cash = dict(BS)["    – Cash and Cash Equivalents"]
ta = dict(BS)["Total Assets"]
tl = dict(BS)["Total Liabilities"]


def safe(f, i):
    try:
        return f(i)
    except (TypeError, ZeroDivisionError):
        return None


drv = [
    ("Revenue y/y %", [safe(lambda i: rev[i]/rev[i+4]-1, i) if i+4 < n else None for i in range(n)]),
    ("Revenue q/q %", [safe(lambda i: rev[i]/rev[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Gross margin % (quarterly)", [safe(lambda i: gp[i]/rev[i], i) for i in range(n)]),
    ("Operating margin % (quarterly)", [safe(lambda i: op[i]/rev[i], i) for i in range(n)]),
    ("Opex as % of revenue", [safe(lambda i: opx[i]/rev[i], i) for i in range(n)]),
    ("Adj. EBITDA (op + D&A + SBC est.)", [safe(lambda i: op[i]+da[i]+86*M, i) for i in range(n)]),
    ("Adj. EBITDA margin %", [safe(lambda i: (op[i]+da[i]+86*M)/rev[i], i) for i in range(n)]),
    ("Operating cash flow margin %", [safe(lambda i: ocf[i]/rev[i], i) for i in range(n)]),
    ("Net income margin %", [safe(lambda i: ni[i]/rev[i], i) for i in range(n)]),
    ("Accounts receivable", ar),
    ("AR / quarterly revenue (x)", [safe(lambda i: ar[i]/rev[i], i) for i in range(n)]),
    ("Cash", cash),
    ("Shareholders' equity (TA - TL)", [safe(lambda i: ta[i]-tl[i], i) for i in range(n)]),
]
r = 3
for name, vals in drv:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if "%" in name else ('0.00"x"' if "(x)" in name else NUMFMT)
        cell.border = THIN
    r += 1

ws = wb.create_sheet("Analyst_Ratings")
ws.column_dimensions["A"].width = 28
for col, w in zip("BCDEF", (12, 12, 12, 18, 14)):
    ws.column_dimensions[col].width = w
ws.cell(row=1, column=1, value="moomoo > Valuation > Analyst Ratings (updated 2026-08-03).").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for k_, v_ in [("Consensus rating", CONSENSUS_RATING["rating"]),
               ("Analysts covering", CONSENSUS_RATING["analysts"]),
               ("Buy / Hold / Sell %", f"{CONSENSUS_RATING['buy']} / {CONSENSUS_RATING['hold']} / {CONSENSUS_RATING['sell']}"),
               ("Target price — High", CONSENSUS_RATING["pt_high"]),
               ("Target price — Average", CONSENSUS_RATING["pt_avg"]),
               ("Target price — Low", CONSENSUS_RATING["pt_low"]),
               ("Current price", CONSENSUS_RATING["current"]),
               ("Implied upside to mean PT", CONSENSUS_RATING["pt_avg"]/CONSENSUS_RATING["current"]-1),
               ("Price vs 52wk high $745.61", CONSENSUS_RATING["current"]/745.61-1)]:
    ws.cell(row=r, column=1, value=k_).font = BOLD
    c = ws.cell(row=r, column=2, value=v_)
    if "upside" in k_ or "vs 52wk" in k_:
        c.number_format = '+0.0%;-0.0%'
    r += 1
r += 1
for j, h in enumerate(["Institution", "Rating", "Prior PT", "New PT", "Action", "Date"], start=1):
    c = ws.cell(row=r, column=j, value=h); c.fill = HDR; c.font = WHITE
r += 1
for name, rating, old, new, action, date in RATINGS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=rating)
    ws.cell(row=r, column=3, value=old)
    ws.cell(row=r, column=4, value=new)
    c = ws.cell(row=r, column=5, value=action)
    if old and new and new > old:
        c.font = Font(color="008000")
    ws.cell(row=r, column=6, value=date)
    r += 1

ws = wb.create_sheet("Market_Data")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 44
ws.cell(row=1, column=1, value="moomoo quote panel + Trade Overview, 2026-08-05 12:06 ET.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.000' if (val is not None and abs(val) < 1000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
ttm_rev = (1.842+1.658+1.405+1.259)*B
ttm_ebitda = (1440+1280+1080+957.68)*M + (33.67+32.74+35.10+47.05)*M + 4*86*M
ttm_ocf = (1.29+1.31+1.05+0.77223)*B
for name, val, fmt in [
        ("Book value = Mkt cap / P/B", 142.43*B/60.267, NUMFMT),
        ("Check: Total assets - Total liabilities", 7.71*B-5.34*B, NUMFMT),
        ("Implied LT debt = Equity x LTD/E 148.68%", 7.71*B*0.3066*1.4868, NUMFMT),
        ("Net debt (LT debt - cash 2.76B)", NET_DEBT, NUMFMT),
        ("Enterprise value", EV_NOW, NUMFMT),
        ("TTM revenue", ttm_rev, NUMFMT),
        ("TTM adj. EBITDA (est.)", ttm_ebitda, NUMFMT),
        ("TTM operating cash flow", ttm_ocf, NUMFMT),
        ("EV / TTM revenue", EV_NOW/ttm_rev, '0.0"x"'),
        ("EV / TTM adj. EBITDA", EV_NOW/ttm_ebitda, '0.0"x"'),
        ("EV / TTM operating cash flow", EV_NOW/ttm_ocf, '0.0"x"'),
        ("P/E TTM (moomoo)", 36.42, '0.0"x"')]:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

ws = wb.create_sheet("Guidance_Street")
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 78
ws.column_dimensions["C"].width = 24
ws.cell(row=1, column=1, value="NOT from the statement screenshots — company guidance and street consensus, web, 2026-08-05.").font = Font(italic=True, size=9, color="C00000")
r = 3
for a, b, c_ in STREET:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    r += 1

ws = wb.create_sheet("Forecast_Q2_2026")
ws.column_dimensions["A"].width = 42
for col in "BCDEF":
    ws.column_dimensions[col].width = 15
ws.cell(row=1, column=1, value=f"Calibration off Q1 2026: adj EBITDA $1.56B = op profit $1.44B + D&A $33.67M + SBC ${SBC_Q1/M:.0f}M").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value=f"Seasonal Q2/Q1 step (2025, only clean comp): {QOQ_2025-1:+.1%} -> ${Q1_26_REV*QOQ_2025/B:.3f}B. Guidance ${GUIDE_LO/B:.3f}-{GUIDE_HI/B:.3f}B implies only {GUIDE_HI/Q1_26_REV-1:+.1%} q/q.").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="Q2 2026 (reports today after close)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k_); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
c = ws.cell(row=r, column=6, value="Guide (hi)"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, ST_REV, GUIDE_HI),
          ("Revenue y/y", "yoy", PCT, 0.541, GUIDE_HI/Q2_25_REV-1),
          ("Revenue q/q", "qoq", PCT, None, None),
          ("Gross profit", "gross_profit", NUMFMT, None, None),
          ("Gross margin", "gross_margin", PCT, None, None),
          ("Operating expense", "opex", NUMFMT, None, None),
          ("Operating profit", "op_profit", NUMFMT, None, None),
          ("Operating margin", "op_margin", PCT, None, None),
          ("Adj. EBITDA", "adj_ebitda", NUMFMT, None, 1.645*B),
          ("Adj. EBITDA margin", "adj_ebitda_margin", PCT, None, 0.85),
          ("Pre-tax profit", "pretax", NUMFMT, None, None),
          ("Net income", "net", NUMFMT, None, None),
          ("EPS (diluted)", "eps", '0.00', ST_EPS, None),
          ("Free cash flow (~70% of revenue)", "fcf", NUMFMT, None, None)]
for label, key, fmt, street, guide in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k_][key]); c.number_format = fmt; c.border = THIN
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    if guide is not None:
        c = ws.cell(row=r, column=6, value=guide); c.number_format = fmt
    r += 1
r += 1
for label, key, base in [("Beat/miss vs street revenue", "revenue", ST_REV),
                         ("Beat/miss vs guidance high", "revenue", GUIDE_HI),
                         ("Beat/miss vs street EPS", "eps", ST_EPS),
                         ("Adj. EBITDA vs guidance high", "adj_ebitda", 1.645*B)]:
    ws.cell(row=r, column=1, value=label).font = BOLD
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k_][key]/base-1); c.number_format = '+0.0%;-0.0%'
    r += 1

ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 44
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="FY2026E built from the quarterly model; FY2027E extrapolated. Target = average of P/E and EV/EBITDA routes on FY2027E.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k_); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt in [
        ("FY2025 actual revenue", lambda k_: FY25, NUMFMT),
        ("FY2026E revenue", lambda k_: FY[k_]["rev26"], NUMFMT),
        ("FY2026E growth", lambda k_: FY[k_]["rev26"]/FY25-1, PCT),
        ("FY2027E revenue", lambda k_: FY[k_]["rev27"], NUMFMT),
        ("FY2027E growth", lambda k_: FY[k_]["rev27"]/FY[k_]["rev26"]-1, PCT),
        ("FY2027E operating margin", lambda k_: FY[k_]["opm27"], PCT),
        ("FY2027E operating profit", lambda k_: VAL[k_]["op27"], NUMFMT),
        ("FY2027E adj. EBITDA", lambda k_: VAL[k_]["ebitda27"], NUMFMT),
        ("FY2027E EPS", lambda k_: VAL[k_]["eps27"], '$0.00'),
        ("P/E multiple applied", lambda k_: FY[k_]["pe"], '0.0"x"'),
        ("EV/EBITDA multiple applied", lambda k_: FY[k_]["ev_ebitda"], '0.0"x"'),
        ("Route 1 — P/E price", lambda k_: VAL[k_]["px_pe"], '$#,##0.00'),
        ("Route 2 — EV/EBITDA price", lambda k_: VAL[k_]["px_ebitda"], '$#,##0.00'),
        ("12-month target (avg)", lambda k_: VAL[k_]["target"], '$#,##0.00'),
        ("Upside vs $423.98", lambda k_: VAL[k_]["upside"], '+0.0%;-0.0%')]:
    c0 = ws.cell(row=r, column=1, value=label)
    hi = "target" in label or "Upside" in label
    if hi:
        c0.font = BOLD
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k_)); c.number_format = fmt; c.border = THIN
        if hi:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k_]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
fy26_eps = 3.56 + Q2["Base"]["eps"] + 4.30 + 5.05
for label, val, fmt in [("EV / FY2026E revenue (base)", EV_NOW/FY["Base"]["rev26"], '0.0"x"'),
                        ("EV / FY2027E revenue (base)", EV_NOW/FY["Base"]["rev27"], '0.0"x"'),
                        ("FY2026E EPS (Q1 actual + model)", fy26_eps, '$0.00'),
                        ("Forward P/E on FY2026E", PX_NOW/fy26_eps, '0.0"x"'),
                        ("Forward P/E on FY2027E (base)", PX_NOW/VAL["Base"]["eps27"], '0.0"x"'),
                        ("Street mean PT $668.84 implies FY27 P/E of", 668.84/VAL["Base"]["eps27"], '0.0"x"'),
                        ("52wk high $745.61 implies FY27 P/E of", 745.61/VAL["Base"]["eps27"], '0.0"x"'),
                        ("52wk low $359.00 implies FY27 P/E of", 359.00/VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

out = "/Users/antaiwei/Desktop/stock/APP_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print(f"\nQ1-2026 calibration: SBC add-back ${SBC_Q1/M:.0f}M")
print(f"Seasonal Q2/Q1 2025 step {QOQ_2025-1:+.1%} -> ${Q1_26_REV*QOQ_2025/B:.3f}B; guidance high implies {GUIDE_HI/Q1_26_REV-1:+.1%} q/q")
print("\n== Q2 2026 forecast (reports today AMC) ==")
print(f"{'':34s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}{'Guide hi':>12s}")
for label, key, _, street, guide in fields:
    vals = "".join(f"{Q2[k_][key]/M:>12,.0f}" if abs(Q2[k_][key]) > 10 else f"{Q2[k_][key]:>12.3f}"
                   for k_ in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.0f}" if abs(street) > 10 else f"{street:>12.3f}")
    g = "" if guide is None else (f"{guide/M:>12,.0f}" if abs(guide) > 10 else f"{guide:>12.3f}")
    print(f"{label:34s}{vals}{s}{g}")
print("\n== Valuation ==")
for k_ in ("Bear", "Base", "Bull"):
    v = VAL[k_]
    print(f"{k_:5s} FY27 rev {FY[k_]['rev27']/B:5.1f}B EBITDA {v['ebitda27']/B:5.1f}B EPS ${v['eps27']:6.2f} | "
          f"P/E ${v['px_pe']:7.2f}  EV/EBITDA ${v['px_ebitda']:7.2f} -> ${v['target']:7.2f} ({v['upside']:+.1%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW}) | street mean $668.84")
print(f"Net debt {NET_DEBT/B:.2f}B | EV {EV_NOW/B:.1f}B | EV/TTM rev {EV_NOW/ttm_rev:.1f}x | "
      f"EV/TTM adj EBITDA {EV_NOW/ttm_ebitda:.1f}x")
print(f"FY26E EPS ${fy26_eps:.2f} -> forward P/E {PX_NOW/fy26_eps:.1f}x | FY27E P/E {PX_NOW/VAL['Base']['eps27']:.1f}x")
