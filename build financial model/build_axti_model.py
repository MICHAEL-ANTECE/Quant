#!/usr/bin/env python3
"""
AXTI (AXT Inc.) model — same framework as MU / NBIS / SHOP / APP / SNDK.

DATA SOURCE CAVEAT: unlike the other five models, this one is built from public filings/
aggregators rather than the user's moomoo terminal screenshots. The income statement and
guidance are exact; the balance-sheet and TTM-ratio layers that the other workbooks carry
are absent. Send the moomoo Financials tabs (Income Statement / Balance Sheet / Cash Flow /
Key Indicators / Analyst Ratings) and those sheets can be filled in.

TIMING: AXT already reported Q2 2026 on 2026-07-30. This model therefore projects Q3 2026
(calendar quarter ending 2026-09-30, reports late October), not a quarter that is imminent.

THE ONE THING THAT MAKES THIS COMPANY DIFFERENT: management's revenue guidance is an
EXPORT-PERMIT-CONSTRAINED FLOOR, not a demand forecast. The Q3 number is explicitly
"revenue for which the company has high confidence based on existing export permits or
non-restricted products." Upside comes from Chinese export licences being granted during
the quarter — which is why the guide is systematically low and why no financial model can
price the tail risk on the other side.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["2026/Q2", "2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2",
        "2025/Q1", "2024/Q4", "2024/Q3"]
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Revenue", row(47.59*M, 26.92*M, 23.04*M, 27.96*M, 17.97*M, 19.36*M, 25.11*M, 23.65*M)),
    ("Gross Profit", row(21.37*M, 7.98*M, 4.83*M, 6.22*M, 1.43*M, -1.24*M, 4.41*M, 5.68*M)),
    ("Operating Income", row(10.42*M, -1.59*M, -3.83*M, -1.12*M, -6.75*M, -10.28*M, -6.17*M, -3.41*M)),
    ("Net Income", row(11.13*M, -1.62*M, -3.55*M, -1.91*M, -7.01*M, -8.80*M, -5.09*M, -2.94*M)),
    ("Diluted EPS", row(0.17, -0.03, -0.08, -0.04, -0.16, -0.20, -0.12, -0.07)),
]
# segment / operating detail disclosed with the Q2 print
Q2_DETAIL = [
    ("Indium phosphide (InP) revenue", 30.7*M, "record; the AI-datacenter optical driver"),
    ("Non-GAAP gross margin", 0.450, "up from 29.9% in Q1 2026"),
    ("Non-GAAP EPS", 0.19, "vs $0.07 consensus — a 171% beat"),
    ("Backlog", 100.0*M, "above $100M"),
    ("Revenue q/q", 0.768, "+77% sequential"),
    ("Revenue y/y", 1.648, "+164% year over year"),
]

# ============================================================== GUIDANCE =====
GUIDE = [
    ("Q3 2026 revenue guidance", "$66M", "explicitly permit-constrained: revenue with 'high confidence based on existing export permits or non-restricted products'"),
    ("Q3 2026 non-GAAP EPS guidance", "$0.30 - $0.32", "on ~66.5M shares"),
    ("Exit-2026 InP quarterly capacity", "~$60M per quarter", "company target"),
    ("Exit-2027 InP quarterly capacity", "~$130M per quarter", "company target — a doubling"),
    ("Q2 2026 report date", "2026-07-30 (already reported)", ""),
    ("Q3 2026 report date", "late October 2026 (est.)", ""),
    ("Analyst mean price target", "$91.60 - $96.50", "stockanalysis / street composite"),
    ("Analyst consensus rating", "Buy / broadly Overweight", ""),
    ("Notable dissent", "B. Riley price target cut to $52", ""),
    ("Structural risk", "Beijing Tongmei subsidiary + Chinese export licences on gallium / germanium / indium gate the revenue", ""),
]

# ========================================================== MARKET DATA ======
PX_NOW = 76.29
SHARES = 63.82e6
SHARES_GUIDE = 66.5e6          # company's own Q3 share-count assumption
MKT = [
    ("Last price (2026-08-06)", 76.29, "USD, +11.19% on the day"),
    ("Market cap", 4.87*B, "USD"),
    ("Shares outstanding", 63.82e6, "shares"),
    ("52-week high", 143.16, "USD"),
    ("52-week low", 1.96, "USD"),
    ("P/E (TTM)", 1063.39, "x — meaningless, TTM EPS is $0.07"),
    ("Forward P/E", 48.05, "x -> implies street forward EPS of ~$1.59"),
    ("TTM revenue", 125.51*M, "+45.8%"),
    ("TTM EPS", 0.07, "USD"),
    ("Analyst consensus", None, "Buy"),
    ("Analyst mean target", 91.60, "USD (+20.1%)"),
]

# ============================================ Q3-2026 FORECAST ENGINE ========
# Calibration: guidance $66M revenue -> $0.30-0.32 non-GAAP EPS on 66.5M shares
#   => guided net income ~= $20.6M, a 31.2% net margin. Backing into it with Q2's
#   $10.95M opex run-rate implies gross margin stepping to ~48-50%.
Q2_REV = 47.59*M
Q2_GM = 21.37/47.59
Q2_OPEX = (21.37 - 10.42) * M
GUIDE_REV = 66.0*M
GUIDE_EPS_LO, GUIDE_EPS_HI = 0.30, 0.32

SCEN = {
    #             revenue, gross margin, opex, other income, tax rate, shares
    "Bear": dict(rev=66.0*M,  gm=0.480, opex=12.0*M, other=1.0*M, tax=0.06, sh=66.5e6,
                 note="no incremental export permits — the guided floor is the ceiling"),
    "Base": dict(rev=75.0*M,  gm=0.500, opex=12.5*M, other=1.2*M, tax=0.08, sh=66.5e6,
                 note="some additional licences clear during the quarter"),
    "Bull": dict(rev=85.0*M,  gm=0.520, opex=13.0*M, other=1.5*M, tax=0.10, sh=66.5e6,
                 note="broad permit relief + InP capacity pulled forward"),
}


def forecast_q3(p):
    gp = p["rev"] * p["gm"]
    op = gp - p["opex"]
    pretax = op + p["other"]
    net = pretax * (1 - p["tax"])
    return dict(revenue=p["rev"], qoq=p["rev"]/Q2_REV-1, gross_profit=gp, gross_margin=p["gm"],
                opex=p["opex"], op_income=op, op_margin=op/p["rev"],
                net=net, net_margin=net/p["rev"], eps=net/p["sh"],
                vs_guide=p["rev"]/GUIDE_REV-1)


Q3 = {k: forecast_q3(v) for k, v in SCEN.items()}

# ================================= FY2026-2028 MODEL + VALUATION ============
FY25_REV = (19.36 + 17.97 + 27.96 + 23.04) * M          # $88.33M actual
FY = {
    "Bear": dict(q4=78*M, rev27=320*M, nm27=0.28, rev28=390*M, nm28=0.30,
                 pe27=18.0, pe28=16.0, sh=67e6),
    "Base": dict(q4=92*M, rev27=430*M, nm27=0.35, rev28=650*M, nm28=0.38,
                 pe27=28.0, pe28=25.0, sh=67e6),
    "Bull": dict(q4=105*M, rev27=540*M, nm27=0.40, rev28=880*M, nm28=0.42,
                 pe27=35.0, pe28=30.0, sh=68e6),
}
PROB = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}
DISCOUNT, YEARS_BACK = 0.15, 1.5


def valuation(k):
    p = FY[k]
    rev26 = 26.92*M + Q2_REV + Q3[k]["revenue"] + p["q4"]
    eps26 = -0.03 + 0.17 + Q3[k]["eps"] + (p["q4"] * p["nm27"] * 0.9) / p["sh"]
    eps27 = p["rev27"] * p["nm27"] / p["sh"]
    eps28 = p["rev28"] * p["nm28"] / p["sh"]
    px_m1 = eps27 * p["pe27"]                                   # FY27 earnings today
    px_m2 = eps28 * p["pe28"] / (1 + DISCOUNT) ** YEARS_BACK    # FY28, discounted back
    target = (px_m1 + px_m2) / 2
    return dict(rev26=rev26, eps26=eps26, eps27=eps27, eps28=eps28,
                px_m1=px_m1, px_m2=px_m2, target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("AXTI — AXT Inc. model", ""),
    ("Built", "2026-08-06, price $76.29 (+11.19% on the day)"),
    ("DATA CAVEAT", "Unlike the MU / NBIS / SHOP / APP / SNDK workbooks, this one is NOT built from the user's moomoo screenshots. Income statement and guidance are exact from filings/aggregators, but the balance-sheet, cash-flow and TTM-ratio layers are absent. Send the moomoo Financials tabs to fill them in."),
    ("TIMING", "Q2 2026 was already reported on 2026-07-30. This model projects Q3 2026 (Sept quarter, reports late October) — not an imminent event."),
    ("Latest REPORTED quarter", "Q2 2026: revenue $47.59M (+77% q/q, +164% y/y), non-GAAP gross margin 45.0%, operating income $10.42M, non-GAAP EPS $0.19 vs $0.07 consensus. First profitable quarter of the cycle."),
    ("THE KEY STRUCTURAL FACT", "Management's revenue guidance is an EXPORT-PERMIT-CONSTRAINED FLOOR, not a demand forecast. Q3's $66M is explicitly 'revenue for which the company has high confidence based on existing export permits or non-restricted products.' Beats come from Chinese export licences clearing mid-quarter. This cuts both ways and is not modellable."),
    ("Demand driver", "Indium phosphide substrates for 800G/1.6T optical transceivers in AI data centres. Q2 InP revenue $30.7M (record), backlog >$100M."),
    ("Capacity path", "Exit-2026 ~$60M/quarter InP capacity, doubling to ~$130M/quarter by end-2027 (company targets)."),
    ("Units", "USD; statement sheets display millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

# income statement + derived
ws = wb.create_sheet("IS_Quarterly")
ws.cell(row=1, column=1, value="Reported quarterly income statement. Calendar quarters (AXT's fiscal year = calendar year).").font = Font(italic=True, size=9, color="7F7F7F")
for j, lab in enumerate(["USD, millions"] + COLS, start=1):
    c = ws.cell(row=2, column=j, value=lab); c.fill = HDR if j == 1 else SUB; c.font = WHITE
    c.alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 40
for j in range(2, n+2):
    ws.column_dimensions[get_column_letter(j)].width = 13
ws.freeze_panes = ws.cell(row=3, column=2)
r = 3
for name, vals in IS:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.number_format = '0.00' if "EPS" in name else NUMFMT
        c.border = THIN
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
rev = dict(IS)["Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Income"]
ni = dict(IS)["Net Income"]
for name, vals, fmt in [
        ("Revenue q/q %", [rev[i]/rev[i+1]-1 if i+1 < n else None for i in range(n)], PCT),
        ("Revenue y/y %", [rev[i]/rev[i+4]-1 if i+4 < n else None for i in range(n)], PCT),
        ("Gross margin %", [gp[i]/rev[i] for i in range(n)], PCT),
        ("Operating margin %", [op[i]/rev[i] for i in range(n)], PCT),
        ("Net margin %", [ni[i]/rev[i] for i in range(n)], PCT),
        ("Implied opex (GP - op income)", [gp[i]-op[i] for i in range(n)], NUMFMT),
        ("Opex as % of revenue", [(gp[i]-op[i])/rev[i] for i in range(n)], PCT)]:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = fmt; c.border = THIN
    r += 1
r += 1
ws.cell(row=r, column=1, value="Q2 2026 DISCLOSED DETAIL").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
for name, val, note in Q2_DETAIL:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = PCT if val < 2 and val > 0 and "margin" in name.lower() or "q/q" in name or "y/y" in name else ('0.00' if val < 2 else NUMFMT)
    ws.cell(row=r, column=4, value=note).font = Font(size=9, color="7F7F7F")
    r += 1

# guidance + market
ws = wb.create_sheet("Guidance_Market")
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 84
r = 1
ws.cell(row=r, column=1, value="COMPANY GUIDANCE").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for a, b, c_ in GUIDE:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c_).alignment = Alignment(wrap_text=True, vertical="top")
    if "permit" in str(c_).lower() or "export" in str(c_).lower():
        for j in range(1, 4):
            ws.cell(row=r, column=j).fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="MARKET DATA").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.00' if (val is not None and abs(val) < 10000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1

# Q3 forecast
ws = wb.create_sheet("Forecast_Q3_2026")
ws.column_dimensions["A"].width = 40
for col in "BCDE":
    ws.column_dimensions[col].width = 16
ws.cell(row=1, column=1, value=f"Q2 2026 base: revenue ${Q2_REV/M:.2f}M, gross margin {Q2_GM:.1%}, implied opex ${Q2_OPEX/M:.2f}M.").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value=f"Guidance ${GUIDE_REV/M:.0f}M / ${GUIDE_EPS_LO}-{GUIDE_EPS_HI} EPS is a PERMIT-CONSTRAINED FLOOR — the bear case is simply 'the floor holds'.").font = Font(italic=True, size=9, color="C00000")
r = 4
ws.cell(row=r, column=1, value="Q3 2026 (reports late Oct)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Guidance"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, GUIDE_REV),
          ("Revenue q/q", "qoq", PCT, GUIDE_REV/Q2_REV-1),
          ("vs guidance", "vs_guide", '+0.0%;-0.0%', 0.0),
          ("Gross profit", "gross_profit", NUMFMT, None),
          ("Gross margin", "gross_margin", PCT, None),
          ("Operating expense", "opex", NUMFMT, None),
          ("Operating income", "op_income", NUMFMT, None),
          ("Operating margin", "op_margin", PCT, None),
          ("Net income", "net", NUMFMT, None),
          ("Net margin", "net_margin", PCT, None),
          ("Non-GAAP EPS", "eps", '0.00', GUIDE_EPS_HI)]
for label, key, fmt, guide in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q3[k][key]); c.number_format = fmt; c.border = THIN
    if guide is not None:
        c = ws.cell(row=r, column=5, value=guide); c.number_format = fmt
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=SCEN[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")

# valuation
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 46
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="Target = average of (FY2027E EPS x P/E) and (FY2028E EPS x P/E, discounted 1.5y at 15%). FY2028 matters because the company's own capacity plan doubles InP output through 2027.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt in [
        ("FY2025 actual revenue", lambda k: FY25_REV, NUMFMT),
        ("Q4 2026E revenue", lambda k: FY[k]["q4"], NUMFMT),
        ("FY2026E revenue", lambda k: VAL[k]["rev26"], NUMFMT),
        ("FY2026E growth", lambda k: VAL[k]["rev26"]/FY25_REV-1, PCT),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT),
        ("FY2027E net margin", lambda k: FY[k]["nm27"], PCT),
        ("FY2027E EPS", lambda k: VAL[k]["eps27"], '$0.00'),
        ("FY2028E revenue", lambda k: FY[k]["rev28"], NUMFMT),
        ("FY2028E EPS", lambda k: VAL[k]["eps28"], '$0.00'),
        ("P/E applied — FY2027", lambda k: FY[k]["pe27"], '0.0"x"'),
        ("P/E applied — FY2028", lambda k: FY[k]["pe28"], '0.0"x"'),
        ("Route 1 — FY2027 earnings", lambda k: VAL[k]["px_m1"], '$#,##0.00'),
        ("Route 2 — FY2028 discounted", lambda k: VAL[k]["px_m2"], '$#,##0.00'),
        ("12-month target (avg)", lambda k: VAL[k]["target"], '$#,##0.00'),
        ("Upside vs $76.29", lambda k: VAL[k]["upside"], '+0.0%;-0.0%')]:
    c0 = ws.cell(row=r, column=1, value=label)
    hi = "target" in label or "Upside" in label
    if hi:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if hi:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
mc = 4.87*B
for label, val, fmt in [
        ("Market cap / TTM revenue", mc/(125.51*M), '0.0"x"'),
        ("Market cap / FY2026E revenue (base)", mc/VAL["Base"]["rev26"], '0.0"x"'),
        ("Market cap / FY2027E revenue (base)", mc/FY["Base"]["rev27"], '0.0"x"'),
        ("Market cap / Q3-guidance annualised ($264M)", mc/(264*M), '0.0"x"'),
        ("P/E on FY2027E (base)", PX_NOW/VAL["Base"]["eps27"], '0.0"x"'),
        ("P/E on FY2028E (base)", PX_NOW/VAL["Base"]["eps28"], '0.0"x"'),
        ("Street mean PT $91.60 implies FY27 P/E of", 91.60/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk high $143.16 implies FY27 P/E of", 143.16/VAL["Base"]["eps27"], '0.0"x"'),
        ("B. Riley PT $52 implies FY27 P/E of", 52.00/VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027E EPS x P/E").font = BOLD
r += 1
mults = [15, 20, 25, 30, 35]
ws.cell(row=r, column=1, value="FY27 EPS \\ P/E").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x"); c.fill = SUB; c.font = WHITE
r += 1
for eps_ in (1.20, 1.80, 2.20, 2.80, 3.40):
    c = ws.cell(row=r, column=1, value=eps_); c.number_format = '$0.00'; c.fill = SUB; c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        c = ws.cell(row=r, column=j, value=eps_*m_)
        c.number_format = '$#,##0.00'
        if abs(eps_*m_ - PX_NOW) < 8:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/AXTI_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print(f"\nQ2 2026 actual: rev ${Q2_REV/M:.2f}M (+{(Q2_REV/(17.97*M)-1):.0%} y/y), GM {Q2_GM:.1%}, "
      f"op ${10.42:.2f}M, non-GAAP EPS $0.19 vs $0.07 consensus")
print(f"Q3 guidance: ${GUIDE_REV/M:.0f}M rev / ${GUIDE_EPS_LO}-{GUIDE_EPS_HI} EPS  (PERMIT-CONSTRAINED FLOOR)")
print("\n== Q3 2026 forecast (reports late Oct) ==")
print(f"{'':28s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Guidance':>12s}")
for label, key, _, guide in fields:
    vals = "".join(f"{Q3[k][key]/M:>12,.1f}" if abs(Q3[k][key]) > 100 else f"{Q3[k][key]:>12.3f}"
                   for k in ("Bear", "Base", "Bull"))
    g = "" if guide is None else (f"{guide/M:>12,.1f}" if abs(guide) > 100 else f"{guide:>12.3f}")
    print(f"{label:28s}{vals}{g}")
print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY26 ${v['rev26']/M:6.1f}M | FY27 ${FY[k]['rev27']/M:6.0f}M EPS ${v['eps27']:5.2f} | "
          f"FY28 EPS ${v['eps28']:5.2f} | M1 ${v['px_m1']:7.2f} M2 ${v['px_m2']:7.2f} -> ${v['target']:7.2f} ({v['upside']:+.1%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW})")
print(f"Street mean $91.60 implies {91.60/VAL['Base']['eps27']:.1f}x FY27E base EPS | "
      f"today = {PX_NOW/VAL['Base']['eps27']:.1f}x")
print(f"Mkt cap/TTM rev {4.87*B/(125.51*M):.1f}x | /FY26E {4.87*B/VAL['Base']['rev26']:.1f}x | "
      f"/FY27E {4.87*B/FY['Base']['rev27']:.1f}x")
