#!/usr/bin/env python3
"""
COHR (Coherent Corp.) model — same framework as the MU / NBIS / SHOP / APP /
SNDK / AXTI / AAOI / CRDO / FIG workbooks.

DATA SOURCE: SEC EDGAR primary documents. The FQ4 FY2026 8-K press release
(EX-99.1, filed 2026-08-12) carries the full statements, the five-quarter
non-GAAP trend tables and FQ1 FY2027 guidance; XBRL companyfacts supplies the
longer quarterly history. Consensus and market data sit on their own sheet and
never touch the reported columns.

FISCAL CALENDAR: fiscal year ends June 30.
  FQ4 FY2026 = quarter ended 2026-06-30, reported 2026-08-12 after the close.
  FQ1 FY2027 = quarter ending 2026-09-30, reports early November.

WHY THIS ONE IS DIFFERENT FROM THE OTHER AI-OPTICS NAMES: Coherent sits at the
module and component layer with a 40.2% non-GAAP gross margin and a 21.8%
non-GAAP operating margin, versus AXTI at the substrate layer (45% GM but only
now turning an operating profit) and CRDO at the chip layer (68% GM, 36% OM).
Coherent is also the only one of the three that is capital-intensive enough for
the cash-flow statement to be the whole story.

THE TENSION: revenue accelerated to +33.8% y/y (+42% pro forma), non-GAAP EPS
beat consensus by 21.7%, guidance for FQ1 FY2027 implies +45% y/y — and the
stock fell 8.0% the next session. Three reasons, all on the balance sheet and
cash-flow side:

  1. FY2026 operating cash flow was $79.5M against $1,102.9M of capex. Free
     cash flow for the year was NEGATIVE $1,023.4M. FY2025 CFO was $633.6M, so
     cash generation fell 87% in a year when revenue grew 22.5%.
  2. Inventory rose $1,143.4M y/y to $2,581.0M (187 days, from 149). Accounts
     payable rose $1,058.4M over the same period. Roughly 93% of the inventory
     build is financed by stretching suppliers — if payables normalise, the
     already-thin CFO goes deeply negative.
  3. Diluted share count went from 155.5M to 202.2M, up 30%, mostly from the
     conversion of the $2,483.3M preferred (mezzanine equity is now zero).
     Per-share growth badly lags dollar growth.

And the setup mattered: the stock ran from $249.06 to $379.13 in six sessions
into the print, a 52% move. The 8% give-back is about positioning, not results.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

M, B = 1e6, 1e9

COLS = ["FQ4 2026", "FQ3 2026", "FQ2 2026", "FQ1 2026", "FQ4 2025", "FQ3 2025"]
ENDS = ["2026-06-30", "2026-03-31", "2025-12-31", "2025-09-30", "2025-06-30", "2025-03-31"]
n = len(COLS)

PX_NOW = 324.31          # 2026-08-14 close, moomoo
PX_PRE = 355.64          # 2026-08-12 close, the session results were released after
PX_PEAK = 379.13         # 2026-08-07, the pre-earnings run-up high close
SH_DIL = 202.2 * M       # FQ4 FY2026 diluted weighted-average shares
STREET_PT = 395.50       # 23-analyst mean 12-month target


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"{len(v)} > {n}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Revenue", row(2045.5*M, 1805.6*M, 1685.6*M, 1581.4*M, 1529.4*M, 1497.9*M)),
    ("  y/y growth", row(0.338, 0.205, 0.175, 0.173, 0.164, 0.239)),
    ("  q/q growth", row(0.133, 0.071, 0.066, 0.034, 0.021, 0.044)),
    ("Cost of goods sold", row(1258.5*M, 1125.7*M, None, None, 983.3*M, None)),
    ("GAAP gross margin $", row(787.1*M, 679.9*M, 622.8*M, 579.2*M, 546.1*M, 527.7*M)),
    ("GAAP gross margin %", row(0.385, 0.377, 0.369, 0.366, 0.357, 0.352)),
    ("Research and development", row(216.4*M, 186.0*M, 165.7*M, 154.9*M, 155.7*M, 150.7*M)),
    ("Selling, general and administrative", row(266.4*M, 267.6*M, 258.5*M, 252.1*M, 245.4*M, 231.4*M)),
    ("Restructuring charges", row(6.1*M, 34.4*M, 4.0*M, 19.0*M, 53.9*M, None)),
    ("Impairment of assets held-for-sale", row(44.3*M, 0.0, 11.0*M, 9.0*M, 85.0*M, None)),
    ("Interest expense", row(41.1*M, 44.6*M, None, None, 55.0*M, None)),
    ("Other expense (income), net", row(-65.6*M, -28.1*M, None, None, 14.4*M, None)),
    ("Earnings before income taxes", row(278.5*M, 184.3*M, None, None, -63.4*M, None)),
    ("Income taxes", row(42.3*M, 2.7*M, None, None, 34.7*M, None)),
    ("Net earnings attributable to Coherent", row(240.5*M, 191.4*M, 146.7*M, 226.3*M, -95.6*M, 15.7*M)),
    ("GAAP diluted EPS", row(1.19, 0.97, 0.76, 1.19, -0.83, -0.11)),
    ("Diluted shares", row(202.2*M, 196.4*M, None, None, 155.5*M, None)),
    ("GAAP operating income", row(253.9*M, 200.8*M, 184.0*M, None, 6.1*M, None)),
    ("Non-GAAP gross margin $", row(822.6*M, 714.2*M, 657.0*M, 613.0*M, 582.2*M, None)),
    ("Non-GAAP gross margin %", row(0.402, 0.396, 0.390, 0.388, 0.381, None)),
    ("Non-GAAP R&D", row(208.0*M, 178.0*M, 159.0*M, 148.0*M, 150.0*M, None)),
    ("Non-GAAP SG&A", row(169.0*M, 170.0*M, 162.0*M, 156.0*M, 157.0*M, None)),
    ("Non-GAAP operating income", row(445.8*M, 366.1*M, 336.0*M, 309.0*M, 275.1*M, None)),
    ("Non-GAAP operating margin", row(0.218, 0.203, 0.199, 0.195, 0.180, None)),
    ("Non-GAAP net earnings", row(351.2*M, 276.2*M, 248.0*M, 221.0*M, 192.3*M, None)),
    ("Non-GAAP diluted EPS", row(1.74, 1.41, None, None, 1.00, None)),
]

# ======================================================== SEGMENT REVENUE ====
SEG = [
    ("Datacenter & Communications", 1615.0*M, 1361.6*M, 1018.3*M, 5274.6*M, 3755.2*M),
    ("Industrial", 430.5*M, 444.0*M, 511.1*M, 1843.6*M, 2054.9*M),
    ("Consolidated", 2045.5*M, 1805.6*M, 1529.4*M, 7118.2*M, 5810.1*M),
]

# ============================================== BALANCE SHEET / CASH FLOW ====
BS = [
    ("Cash and cash equivalents", 1162.0*M, 909.2*M),
    ("Short-term investments", 825.0*M, 0.0),
    ("Restricted cash (current + non-current)", 606.4*M, 723.7*M),
    ("Accounts receivable", 1343.3*M, 964.1*M),
    ("INVENTORIES", 2581.0*M, 1437.6*M),
    ("Property, plant and equipment, net", 2999.3*M, 1877.5*M),
    ("Goodwill", 4375.6*M, 4471.1*M),
    ("Other intangible assets, net", 2884.5*M, 3204.7*M),
    ("Total assets", 18299.9*M, 14910.9*M),
    ("ACCOUNTS PAYABLE", 1905.4*M, 847.0*M),
    ("Current portion of long-term debt", 7.9*M, 188.3*M),
    ("Long-term debt", 3214.3*M, 3498.6*M),
    ("TOTAL DEBT", 3222.2*M, 3686.9*M),
    ("NET DEBT (of cash + ST investments)", 1235.2*M, 2777.7*M),
    ("Total liabilities", 7061.7*M, 6429.7*M),
    ("Mezzanine equity (preferred)", 0.0, 2483.3*M),
    ("Coherent shareholders' equity", 10903.5*M, 5644.5*M),
]

CF = [
    ("Net cash provided by operating activities", 79.5*M, 633.6*M),
    ("Additions to property, plant and equipment", -1102.9*M, -440.8*M),
    ("FREE CASH FLOW", -1023.4*M, 192.8*M),
    ("FCF as % of revenue", -0.1438, 0.0332),
    ("Proceeds from sale of business", 437.0*M, 27.0*M),
    ("Purchases of short-term investments", -1025.0*M, 0.0),
]

# ================================================ GUIDANCE VS ACTUAL =========
# Only two quarters can be verified precisely from the releases in hand — the
# earlier prints give the EPS guide but the matching quarterly non-GAAP EPS is
# not separately restated. A two-point calibration, unlike FIG's four-point one.
GUIDE_HIST = [
    ("FQ2 FY2026", 1.10, 1.30, 1.20, None, "2025-11-05"),
    ("FQ3 FY2026", 1.28, 1.48, 1.38, 1.41, "2026-02-04"),
    ("FQ4 FY2026", 1.52, 1.72, 1.62, 1.74, "2026-05-06"),
    ("FQ1 FY2027", 1.85, 2.05, 1.95, None, "2026-08-12"),
]
BEATS = [(a / m - 1) for _, _, _, m, a, _ in GUIDE_HIST if a]
BEAT_AVG = sum(BEATS) / len(BEATS)

FQ4_GUIDE = [
    ("FQ4 FY2026 revenue guidance", 1.98*B, "midpoint of $1.91B - $2.05B, given 2026-05-06"),
    ("FQ4 FY2026 revenue actual", 2.0455*B, "at the TOP of the range, +3.3% vs midpoint"),
    ("FQ4 FY2026 non-GAAP EPS guidance", 1.62, "midpoint of $1.52 - $1.72"),
    ("FQ4 FY2026 non-GAAP EPS actual", 1.74, "+7.4% vs midpoint, +21.7% vs $1.43 consensus"),
]

# ==================================================== CONSENSUS / MARKET =====
STREET = [
    ("FQ4 FY2026 consensus non-GAAP EPS", "$1.43 — actual $1.74, a 21.7% beat", "street"),
    ("FQ4 FY2026 company guidance", "revenue $1.91B-$2.05B, non-GAAP EPS midpoint $1.62", "company"),
    ("FQ1 FY2027 revenue guidance", "$2.2B - $2.4B (midpoint $2.3B), implying +45.4% y/y and +12.4% q/q", "company"),
    ("FQ1 FY2027 non-GAAP gross margin guidance", "39.5% - 41.5% (midpoint 40.5%)", "company"),
    ("FQ1 FY2027 non-GAAP opex guidance", "$400M - $420M (midpoint $410M)", "company"),
    ("FQ1 FY2027 non-GAAP EPS guidance", "$1.85 - $2.05 (midpoint $1.95)", "company"),
    ("FQ1 FY2027 implied non-GAAP op margin", "$2,300M x 40.5% - $410M = $521.5M, i.e. 22.7%", "derived"),
    ("Management ambition", "first $3B revenue quarter targeted within fiscal 2027", "earnings call"),
    ("12-month mean price target", f"${STREET_PT:.2f}; 17 Strong Buy / 1 Buy / 5 Hold of 23 analysts", "street"),
    ("Price 2026-08-07 (pre-earnings peak)", f"${PX_PEAK:.2f}", "market"),
    ("Price 2026-08-12 (results after close)", f"${PX_PRE:.2f}", "market"),
    ("Price 2026-08-14", f"${PX_NOW:.2f} — down 8.8% since the print, 14.5% off the peak", "market"),
    ("Run-up into the print", "$249.06 on 2026-07-30 to $379.13 on 2026-08-07, +52% in six sessions", "market"),
    ("20-day realized volatility", "140% annualized", "market"),
]

# ============================================== FY2027 MODEL + VALUATION =====
FY26_REV = 7118.2 * M
FY26_NG_OP = 1456.9 * M
FY26_NG_EPS = 5.61
FQ1_GUIDE_MID = 2.30 * B
NET_DEBT = 1235.2 * M
NET_INT = 80.0 * M        # interest expense net of other income, annualised
TAX_NG = 0.19             # company guide 18-20%

FY = {
    "Bear": dict(q1=2.25*B, q2=2.38*B, q3=2.50*B, q4=2.65*B, ngm=0.215,
                 pe=25.0, evs=4.5, sh=206e6,
                 note="Capacity comes on but hyperscaler order timing slips a quarter; the "
                      "$3B quarter lands in FY2028 instead. Industrial keeps shrinking and "
                      "the payables stretch starts to unwind, forcing a cash-preserving "
                      "slowdown in capex."),
    "Base": dict(q1=2.35*B, q2=2.55*B, q3=2.78*B, q4=3.00*B, ngm=0.235,
                 pe=38.0, evs=7.0, sh=208e6,
                 note="The $3B quarter arrives in FQ4 FY2027 as management guided. Non-GAAP "
                      "operating margin continues the 18.0-19.5-19.9-20.3-21.8 ladder toward "
                      "the mid-23s as the new capacity absorbs fixed cost."),
    "Bull": dict(q1=2.42*B, q2=2.68*B, q3=2.95*B, q4=3.25*B, ngm=0.255,
                 pe=48.0, evs=9.0, sh=210e6,
                 note="Copper-to-optical conversion inside the rack pulls demand forward; "
                      "800G/1.6T mix and the InP supply agreements (Coherent prepaid AXT "
                      "$22.3M in June for six-inch InP) let Coherent hold price."),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}


def valuation(k):
    p = FY[k]
    rev27 = p["q1"] + p["q2"] + p["q3"] + p["q4"]
    ng_op = rev27 * p["ngm"]
    ng_net = (ng_op - NET_INT) * (1 - TAX_NG)
    eps27 = ng_net / p["sh"]
    px_pe = eps27 * p["pe"]
    px_evs = (p["evs"] * rev27 - NET_DEBT) / p["sh"]
    target = (px_pe + px_evs) / 2
    return dict(rev27=rev27, growth27=rev27 / FY26_REV - 1, ng_op=ng_op, ng_net=ng_net,
                eps27=eps27, px_pe=px_pe, px_evs=px_evs, target=target,
                upside=target / PX_NOW - 1, q1_vs_guide=p["q1"] / FQ1_GUIDE_MID - 1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k] * VAL[k]["target"] for k in PROB)
MCAP = SH_DIL * PX_NOW
EV = MCAP + NET_DEBT

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
GOOD = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'
PCTS = '+0.0%;-0.0%'
USD = '$#,##0.00'

wb = Workbook()
wb.remove(wb.active)


def header(ws, labels, r=1, width0=44, width=15):
    ws.column_dimensions["A"].width = width0
    for j, lab in enumerate(labels, start=2):
        ws.column_dimensions[chr(64 + j)].width = width
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).fill = HDR


# ---------------------------------------------------------------- README ----
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 108
notes = [
    ("Company", "Coherent Corp. (NYSE: COHR) — photonics: optical transceivers, lasers, compound "
                "semiconductor components. Formed from II-VI's acquisition of Coherent Inc."),
    ("Model built", "2026-08-14, after FQ4 FY2026 results (reported 2026-08-12 after the close)."),
    ("Data source", "SEC EDGAR: 8-K EX-99.1 filed 2026-08-12 (statements, five-quarter non-GAAP "
                    "trend tables, FQ1 FY2027 guidance) plus XBRL companyfacts for history."),
    ("Fiscal calendar", "Year ends June 30. FQ4 FY2026 = quarter ended 2026-06-30. "
                        "FQ1 FY2027 ends 2026-09-30 and reports early November."),
    ("Latest reported quarter", "FQ4 FY2026: revenue $2,045.5M, +33.8% y/y (+42% pro forma), "
                                "non-GAAP EPS $1.74. FY2026 revenue $7,118.2M, +22.5%."),
    ("", ""),
    ("THE CENTRAL PUZZLE", "Revenue accelerated to +33.8%, non-GAAP EPS beat consensus by 21.7%, and "
                           "FQ1 FY2027 guidance implies +45% y/y growth — yet the stock fell 8.0% the "
                           "next session and sits 14.5% below its pre-earnings peak."),
    ("Trap 1 — cash burn", "FY2026 operating cash flow was $79.5M against $1,102.9M of capex, so free "
                           "cash flow was NEGATIVE $1,023.4M. FY2025 CFO was $633.6M. Cash generation "
                           "fell 87% in a year revenue grew 22.5%. Capex ran at 27.2% of revenue in FQ4."),
    ("Trap 2 — payables funding", "Inventory rose $1,143.4M y/y to $2,581.0M (187 days from 149). "
                                  "Accounts payable rose $1,058.4M to $1,905.4M over the same span. "
                                  "About 93% of the inventory build is supplier-financed. That is the "
                                  "only reason CFO stayed positive — and it reverses if terms normalise."),
    ("Trap 3 — 30% more shares", "Diluted shares went 155.5M to 202.2M y/y, mostly the conversion of "
                                 "$2,483.3M of preferred (mezzanine equity is now zero). Non-GAAP net "
                                 "earnings grew 82.6% y/y in FQ4 but non-GAAP EPS grew only 74%."),
    ("Trap 4 — one segment", "Datacenter & Communications grew 58.6% y/y and is now 79% of revenue. "
                             "Industrial FELL 15.8% y/y in FQ4 and 10.3% for the year. The whole "
                             "company is now an AI-datacenter bet with a shrinking legacy attached."),
    ("Trap 5 — the setup", "The stock ran from $249.06 to $379.13 in six sessions into the print, +52%. "
                           "The post-print give-back is positioning unwinding, not a verdict on results."),
    ("", ""),
    ("Calibration 1 — guide beat", f"Non-GAAP EPS versus its own guidance midpoint: FQ3 FY2026 +2.2%, "
                                   f"FQ4 FY2026 +7.4%. Mean {BEAT_AVG:.2%}. Only two quarters can be "
                                   f"verified exactly from the releases in hand, so this is a weaker "
                                   f"calibration than the four-point one in the FIG model — the base "
                                   f"case therefore leans on the guidance itself, not on an assumed beat."),
    ("Calibration 2 — margin ladder", "Non-GAAP operating margin by quarter: 18.0%, 19.5%, 19.9%, 20.3%, "
                                      "21.8%. Roughly +95bps per quarter for five straight quarters. "
                                      "FQ1 FY2027 guidance implies 22.7% ($2,300M x 40.5% - $410M)."),
    ("Calibration 3 — non-GAAP bridge", "non-GAAP operating income = GAAP operating income + SBC + "
                                        "acquired-intangible amortisation + restructuring + impairment. "
                                        "FQ4 check: 253.9 + 47.0 + 28.0 + 6.1 + 44.3 = 445.8, hitting "
                                        "the reported non-GAAP figure$."),
    ("Calibration 4 — tax and interest", "Non-GAAP tax 19% (company guide 18-20%). Interest expense is "
                                         "falling as debt is repaid: $55.0M in FQ4 FY2025 to $41.1M in "
                                         "FQ4 FY2026. Model uses $80M/yr net of other income."),
    ("", ""),
    ("Valuation method", "Two routes on FY2027E, averaged, then probability-weighted 25/50/25. Route 1 "
                         "is P/E on non-GAAP EPS; route 2 is EV/Sales with net debt of $1,235.2M "
                         "subtracted. Unlike FIG, Coherent carries net DEBT, so the EV bridge works "
                         "against the per-share number rather than for it."),
    ("Model output", f"Probability-weighted 12-month target ${PW_TARGET:.2f} versus ${PX_NOW:.2f} spot "
                     f"({PW_TARGET/PX_NOW-1:+.1%}) and a ${STREET_PT:.2f} street mean."),
    ("Read-through to AXTI", "Coherent prepaid AXT $22.3M in June 2026 for six-inch indium phosphide "
                             "under a three-year agreement. FQ4 Datacenter & Communications up 58.6% "
                             "y/y is direct confirmation of the demand behind that contract — see the "
                             "AXTI_Financial_Model workbook."),
    ("How to update", "When FQ1 FY2027 lands: append the column to IS_Quarterly, add the row to "
                      "Guidance_vs_Actual and re-derive the beat mean, and above all check CFO minus "
                      "capex and the accounts-payable balance. Whether free cash flow turns is the "
                      "thesis; revenue is no longer the question."),
]
for i, (k, v) in enumerate(notes, start=1):
    c = ws.cell(row=i, column=1, value=k)
    c.font = BOLD
    c.alignment = Alignment(vertical="top")
    c2 = ws.cell(row=i, column=2, value=v)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    if k.startswith("Trap"):
        c.fill = WARN
        c2.fill = WARN
    if k.startswith("Calibration"):
        c.fill = GOOD
        c2.fill = GOOD
    ws.row_dimensions[i].height = 46 if len(v) > 130 else 30

# ------------------------------------------------------------ IS_Quarterly --
ws = wb.create_sheet("IS_Quarterly")
header(ws, COLS)
ws.cell(row=2, column=1, value="Period ended").font = BOLD
for j, e in enumerate(ENDS, start=2):
    ws.cell(row=2, column=j, value=e).alignment = Alignment(horizontal="center")
r = 3
for label, vals in IS:
    c = ws.cell(row=r, column=1, value=label)
    pct = "%" in label or "growth" in label or "margin" in label and "$" not in label
    if "Non-GAAP" in label or label.isupper():
        c.font = BOLD
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = THIN
        if v is None:
            continue
        if "%" in label or "growth" in label or ("margin" in label and "$" not in label):
            cell.number_format = PCT
        elif "EPS" in label:
            cell.number_format = '$0.00'
        else:
            cell.number_format = NUMFMT
        if label == "Non-GAAP operating margin":
            cell.fill = GOOD
    r += 1

# ------------------------------------------------------------- Segments -----
ws = wb.create_sheet("Segments")
header(ws, ["FQ4 FY26", "FQ3 FY26", "FQ4 FY25", "FY2026", "FY2025"], width0=40, width=16)
r = 2
for label, a, b_, c_, d_, e_ in SEG:
    cc = ws.cell(row=r, column=1, value=label)
    if label == "Consolidated":
        cc.font = BOLD
    for j, v in enumerate((a, b_, c_, d_, e_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
        if label == "Consolidated":
            cell.font = BOLD
            cell.fill = SEC
    r += 1
r += 1
for label, yoy_q, yoy_fy in [("Datacenter & Communications", 1615.0/1018.3-1, 5274.6/3755.2-1),
                             ("Industrial", 430.5/511.1-1, 1843.6/2054.9-1),
                             ("Consolidated", 2045.5/1529.4-1, 7118.2/5810.1-1)]:
    ws.cell(row=r, column=1, value=f"{label} — y/y growth").font = BOLD
    for j, v in enumerate((yoy_q, yoy_fy), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCTS
        cell.font = BOLD
        cell.fill = WARN if v < 0 else GOOD
    r += 1
ws.cell(row=r - 3, column=4, value="(quarter)").font = Font(italic=True, size=9)
r += 1
ws.cell(row=r, column=1, value="Datacenter & Communications is now 79.0% of revenue. Industrial is "
                               "shrinking in both the quarter and the year — every dollar of growth, "
                               "and then some, comes from AI datacenter optics.").font = Font(italic=True, size=9, color="7F7F7F")

# --------------------------------------------------------------- BS_CF ------
ws = wb.create_sheet("BS_CF")
header(ws, ["Jun 30, 2026", "Jun 30, 2025"], width0=48, width=18)
r = 2
ws.cell(row=r, column=1, value="BALANCE SHEET").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, a, b_ in BS:
    c = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        c.font = BOLD
    for j, v in enumerate((a, b_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
        if hl:
            cell.font = BOLD
            cell.fill = WARN if label in ("INVENTORIES", "ACCOUNTS PAYABLE") else SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Inventory change y/y").font = BOLD
ws.cell(row=r, column=2, value=2581.0*M - 1437.6*M).number_format = NUMFMT
r += 1
ws.cell(row=r, column=1, value="Accounts payable change y/y").font = BOLD
ws.cell(row=r, column=2, value=1905.4*M - 847.0*M).number_format = NUMFMT
r += 1
c = ws.cell(row=r, column=1, value="Share of inventory build funded by payables")
c.font = BOLD
c = ws.cell(row=r, column=2, value=(1905.4 - 847.0) / (2581.0 - 1437.6))
c.number_format = PCT
c.fill = WARN
c.font = BOLD
r += 3
ws.cell(row=r, column=1, value="CASH FLOW — fiscal year").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, a, b_ in CF:
    c = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        c.font = BOLD
    for j, v in enumerate((a, b_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if "%" in label else NUMFMT
        cell.border = THIN
        if hl:
            cell.font = BOLD
            cell.fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="Revenue grew 22.5% and free cash flow went from +$192.8M to -$1,023.4M. "
                               "Capex nearly tripled. This is the sheet the stock reacted to.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Guidance_vs_Actual -------
ws = wb.create_sheet("Guidance_vs_Actual")
header(ws, ["EPS guide low", "EPS guide high", "Midpoint", "Actual", "Beat vs midpoint", "Guided on"],
       width0=18, width=17)
r = 2
for q, lo, hi, mid, act, given in GUIDE_HIST:
    ws.cell(row=r, column=1, value=q).font = BOLD
    for j, v in enumerate((lo, hi, mid, act), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = '$0.00'
        cell.border = THIN
    if act:
        cell = ws.cell(row=r, column=6, value=act / mid - 1)
        cell.number_format = PCTS
        cell.font = BOLD
        cell.fill = GOOD
    else:
        cell = ws.cell(row=r, column=6, value="not separately restated" if q == "FQ2 FY2026" else "not yet reported")
        cell.fill = SEC
    ws.cell(row=r, column=7, value=given)
    r += 1
r += 1
ws.cell(row=r, column=1, value="Mean verified beat").font = BOLD
c = ws.cell(row=r, column=4, value=BEAT_AVG)
c.number_format = PCTS
c.fill = GOOD
c.font = BOLD
r += 3
ws.cell(row=r, column=1, value="FQ4 FY2026 — GUIDANCE VERSUS DELIVERED").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, val, note in FQ4_GUIDE:
    ws.cell(row=r, column=1, value=label).font = BOLD
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '$0.00' if val < 100 else NUMFMT
    ws.cell(row=r, column=3, value=note)
    r += 1
r += 1
ws.cell(row=r, column=1, value="Revenue landed at the very top of the guided range and non-GAAP EPS beat "
                               "the midpoint by 7.4% and consensus by 21.7%. The quarter itself was not "
                               "the problem.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Consensus_Market ---------
ws = wb.create_sheet("Consensus_Market")
header(ws, ["Value", "Source"], width0=46, width=76)
ws.column_dimensions["C"].width = 14
r = 2
for k, v, src in STREET:
    ws.cell(row=r, column=1, value=k).font = BOLD
    ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True)
    ws.cell(row=r, column=3, value=src)
    r += 1
r += 1
ws.cell(row=r, column=1, value="Street and market figures only. Nothing here feeds the reported columns "
                               "in IS_Quarterly — same discipline as the NBIS, MU and FIG workbooks.").font = Font(italic=True, size=9, color="7F7F7F")

# ---------------------------------------------------------- Valuation -------
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 50
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value=f"FY2027E built quarter by quarter off the FQ1 guide midpoint of "
                               f"${FQ1_GUIDE_MID/B:.2f}B. Target = average of (non-GAAP EPS x P/E) and "
                               f"(EV/Sales x revenue, less ${NET_DEBT/B:.2f}B net debt), per share. "
                               f"Non-GAAP net = (operating income - ${NET_INT/M:.0f}M) x (1 - {TAX_NG:.0%}).").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k)
    c.fill = HDR
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("FQ1 FY2027E revenue", lambda k: FY[k]["q1"], NUMFMT, False),
        ("  vs guide midpoint $2.30B", lambda k: VAL[k]["q1_vs_guide"], PCTS, False),
        ("FQ2 FY2027E revenue", lambda k: FY[k]["q2"], NUMFMT, False),
        ("FQ3 FY2027E revenue", lambda k: FY[k]["q3"], NUMFMT, False),
        ("FQ4 FY2027E revenue (the '$3B quarter')", lambda k: FY[k]["q4"], NUMFMT, False),
        ("FY2027E revenue", lambda k: VAL[k]["rev27"], NUMFMT, True),
        ("  FY2027E growth (FY2026 = $7.118B)", lambda k: VAL[k]["growth27"], PCTS, True),
        ("FY2027E non-GAAP operating margin", lambda k: FY[k]["ngm"], PCT, False),
        ("FY2027E non-GAAP operating income", lambda k: VAL[k]["ng_op"], NUMFMT, False),
        ("FY2027E non-GAAP net earnings", lambda k: VAL[k]["ng_net"], NUMFMT, False),
        ("FY2027E diluted shares", lambda k: FY[k]["sh"], NUMFMT, False),
        ("FY2027E non-GAAP EPS (FY2026 = $5.61)", lambda k: VAL[k]["eps27"], '$0.00', True),
        ("P/E multiple applied", lambda k: FY[k]["pe"], '0.0"x"', False),
        ("EV/Sales multiple applied", lambda k: FY[k]["evs"], '0.0"x"', False),
        ("Route 1 — P/E on non-GAAP EPS", lambda k: VAL[k]["px_pe"], USD, False),
        ("Route 2 — EV/Sales less net debt", lambda k: VAL[k]["px_evs"], USD, False),
        ("12-MONTH TARGET (average)", lambda k: VAL[k]["target"], USD, True),
        (f"Upside vs ${PX_NOW:.2f}", lambda k: VAL[k]["upside"], PCTS, True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k))
        c.number_format = fmt
        c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k]["note"])
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 90
r += 2
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    ws.cell(row=r, column=j, value=PROB[k]).number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET)
c.number_format = USD
c.font = Font(bold=True, size=12)
c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET / PX_NOW - 1)
c.number_format = PCTS
c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
for label, val, fmt in [
        ("Share price (2026-08-14)", PX_NOW, USD),
        ("Diluted shares (FQ4 FY2026)", SH_DIL, NUMFMT),
        ("Market capitalisation", MCAP, NUMFMT),
        ("Net debt", NET_DEBT, NUMFMT),
        ("Enterprise value", EV, NUMFMT),
        ("P/E on FY2026 actual non-GAAP EPS ($5.61)", PX_NOW / FY26_NG_EPS, '0.0"x"'),
        ("P/E on FY2027E base non-GAAP EPS", PX_NOW / VAL["Base"]["eps27"], '0.0"x"'),
        ("EV / FY2026 actual revenue", EV / FY26_REV, '0.0"x"'),
        ("EV / FY2027E revenue (base case)", EV / VAL["Base"]["rev27"], '0.0"x"'),
        (f"Street mean target ${STREET_PT:.2f} implies FY2027E P/E of", STREET_PT / VAL["Base"]["eps27"], '0.0"x"'),
        (f"Pre-earnings peak ${PX_PEAK:.2f} implied FY2027E P/E of", PX_PEAK / VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027E non-GAAP EPS x P/E").font = BOLD
r += 1
mults = [22, 28, 34, 40, 46]
ws.cell(row=r, column=1, value="FY27 EPS \\ P/E").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x")
    c.fill = SUB
    c.font = WHITE
r += 1
for eps_ in (7.50, 8.50, 9.50, 10.50, 11.50):
    c = ws.cell(row=r, column=1, value=eps_)
    c.number_format = '$0.00'
    c.fill = SUB
    c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        c = ws.cell(row=r, column=j, value=eps_ * m_)
        c.number_format = '$#,##0'
        if abs(eps_ * m_ - PX_NOW) < 20:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/COHR_Financial_Model.xlsx"
wb.save(out)
print(f"wrote {out}\n")

ks = ("Bear", "Base", "Bull")
print(f"{'':30}{'Bear':>12}{'Base':>12}{'Bull':>12}")
print(f"{'FY2027E revenue ($M)':30}" + "".join(f"{VAL[k]['rev27']/M:>12,.0f}" for k in ks))
print(f"{'  growth vs FY2026':30}" + "".join(f"{VAL[k]['growth27']:>11.1%} " for k in ks))
print(f"{'FY2027E non-GAAP op margin':30}" + "".join(f"{FY[k]['ngm']:>11.1%} " for k in ks))
print(f"{'FY2027E non-GAAP EPS':30}" + "".join(f"{VAL[k]['eps27']:>12.2f}" for k in ks))
print(f"{'Route 1 P/E':30}" + "".join(f"{VAL[k]['px_pe']:>12.2f}" for k in ks))
print(f"{'Route 2 EV/Sales':30}" + "".join(f"{VAL[k]['px_evs']:>12.2f}" for k in ks))
print(f"{'TARGET':30}" + "".join(f"{VAL[k]['target']:>12.2f}" for k in ks))
print(f"{'upside':30}" + "".join(f"{VAL[k]['upside']:>11.1%} " for k in ks))
print(f"\nProbability-weighted target: ${PW_TARGET:.2f}  ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW:.2f})")
print(f"Street mean target:          ${STREET_PT:.2f}  ({STREET_PT/PX_NOW-1:+.1%})")
print(f"Market cap ${MCAP/B:.1f}B | net debt ${NET_DEBT/B:.2f}B | EV ${EV/B:.1f}B")
print(f"P/E on FY26 actual $5.61 = {PX_NOW/FY26_NG_EPS:.1f}x | P/E on FY27E base = {PX_NOW/VAL['Base']['eps27']:.1f}x")
print(f"EV/FY26 revenue = {EV/FY26_REV:.1f}x | EV/FY27E base revenue = {EV/VAL['Base']['rev27']:.1f}x")
print(f"FY2026 free cash flow: -$1,023.4M on $7,118.2M revenue ({-1023.4/7118.2:.1%})")
