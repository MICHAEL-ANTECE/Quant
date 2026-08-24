#!/usr/bin/env python3
"""
CRDO (Credo Technology Group) model — same framework as the MU / NBIS / SHOP / APP /
SNDK / AXTI / AAOI workbooks.

DATA CAVEAT: built from filings/aggregators, not the user's moomoo screenshots. Income
statement, guidance and market data are exact; the balance-sheet / cash-flow / TTM-ratio
layers the screenshot-based workbooks carry are absent.

FISCAL CALENDAR: Credo's fiscal year ends in late April / early May.
  FQ4 FY2026 = quarter ended 2026-05-02, reported 2026-06-01 (latest reported)
  FQ1 FY2027 = quarter ended 2026-08-01 — already CLOSED, reports 2026-09-02 (~4 weeks out)

WHY THIS ONE IS DIFFERENT FROM AXTI / AAOI: same AI-optics supply chain, but Credo sits at
the chip layer (SerDes, optical DSP, AEC controllers) and earns a 68% gross margin with a
36% operating margin. AXTI (substrate) earns 45% GM; AAOI (module assembly) earns 29% GM
and still loses money. The layer of the stack determines the economics.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["FQ4 2026", "FQ3 2026", "FQ2 2026", "FQ1 2026",
        "FQ4 2025", "FQ3 2025", "FQ2 2025", "FQ1 2025"]
ENDS = ["2026-05-02", "2026-01-31", "2025-11-01", "2025-08-02",
        "2025-05-03", "2025-02-01", "2024-11-02", "2024-08-03"]
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Revenue", row(437.00*M, 407.01*M, 268.03*M, 223.07*M,
                    170.03*M, 135.00*M, 72.03*M, 59.71*M)),
    ("Gross Profit", row(298.07*M, 278.87*M, 181.05*M, 150.37*M,
                         114.19*M, 85.93*M, 45.51*M, 37.28*M)),
    ("Operating Income", row(155.85*M, 149.62*M, 78.80*M, 60.74*M,
                             33.79*M, 26.19*M, -8.41*M, -14.45*M)),
    ("Net Income (GAAP)", row(169.10*M, 157.14*M, 82.64*M, 63.40*M,
                              36.59*M, 29.36*M, -4.23*M, -9.54*M)),
    ("Diluted EPS (GAAP)", row(0.88, 0.82, 0.44, 0.34, 0.20, 0.16, -0.03, -0.06)),
]

# ============================================================== GUIDANCE =====
GUIDE = [
    ("FQ1 FY2027 report date", "2026-09-02 (quarter already closed 2026-08-01)", "company"),
    ("FQ1 FY2027 revenue guidance", "$465M - $475M (midpoint $470M)", "company, given 2026-06-01"),
    ("FQ1 FY2027 consensus revenue", "$470.4M — virtually identical to the guidance midpoint", "street"),
    ("FY2027 revenue guidance", "more than 80% growth -> $2.41B+", "management"),
    ("FY2027 growth driver", "optical DSP and other optics products ramping", "management"),
    ("FQ4 FY2026 actual", "revenue $437M (+157% y/y) vs $433.3M expected; non-GAAP EPS $1.16 vs $1.03", "reported 2026-06-01"),
    ("FY2026 full year", "revenue $1.34B (more than tripled); non-GAAP net income $662M (up 5x+)", "company"),
    ("Post-earnings reaction", "stock SOLD OFF on 2026-06-01 — the beat was fine but guidance was merely inline", "market"),
    ("Analyst consensus", "Strong Buy, average target $279.29 (+13.9%)", "stockanalysis"),
    ("Analyst target range", "~$175 - $300; Mizuho raised to $290 (Outperform) post-print", "street"),
]

# ========================================================== MARKET DATA ======
PX_NOW = 245.22
SHARES_OUT = 186.48e6
DILUTED_FQ4 = 169.10 / 0.88 * 1e6      # implied diluted share count from reported EPS
MKT = [
    ("Last price (2026-08-06)", 245.22, "USD, +6.42% on the day"),
    ("Market cap", 45.73*B, "USD"),
    ("Shares outstanding", 186.48e6, "shares"),
    ("Implied diluted shares (FQ4 net / EPS)", DILUTED_FQ4, "shares"),
    ("52-week high", 308.67, "USD"),
    ("52-week low", 86.49, "USD"),
    ("P/E (TTM, GAAP)", 97.70, "x"),
    ("Forward P/E", 40.03, "x -> implies street FY2027 non-GAAP EPS of ~$6.13"),
    ("TTM revenue", 1.34*B, "+205.7%"),
    ("TTM EPS (GAAP)", 2.51, "+765.5%"),
    ("Analyst mean target", 279.29, "USD (+13.9%)"),
    ("Next earnings", None, "2026-09-02"),
]

# ==================================== FQ1 FY2027 FORECAST ENGINE =============
FQ4_REV = 437.00*M
FQ4_GM = 298.07/437.00
FQ4_OPEX = (298.07 - 155.85) * M          # $142.22M
GUIDE_LO, GUIDE_HI, GUIDE_MID = 465.0*M, 475.0*M, 470.0*M
ST_REV = 470.4*M
# non-GAAP / GAAP EPS ratio, calibrated on FQ4 FY2026: $1.16 non-GAAP vs $0.88 GAAP
NONGAAP_RATIO = 1.16 / 0.88
ST_EPS_NONGAAP = 1.25                      # street FQ1 estimate, derived from the FY27 path

SCEN = {
    #             revenue, gross margin, opex, other income, tax, diluted shares
    "Bear": dict(rev=470.0*M, gm=0.678, opex=155.0*M, other=11.0*M, tax=0.04, sh=194.0e6,
                 note="guide midpoint exactly; optical mix starts to dilute gross margin"),
    "Base": dict(rev=482.0*M, gm=0.685, opex=152.0*M, other=12.0*M, tax=0.03, sh=193.0e6,
                 note="modest beat, gross margin holds at the FQ4 level"),
    "Bull": dict(rev=500.0*M, gm=0.692, opex=150.0*M, other=13.0*M, tax=0.03, sh=192.5e6,
                 note="optical DSP pulls forward; the H2 reacceleration starts early"),
}


def forecast_fq1(p):
    gp = p["rev"] * p["gm"]
    op = gp - p["opex"]
    pretax = op + p["other"]
    net = pretax * (1 - p["tax"])
    eps_gaap = net / p["sh"]
    return dict(revenue=p["rev"], qoq=p["rev"]/FQ4_REV-1, yoy=p["rev"]/(223.07*M)-1,
                gross_profit=gp, gross_margin=p["gm"], opex=p["opex"],
                op_income=op, op_margin=op/p["rev"], net=net,
                eps_gaap=eps_gaap, eps_nongaap=eps_gaap*NONGAAP_RATIO,
                vs_street=p["rev"]/ST_REV-1)


FQ1 = {k: forecast_fq1(v) for k, v in SCEN.items()}

# ================================= FY2027-2028 MODEL + VALUATION ============
FY26_REV = (437.00 + 407.01 + 268.03 + 223.07) * M       # $1,335.11M actual
FY26_NONGAAP_NI = 662.0*M
FY26_NONGAAP_MARGIN = FY26_NONGAAP_NI / FY26_REV          # 49.6%

FY = {
    "Bear": dict(rev27=2.15*B, nm27=0.480, rev28=2.75*B, nm28=0.500,
                 pe27=25.0, pe28=22.0, sh27=195e6, sh28=202e6,
                 note="FY27 falls short of the >80% guide; a top customer in-sources AECs"),
    "Base": dict(rev27=2.45*B, nm27=0.520, rev28=3.55*B, nm28=0.550,
                 pe27=38.0, pe28=32.0, sh27=195e6, sh28=200e6,
                 note="the >80% FY27 guide lands; optical DSP ramps as promised"),
    "Bull": dict(rev27=2.70*B, nm27=0.545, rev28=4.60*B, nm28=0.575,
                 pe27=48.0, pe28=42.0, sh27=194e6, sh28=198e6,
                 note="scale-up Ethernet share gains on top of the optics ramp"),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}
DISCOUNT, YEARS_BACK = 0.12, 1.5


def valuation(k):
    p = FY[k]
    eps27 = p["rev27"] * p["nm27"] / p["sh27"]
    eps28 = p["rev28"] * p["nm28"] / p["sh28"]
    px_m1 = eps27 * p["pe27"]
    px_m2 = eps28 * p["pe28"] / (1 + DISCOUNT) ** YEARS_BACK
    target = (px_m1 + px_m2) / 2
    return dict(growth27=p["rev27"]/FY26_REV-1, growth28=p["rev28"]/p["rev27"]-1,
                eps27=eps27, eps28=eps28, px_m1=px_m1, px_m2=px_m2,
                target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("CRDO — Credo Technology Group model", ""),
    ("Built", "2026-08-06, price $245.22 (+6.42% on the day)"),
    ("DATA CAVEAT", "Built from filings/aggregators, NOT the user's moomoo screenshots. Income statement, guidance and market data are exact; balance-sheet / cash-flow / TTM-ratio sheets are absent. Send the moomoo Financials tabs to fill them in."),
    ("FISCAL CALENDAR", "Fiscal year ends late April / early May. FQ4 FY2026 = quarter ended 2026-05-02 (reported 2026-06-01). FQ1 FY2027 = quarter ended 2026-08-01 — ALREADY CLOSED, reports 2026-09-02."),
    ("TIMING", "Unlike SNDK / APP / AAOI, this is NOT an imminent event — the print is roughly four weeks away."),
    ("Latest REPORTED quarter", "FQ4 FY2026: revenue $437M (+157% y/y), gross margin 68.21%, operating income $155.85M (35.7%), GAAP EPS $0.88 / non-GAAP $1.16 vs $1.03 consensus."),
    ("THE RAMP", "Revenue went from $59.71M (FQ1 FY2025) to $437M (FQ4 FY2026) — 7.3x in eight quarters — while flipping from a -$14.45M operating loss to a +$155.85M operating profit. FY2026 revenue more than tripled to $1.34B; non-GAAP net income rose 5x+ to $662M (a 49.6% non-GAAP net margin)."),
    ("WHY IT IS NOT AXTI OR AAOI", "Same AI-optics supply chain, three different layers. Credo sits at the CHIP layer (SerDes, optical DSP, AEC controllers): 68% gross margin, 36% operating margin. AXTI at the SUBSTRATE layer: 45% GM. AAOI at MODULE ASSEMBLY: 29% GM and still loss-making. The layer of the stack sets the economics."),
    ("THE TENSION IN THE GUIDE", "FQ1 guidance implies only +7.6% q/q, but the FY2027 '>80% growth' guide needs roughly +18%/+15%/+12% sequential steps across FQ2-FQ4. The whole year is back-loaded onto the optical DSP ramp."),
    ("Units", "USD; statement sheet displays millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("IS_Quarterly")
ws.cell(row=1, column=1, value="Reported quarterly income statement. Fiscal quarters — see row 3 for period-end dates.").font = Font(italic=True, size=9, color="7F7F7F")
for j, lab in enumerate(["USD, millions"] + COLS, start=1):
    c = ws.cell(row=2, column=j, value=lab); c.fill = HDR if j == 1 else SUB; c.font = WHITE
    c.alignment = Alignment(horizontal="center")
ws.cell(row=3, column=1, value="period ended").font = Font(italic=True, size=9, color="7F7F7F")
for j, e in enumerate(ENDS, start=2):
    ws.cell(row=3, column=j, value=e).font = Font(italic=True, size=9, color="7F7F7F")
ws.column_dimensions["A"].width = 42
for j in range(2, n+2):
    ws.column_dimensions[get_column_letter(j)].width = 14
ws.freeze_panes = ws.cell(row=4, column=2)
r = 4
for name, vals in IS:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.number_format = '0.00' if "EPS" in name else NUMFMT
        c.border = THIN
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
rev = dict(IS)["Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Income"]
ni = dict(IS)["Net Income (GAAP)"]
for name, vals, fmt in [
        ("Revenue q/q %", [rev[i]/rev[i+1]-1 if i+1 < n else None for i in range(n)], PCT),
        ("Revenue y/y %", [rev[i]/rev[i+4]-1 if i+4 < n else None for i in range(n)], PCT),
        ("Gross margin %", [gp[i]/rev[i] for i in range(n)], PCT),
        ("Operating margin %", [op[i]/rev[i] for i in range(n)], PCT),
        ("Net margin % (GAAP)", [ni[i]/rev[i] for i in range(n)], PCT),
        ("Implied opex (GP - op income)", [gp[i]-op[i] for i in range(n)], NUMFMT),
        ("Opex as % of revenue", [(gp[i]-op[i])/rev[i] for i in range(n)], PCT),
        ("Opex q/q %", [(gp[i]-op[i])/(gp[i+1]-op[i+1])-1 if i+1 < n else None for i in range(n)], PCT),
        ("Incremental operating margin (d-op / d-rev)",
         [(op[i]-op[i+1])/(rev[i]-rev[i+1]) if i+1 < n else None for i in range(n)], PCT)]:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = fmt; c.border = THIN
    r += 1

ws = wb.create_sheet("Guidance_Market")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 58
ws.column_dimensions["C"].width = 34
r = 1
ws.cell(row=r, column=1, value="COMPANY GUIDANCE & STREET").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for a, b, c_ in GUIDE:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    r += 1
r += 1
ws.cell(row=r, column=1, value="MARKET DATA").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.00' if (val is not None and abs(val) < 10000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1

ws = wb.create_sheet("Forecast_FQ1_2027")
ws.column_dimensions["A"].width = 42
for col in "BCDEF":
    ws.column_dimensions[col].width = 15
ws.cell(row=1, column=1, value=f"FQ4 base: revenue ${FQ4_REV/M:.0f}M, gross margin {FQ4_GM:.2%}, opex ${FQ4_OPEX/M:.2f}M, GAAP EPS $0.88 on {DILUTED_FQ4/1e6:.1f}M diluted shares.").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value=f"Non-GAAP/GAAP EPS ratio calibrated on FQ4: $1.16 / $0.88 = {NONGAAP_RATIO:.3f}. Guidance ${GUIDE_LO/M:.0f}-{GUIDE_HI/M:.0f}M; consensus ${ST_REV/M:.1f}M sits on the midpoint.").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="FQ1 FY2027 (closed 8/1, reports 9/2)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
c = ws.cell(row=r, column=6, value="Guide (hi)"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, ST_REV, GUIDE_HI),
          ("Revenue q/q", "qoq", PCT, ST_REV/FQ4_REV-1, GUIDE_HI/FQ4_REV-1),
          ("Revenue y/y", "yoy", PCT, ST_REV/(223.07*M)-1, None),
          ("vs street", "vs_street", '+0.0%;-0.0%', None, None),
          ("Gross profit", "gross_profit", NUMFMT, None, None),
          ("Gross margin", "gross_margin", PCT, None, None),
          ("Operating expense", "opex", NUMFMT, None, None),
          ("Operating income", "op_income", NUMFMT, None, None),
          ("Operating margin", "op_margin", PCT, None, None),
          ("Net income (GAAP)", "net", NUMFMT, None, None),
          ("EPS (GAAP)", "eps_gaap", '0.00', None, None),
          ("EPS (non-GAAP)", "eps_nongaap", '0.00', ST_EPS_NONGAAP, None)]
for label, key, fmt, street, guide in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=FQ1[k][key]); c.number_format = fmt; c.border = THIN
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    if guide is not None:
        c = ws.cell(row=r, column=6, value=guide); c.number_format = fmt
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=SCEN[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")
r += 3
ws.cell(row=r, column=1, value="THE FY2027 BACK-LOADING PROBLEM").font = Font(bold=True, color="C00000")
r += 1
ws.cell(row=r, column=1, value="FY2027 guide of >80% growth on $1.335B = $2.41B+. With FQ1 at the $470M guide, the remaining three quarters must average $647M:").alignment = Alignment(wrap_text=True)
r += 1
path = [470, 555, 638, 715]
for j, (lab, v) in enumerate(zip(["FQ1 (guided)", "FQ2E", "FQ3E", "FQ4E"], path), start=1):
    ws.cell(row=r, column=j, value=lab).font = BOLD
    c = ws.cell(row=r+1, column=j, value=v*M); c.number_format = NUMFMT
    if j > 1:
        c2 = ws.cell(row=r+2, column=j, value=path[j-1]/path[j-2]-1); c2.number_format = PCT
ws.cell(row=r+2, column=1, value="implied q/q").font = Font(italic=True, size=9)
r += 4
ws.cell(row=r, column=1, value=f"Sum = ${sum(path)}M vs the $2,410M the guide requires. FQ1's guided +7.6% q/q has to become +18% / +15% / +12%.").font = Font(italic=True, size=9, color="C00000")

ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 48
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value=f"Non-GAAP net margin anchored on FY2026 actual: $662M on $1,335M = {FY26_NONGAAP_MARGIN:.1%}. Target = average of (FY2027E EPS x P/E) and (FY2028E EPS x P/E, discounted 1.5y at 12%).").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("FY2026 actual revenue", lambda k: FY26_REV, NUMFMT, False),
        ("FY2026 actual non-GAAP net income", lambda k: FY26_NONGAAP_NI, NUMFMT, False),
        ("FY2026 non-GAAP net margin", lambda k: FY26_NONGAAP_MARGIN, PCT, False),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT, False),
        ("FY2027E growth (guide: >80%)", lambda k: VAL[k]["growth27"], PCT, True),
        ("FY2027E non-GAAP net margin", lambda k: FY[k]["nm27"], PCT, False),
        ("FY2027E non-GAAP EPS", lambda k: VAL[k]["eps27"], '$0.00', False),
        ("FY2028E revenue", lambda k: FY[k]["rev28"], NUMFMT, False),
        ("FY2028E growth", lambda k: VAL[k]["growth28"], PCT, False),
        ("FY2028E non-GAAP EPS", lambda k: VAL[k]["eps28"], '$0.00', False),
        ("P/E applied — FY2027", lambda k: FY[k]["pe27"], '0.0"x"', False),
        ("P/E applied — FY2028", lambda k: FY[k]["pe28"], '0.0"x"', False),
        ("Route 1 — FY2027 earnings", lambda k: VAL[k]["px_m1"], '$#,##0.00', False),
        ("Route 2 — FY2028 discounted", lambda k: VAL[k]["px_m2"], '$#,##0.00', False),
        ("12-month target (avg)", lambda k: VAL[k]["target"], '$#,##0.00', True),
        ("Upside vs $245.22", lambda k: VAL[k]["upside"], '+0.0%;-0.0%', True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")
r += 3
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
mc = 45.73*B
for label, val, fmt in [
        ("Market cap / TTM revenue ($1.335B)", mc/FY26_REV, '0.0"x"'),
        ("Market cap / FY2027E revenue (base)", mc/FY["Base"]["rev27"], '0.0"x"'),
        ("Market cap / FY2028E revenue (base)", mc/FY["Base"]["rev28"], '0.0"x"'),
        ("P/E on FY2027E base non-GAAP EPS", PX_NOW/VAL["Base"]["eps27"], '0.0"x"'),
        ("P/E on FY2028E base non-GAAP EPS", PX_NOW/VAL["Base"]["eps28"], '0.0"x"'),
        ("Street mean PT $279.29 implies FY27 P/E of", 279.29/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk high $308.67 implies FY27 P/E of", 308.67/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk low $86.49 implies FY27 P/E of", 86.49/VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027E non-GAAP EPS x P/E").font = BOLD
r += 1
mults = [25, 30, 35, 40, 45]
ws.cell(row=r, column=1, value="FY27 EPS \\ P/E").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x"); c.fill = SUB; c.font = WHITE
r += 1
for eps_ in (5.00, 5.75, 6.50, 7.25, 8.00):
    c = ws.cell(row=r, column=1, value=eps_); c.number_format = '$0.00'; c.fill = SUB; c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        c = ws.cell(row=r, column=j, value=eps_*m_)
        c.number_format = '$#,##0'
        if abs(eps_*m_ - PX_NOW) < 20:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/CRDO_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print("\nEight-quarter ramp:")
for i in range(n-1, -1, -1):
    inc = (op[i]-op[i+1])/(rev[i]-rev[i+1]) if i+1 < n else float('nan')
    print(f"   {COLS[i]:9s} ({ENDS[i]})  rev ${rev[i]/M:7.2f}M  GM {gp[i]/rev[i]:6.2%}  "
          f"op {op[i]/M:+8.2f}M ({op[i]/rev[i]:+6.1%})  incr-op-margin {inc:6.1%}")
print(f"\nFQ4 non-GAAP/GAAP EPS ratio = {NONGAAP_RATIO:.3f}")
print(f"FQ1 FY2027 guidance ${GUIDE_LO/M:.0f}-{GUIDE_HI/M:.0f}M | consensus ${ST_REV/M:.1f}M (on the midpoint)")
print("\n== FQ1 FY2027 forecast (closed 8/1, reports 9/2) ==")
print(f"{'':30s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}")
for label, key, _, street, guide in fields:
    vals = "".join(f"{FQ1[k][key]/M:>12,.1f}" if abs(FQ1[k][key]) > 100 else f"{FQ1[k][key]:>12.3f}"
                   for k in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.1f}" if abs(street) > 100 else f"{street:>12.3f}")
    print(f"{label:30s}{vals}{s}")
print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY27 ${FY[k]['rev27']/B:4.2f}B (+{v['growth27']:.0%}) EPS ${v['eps27']:5.2f} | "
          f"FY28 ${FY[k]['rev28']/B:4.2f}B EPS ${v['eps28']:5.2f} | "
          f"M1 ${v['px_m1']:7.2f} M2 ${v['px_m2']:7.2f} -> ${v['target']:7.2f} ({v['upside']:+.1%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW}) | street $279.29")
print(f"Today = {PX_NOW/VAL['Base']['eps27']:.1f}x FY27E base EPS | street PT = {279.29/VAL['Base']['eps27']:.1f}x")
print(f"Mkt cap/TTM rev {mc/FY26_REV:.1f}x | /FY27E {mc/FY['Base']['rev27']:.1f}x | /FY28E {mc/FY['Base']['rev28']:.1f}x")
