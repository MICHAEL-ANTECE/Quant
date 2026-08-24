#!/usr/bin/env python3
"""
FIG (Figma, Inc.) model — same framework as the MU / NBIS / SHOP / APP / SNDK /
AXTI / AAOI / CRDO workbooks.

DATA SOURCE: SEC EDGAR primary documents, not moomoo screenshots — the Q2 2026
8-K press release (EX-99.1, filed 2026-08-05) carries the full three statements
plus the GAAP-to-non-GAAP reconciliation, and the 10-Q was filed the same day so
XBRL companyfacts already covers the quarter. Every figure below is traceable to
one of those two. Consensus and guidance live on their own sheet and are never
mixed into the reported columns.

FISCAL CALENDAR: calendar year. Q2 2026 = quarter ended 2026-06-30, reported
2026-08-05 after the close. Q3 2026 ends 2026-09-30.

WHY THIS ONE IS DIFFERENT: Figma beat revenue AND raised full-year guidance, and
the stock fell 14.8% the next session (-17.3% over two). The model exists to
answer why. Three things the headline hides:

  1. The full-year non-GAAP operating income guide did NOT move ($125-135M in
     May, $125-135M in August) even though the revenue guide went up $40M. The
     incremental operating margin on the raise is ZERO. One quarter earlier the
     same raise carried 45% incremental margin.
  2. Billings (revenue + change in deferred revenue) grew 33.7% y/y in Q2 while
     revenue grew 48.2%. Deferred revenue actually FELL $0.9M sequentially. The
     revenue acceleration is being fed by the deferred balance, not by bookings.
  3. Reported Free Cash Flow ignores the cash paid to settle RSU tax
     withholding, which sits in financing. Q2 FCF of $53.2M is $7.7M after that
     $45.5M outflow; first-half "FCF" of $141.8M is NEGATIVE $19.8M.

SBC is the other elephant: $147.6M in Q2 = 39.9% of revenue. It is decaying fast
off the IPO catch-up (2025Q3 was $1,138M, 415% of revenue) but the entire gap
between the $(117.3)M GAAP operating loss and the $36.1M non-GAAP operating
income is compensation.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

M, B = 1e6, 1e9

COLS = ["2026Q2", "2026Q1", "2025Q4", "2025Q3", "2025Q2", "2025Q1"]
ENDS = ["2026-06-30", "2026-03-31", "2025-12-31", "2025-09-30", "2025-06-30", "2025-03-31"]
n = len(COLS)

PX_NOW = 23.29          # 2026-08-07 close, moomoo
PX_PRE = 28.15          # 2026-08-05 close, the session before results
SH_DIL = 542.583 * M    # Q2 2026 non-GAAP diluted weighted-average shares
STREET_PT = 30.50       # 14-analyst mean 12-month target


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"{len(v)} > {n}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
# Q2/Q1 2026 and Q2/Q1 2025 split out of the press-release three- and six-month
# columns; 2025Q3/Q4 from XBRL companyfacts.
IS = [
    ("Revenue", row(370.083*M, 333.439*M, 303.8*M, 274.2*M, 249.640*M, 228.199*M)),
    ("Cost of revenue", row(60.472*M, 68.666*M, 54.3*M, 83.9*M, 27.889*M, 19.452*M)),
    ("Gross profit", row(309.611*M, 264.773*M, 249.5*M, 190.3*M, 221.751*M, 208.747*M)),
    ("Gross margin", row(0.8366, 0.7941, 0.8213, 0.6940, 0.8883, 0.9147)),
    ("Research and development", row(167.329*M, 172.974*M, 195.8*M, 680.9*M, 83.052*M, 69.925*M)),
    ("Sales and marketing", row(154.856*M, 125.568*M, None, None, 97.701*M, 68.840*M)),
    ("General and administrative", row(104.715*M, 103.629*M, None, None, 38.922*M, 30.233*M)),
    ("Total operating expenses", row(426.900*M, 402.171*M, 445.0*M, 1327.1*M, 219.675*M, 169.022*M)),
    ("GAAP operating income (loss)", row(-117.289*M, -137.398*M, -195.5*M, -1136.8*M, 2.076*M, 39.749*M)),
    ("GAAP operating margin", row(-0.3169, -0.4121, -0.6435, -4.1459, 0.0083, 0.1742)),
    ("Other income, net", row(7.614*M, -4.325*M, None, None, 36.978*M, 7.274*M)),
    ("Provision for income taxes", row(2.477*M, 0.678*M, None, None, 10.827*M, 2.141*M)),
    ("GAAP net income (loss)", row(-112.152*M, -142.401*M, -226.6*M, -1097.0*M, 28.227*M, 44.882*M)),
    ("GAAP diluted EPS", row(-0.21, -0.27, -0.44, -2.72, 0.00, 0.04)),
    ("Stock-based compensation", row(147.554*M, 168.998*M, 218.3*M, 1138.3*M, 7.310*M, 0.197*M)),
    ("SBC as % of revenue", row(0.3987, 0.5068, 0.7186, 4.1514, 0.0293, 0.0009)),
    ("Non-GAAP gross profit", row(313.999*M, 274.559*M, None, None, 223.954*M, 208.833*M)),
    ("Non-GAAP operating income", row(36.093*M, 52.133*M, None, None, 11.472*M, 40.032*M)),
    ("Non-GAAP operating margin", row(0.0975, 0.1563, None, None, 0.0460, 0.1754)),
    ("Non-GAAP net income", row(42.567*M, 56.503*M, None, None, 19.783*M, 41.726*M)),
    ("Non-GAAP diluted EPS", row(0.08, 0.10, None, None, 0.00, 0.07)),
    ("Diluted shares (non-GAAP)", row(542.583*M, 543.704*M, None, None, 231.702*M, 231.386*M)),
]

# ============================================== BALANCE SHEET / CASH FLOW ====
BS = [
    ("Cash and cash equivalents", 445.845*M, 403.469*M),
    ("Marketable securities", 1221.293*M, 1252.474*M),
    ("Digital assets (current + non-current)", 25.935*M, 30.691*M),
    ("Restricted cash", 9.800*M, 9.799*M),
    ("TOTAL CASH AND INVESTMENTS", 1702.873*M, 1696.433*M),
    ("Accounts receivable, net", 190.876*M, 247.915*M),
    ("Goodwill", 101.396*M, 101.396*M),
    ("Total assets", 2351.274*M, 2348.207*M),
    ("Deferred revenue", 626.783*M, 595.334*M),
    ("Accrued compensation and benefits", 53.528*M, 107.105*M),
    ("Total liabilities", 873.367*M, 837.566*M),
    ("Total debt", 0.0, 0.0),
    ("Additional paid-in capital", 3178.403*M, 2950.007*M),
    ("Accumulated deficit", -1697.926*M, -1443.373*M),
    ("Total stockholders' equity", 1477.907*M, 1510.641*M),
]

# Q2 2026, Q1 2026, Q2 2025 — the cash-flow story with the RSU-tax correction
CF = [
    ("Net cash provided by operating activities", 60.893*M, 97.308*M, 62.455*M),
    ("Capital expenditures", -6.688*M, -7.812*M, -1.134*M),
    ("Capitalized internal-use software", -0.995*M, -0.888*M, -0.718*M),
    ("REPORTED FREE CASH FLOW", 53.210*M, 88.608*M, 60.603*M),
    ("Reported FCF margin", 0.1438, 0.2657, 0.2427),
    ("Taxes paid on net share settlement of RSUs", -45.489*M, -116.159*M, 0.0),
    ("FREE CASH FLOW AFTER RSU TAX SETTLEMENT", 7.721*M, -27.551*M, 60.603*M),
    ("Adjusted FCF margin", 0.0209, -0.0826, 0.2427),
]

# ===================================================== KPI / UNIT ECONOMICS ==
# Billings = revenue + change in deferred revenue.  Deferred revenue moved
# +$32.330M in Q1 2026 and -$0.881M in Q2 2026 (six-month change +$31.449M).
KPI = [
    ("Revenue y/y growth", 0.4825, 0.4612, 0.4000, 0.3800, 0.4093, None),
    ("Revenue q/q growth", 0.1100, 0.0977, 0.1080, 0.0985, 0.0939, None),
    ("Change in deferred revenue", -0.881*M, 32.330*M, None, None, 26.511*M, 25.273*M),
    ("Billings (revenue + change in DR)", 369.202*M, 365.769*M, None, None, 276.151*M, 253.472*M),
    ("Billings y/y growth", 0.3370, 0.4430, None, None, None, None),
    ("Net Dollar Retention Rate", 1.36, None, None, None, None, None),
    ("Paid customers > $10K ARR", 15964, None, None, None, 11913, None),
    ("Paid customers > $100K ARR", 1635, None, None, None, 1120, None),
]

# ================================================ GUIDANCE VS ACTUAL =========
# The single most useful calibration in this model: Figma has beaten its own
# revenue guidance every quarter as a public company, and the beat is WIDENING.
GUIDE_HIST = [
    # quarter, guide low, guide high, midpoint, actual, given on
    ("Q3 2025", 263.0*M, 265.0*M, 264.0*M, 274.2*M, "2025-09-03"),
    ("Q4 2025", 292.0*M, 294.0*M, 293.0*M, 303.8*M, "2025-11-05"),
    ("Q1 2026", 315.0*M, 317.0*M, 316.0*M, 333.439*M, "2026-02-18"),
    ("Q2 2026", 348.0*M, 350.0*M, 349.0*M, 370.083*M, "2026-05-14"),
    ("Q3 2026", 373.0*M, 375.0*M, 374.0*M, None, "2026-08-05"),
]
BEATS = [(a / m - 1) for _, _, _, m, a, _ in GUIDE_HIST if a]
BEAT_AVG = sum(BEATS) / len(BEATS)
BEAT_L2 = sum(BEATS[-2:]) / 2

# Full-year guidance walk — where the margin story broke
FY_WALK = [
    ("FY2025 @ Q2 2025", 1.023*B, 93.0*M, "2025-09-03"),
    ("FY2025 @ Q3 2025", 1.045*B, 114.5*M, "2025-11-05"),
    ("FY2025 actual", 1.055839*B, None, "reported"),
    ("FY2026 @ Q4 2025", 1.370*B, 105.0*M, "2026-02-18"),
    ("FY2026 @ Q1 2026", 1.425*B, 130.0*M, "2026-05-14"),
    ("FY2026 @ Q2 2026", 1.465*B, 130.0*M, "2026-08-05"),
]

# ==================================================== CONSENSUS / MARKET =====
STREET = [
    ("Q2 2026 consensus revenue", "$352M — actual $370.1M, a 5.1% beat", "street"),
    ("Q2 2026 consensus non-GAAP EPS", "$0.04 — actual $0.08, a 100% beat", "street"),
    ("Q3 2026 company guidance", "$373.0M - $375.0M (midpoint $374.0M), implying 36% y/y", "company"),
    ("FY2026 company guidance", "$1.463B - $1.467B, +39% y/y, raised $40M from May", "company"),
    ("FY2026 non-GAAP operating income guidance", "$125M - $135M — UNCHANGED from May despite the $40M revenue raise", "company"),
    ("12-month mean price target", f"${STREET_PT:.2f} across 14 analysts, consensus rating Buy", "street"),
    ("Morgan Stanley post-print action", "price target cut to $33 from $38", "street"),
    ("Share price before results (2026-08-05)", f"${PX_PRE:.2f}", "market"),
    ("Share price after (2026-08-07)", f"${PX_NOW:.2f} — down 17.3% over two sessions", "market"),
    ("20-day realized volatility", "100% annualized", "market"),
    ("Stock year-to-date 2026", "down roughly 44%; down about 80% from the post-IPO peak", "market"),
]

# ========================================================= Q3 2026 MODEL ====
Q2_REV = 370.083 * M
Q3_GUIDE_MID = 374.0 * M
Q3_LY = 274.2 * M

SCEN_Q3 = {
    "Bear": dict(beat=0.020, ngm=0.070,
                 note="Beat narrows toward 2% as Config spend and AI credit COGS bite; "
                      "the billings slowdown proves to be real demand, not billing timing."),
    "Base": dict(beat=BEAT_L2, ngm=0.095,
                 note="The last-two-quarter average beat of 5.8% repeats. Seat expansion "
                      "plus the first full quarters of AI credit monetization hold."),
    "Bull": dict(beat=0.080, ngm=0.120,
                 note="Beat widens again as it has every quarter; Code Layers and the Figma "
                      "agent pull AI credit consumption above plan."),
}


def forecast_q3(p):
    rev = Q3_GUIDE_MID * (1 + p["beat"])
    ng_op = rev * p["ngm"]
    return dict(revenue=rev, qoq=rev / Q2_REV - 1, yoy=rev / Q3_LY - 1,
                vs_guide=rev / Q3_GUIDE_MID - 1, ng_op=ng_op, ng_margin=p["ngm"])


Q3 = {k: forecast_q3(v) for k, v in SCEN_Q3.items()}

# FY2026 = H1 actual + Q3 scenario + Q4 (Q3 grown at the scenario's q/q rate)
H1_26 = 703.522 * M
FY25 = 1.055839 * B
FY26_GUIDE = 1.465 * B

Q4_QOQ = {"Bear": 0.045, "Base": 0.070, "Bull": 0.090}
FY26 = {}
for k in Q3:
    q4 = Q3[k]["revenue"] * (1 + Q4_QOQ[k])
    fy = H1_26 + Q3[k]["revenue"] + q4
    FY26[k] = dict(q4=q4, fy=fy, growth=fy / FY25 - 1, vs_guide=fy / FY26_GUIDE - 1)

# ============================================= FY2027 MODEL + VALUATION =====
NET_CASH = 1702.873 * M
INTEREST = 62.0 * M       # ~3.6% on $1.70B of cash and securities
TAX_NG = 0.145            # company's stated non-GAAP tax rate

FY = {
    "Bear": dict(g27=0.26, ngm27=0.10, evs=4.0, pe=30.0, sh=550e6,
                 note="AI-native design tools and model-vendor bundling compress seat growth; "
                      "NDR slides toward 120% and the margin guide stays parked at 9-10%."),
    "Base": dict(g27=0.33, ngm27=0.14, evs=7.0, pe=50.0, sh=555e6,
                 note="Growth decays gracefully from 43% to 33%; opex leverage finally arrives "
                      "once the Config and IPO-cohort hiring anniversaries pass."),
    "Bull": dict(g27=0.40, ngm27=0.18, evs=9.5, pe=65.0, sh=565e6,
                 note="AI credits become a genuine second revenue engine on top of seats; "
                      "Figma becomes the system of record for full-stack creation."),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}


def valuation(k):
    p = FY[k]
    rev27 = FY26[k]["fy"] * (1 + p["g27"])
    ng_op = rev27 * p["ngm27"]
    ng_net = (ng_op + INTEREST) * (1 - TAX_NG)
    eps27 = ng_net / p["sh"]
    px_evs = (p["evs"] * rev27 + NET_CASH) / p["sh"]     # route 1: EV/Sales
    px_pe = eps27 * p["pe"]                              # route 2: P/E
    target = (px_evs + px_pe) / 2
    return dict(rev27=rev27, ng_op=ng_op, ng_net=ng_net, eps27=eps27,
                px_evs=px_evs, px_pe=px_pe, target=target, upside=target / PX_NOW - 1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k] * VAL[k]["target"] for k in PROB)
MCAP = SH_DIL * PX_NOW
EV = MCAP - NET_CASH

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
GOOD = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'
PCTS = '+0.0%;-0.0%'
USD = '$#,##0.00'

wb = Workbook()
wb.remove(wb.active)


def header(ws, labels, r=1, width0=46, width=15):
    ws.column_dimensions["A"].width = width0
    for j, lab in enumerate(labels, start=2):
        ws.column_dimensions[chr(64 + j)].width = width
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).fill = HDR


# ---------------------------------------------------------------- README ----
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 108
notes = [
    ("Company", "Figma, Inc. (NYSE: FIG) — collaborative design and build platform. IPO July 2025."),
    ("Model built", "2026-08-10, after Q2 2026 results (reported 2026-08-05 after the close)."),
    ("Data source", "SEC EDGAR: 8-K EX-99.1 press release filed 2026-08-05 plus XBRL companyfacts. "
                    "Not moomoo screenshots — EDGAR is the primary document here and the 10-Q was "
                    "filed the same day, so the quarter is fully tagged."),
    ("Latest reported quarter", "2026Q2 (ended 2026-06-30): revenue $370.083M, +48.2% y/y."),
    ("Next report", "Q3 2026, quarter ends 2026-09-30, expected early November."),
    ("", ""),
    ("THE CENTRAL PUZZLE", "Revenue beat, full-year guidance raised $40M, and the stock fell 14.8% "
                           "the next session and 17.3% over two. The three sheets below explain it."),
    ("Trap 1 — margin guide frozen", "FY2026 non-GAAP operating income guidance is $125-135M today and was "
                                     "$125-135M in May, while the revenue guide went from $1.425B to $1.465B. "
                                     "Incremental operating margin on the raise: ZERO. The prior raise "
                                     "(Feb to May) carried +$25M op income on +$55M revenue, i.e. 45%."),
    ("Trap 2 — billings", "Billings (revenue + change in deferred revenue) grew 33.7% y/y in Q2 versus "
                          "revenue at 48.2%. Deferred revenue FELL $0.9M sequentially. Billings growth "
                          "decelerated from 44.3% in Q1. Revenue acceleration is drawing down the "
                          "deferred balance rather than being fed by new bookings."),
    ("Trap 3 — the FCF headline", "Reported Q2 Free Cash Flow of $53.2M excludes $45.5M of cash paid to "
                                  "settle RSU tax withholding, which GAAP puts in financing. True Q2 cash "
                                  "generation is $7.7M. First-half reported FCF of $141.8M is NEGATIVE "
                                  "$19.8M after $161.6M of RSU tax settlement."),
    ("Trap 4 — SBC scale", "SBC was $147.6M in Q2 = 39.9% of revenue. The entire gap between the "
                           "$(117.3)M GAAP operating loss and $36.1M non-GAAP operating income is "
                           "compensation. It IS decaying (2025Q3 $1,138M / 415% of revenue at the IPO "
                           "catch-up, then 71.9%, 50.7%, 39.9%) but from an extreme base."),
    ("", ""),
    ("Calibration 1 — guidance beat", f"Figma has beaten its own revenue guidance all four quarters as a public "
                                      f"company, and the beat is WIDENING: 3.9%, 3.7%, 5.5%, 6.0%. "
                                      f"Four-quarter mean {BEAT_AVG:.2%}; last two {BEAT_L2:.2%}. "
                                      f"The base case applies the last-two mean to the Q3 guide."),
    ("Calibration 2 — non-GAAP bridge", "non-GAAP operating income = GAAP operating income + SBC + employer "
                                        "payroll taxes on stock + acquired-intangible amortisation "
                                        "(+ impairments). Q2 check: -117.289 + 147.554 + 0.308 + 3.486 "
                                        "+ 2.034 = 36.093, matching the reported figure exactly."),
    ("Calibration 3 — non-GAAP tax", "14.5%, the company's stated projected non-GAAP rate (cut from 25% after "
                                     "a valuation-allowance release at end-2025)."),
    ("", ""),
    ("Valuation method", "Two routes on FY2027E, averaged, then probability-weighted 25/50/25. "
                         "Route 1 is EV/Sales (the right anchor while the margin is immature); "
                         "route 2 is P/E on non-GAAP EPS. This deviates from the CRDO/NBIS template's "
                         "FY+1 / FY+2-discounted pair because Figma's FY2028 margin is too speculative "
                         "to carry a target."),
    ("Model output", f"Probability-weighted 12-month target ${PW_TARGET:.2f} versus ${PX_NOW:.2f} spot "
                     f"({PW_TARGET/PX_NOW-1:+.1%}) and a ${STREET_PT:.2f} street mean."),
    ("How to update", "When Q3 2026 actuals land: append the column to IS_Quarterly, add the Q3 row to "
                      "Guidance_vs_Actual and re-derive BEAT_L2, recompute billings from the new "
                      "deferred-revenue balance, and check whether the FY non-GAAP operating income "
                      "guide finally moves with the revenue guide. That last one is the thesis."),
]
for i, (k, v) in enumerate(notes, start=1):
    c = ws.cell(row=i, column=1, value=k)
    c.font = BOLD
    c.alignment = Alignment(vertical="top")
    c2 = ws.cell(row=i, column=2, value=v)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    if k.startswith("Trap"):
        c.fill = WARN
        c2.fill = WARN
    if k.startswith("Calibration"):
        c.fill = GOOD
        c2.fill = GOOD
    ws.row_dimensions[i].height = 46 if len(v) > 130 else 30

# ------------------------------------------------------------ IS_Quarterly --
ws = wb.create_sheet("IS_Quarterly")
header(ws, COLS)
ws.cell(row=2, column=1, value="Period ended").font = BOLD
for j, e in enumerate(ENDS, start=2):
    ws.cell(row=2, column=j, value=e).alignment = Alignment(horizontal="center")
r = 3
for label, vals in IS:
    c = ws.cell(row=r, column=1, value=label)
    pct = "margin" in label or "%" in label
    if label.isupper() or "Non-GAAP" in label or "GAAP net" in label:
        c.font = BOLD
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = THIN
        if v is None:
            continue
        if pct:
            cell.number_format = PCT
        elif "EPS" in label:
            cell.number_format = '$0.00'
        else:
            cell.number_format = NUMFMT
    r += 1
ws.cell(row=r + 1, column=1, value="Values in USD. Millions display via number format. 2025Q3/Q4 sales-and-marketing "
                                   "and G&A are not separately split in XBRL for those quarters; total opex is exact.").font = Font(italic=True, size=9, color="7F7F7F")

# --------------------------------------------------------------- BS_CF ------
ws = wb.create_sheet("BS_CF")
header(ws, ["2026-06-30", "2025-12-31"], width0=48, width=18)
r = 2
ws.cell(row=r, column=1, value="BALANCE SHEET").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, a, b_ in BS:
    c = ws.cell(row=r, column=1, value=label)
    if label.isupper():
        c.font = BOLD
    for j, v in enumerate((a, b_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
        if label.isupper():
            cell.font = BOLD
            cell.fill = SEC
    r += 1
r += 2
ws.cell(row=r, column=1, value="CASH FLOW — and the correction the headline omits").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for j, lab in enumerate(["2026Q2", "2026Q1", "2025Q2"], start=2):
    c = ws.cell(row=r, column=j, value=lab)
    c.fill = SUB
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, a, b_, c_ in CF:
    cc = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        cc.font = BOLD
    for j, v in enumerate((a, b_, c_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if "margin" in label else NUMFMT
        cell.border = THIN
        if hl:
            cell.font = BOLD
            cell.fill = WARN if "AFTER" in label else SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="RSU tax settlement is a real, recurring cash cost of compensation that GAAP "
                               "classifies in financing rather than operating. Excluding it flatters the "
                               "Free Cash Flow headline by $45.5M in Q2 and $161.6M in the first half.").font = Font(italic=True, size=9, color="7F7F7F")

# ----------------------------------------------------------------- KPI ------
ws = wb.create_sheet("KPI")
header(ws, COLS)
r = 2
for label, *vals in KPI:
    c = ws.cell(row=r, column=1, value=label)
    if "Billings" in label:
        c.font = BOLD
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = THIN
        if v is None:
            continue
        if "growth" in label or "Retention" in label:
            cell.number_format = PCT
        elif "customers" in label:
            cell.number_format = '#,##0'
        else:
            cell.number_format = NUMFMT
        if "Billings y/y" in label:
            cell.fill = WARN
            cell.font = BOLD
    r += 1
r += 1
ws.cell(row=r, column=1, value="Billings y/y decelerated from 44.3% to 33.7% while revenue y/y ACCELERATED from "
                               "46.1% to 48.2%. That divergence is the single most important number on this "
                               "sheet and is not mentioned in the press release.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Guidance_vs_Actual -------
ws = wb.create_sheet("Guidance_vs_Actual")
header(ws, ["Guide low", "Guide high", "Midpoint", "Actual", "Beat vs midpoint", "Guided on"], width0=18, width=17)
r = 2
for q, lo, hi, mid, act, given in GUIDE_HIST:
    ws.cell(row=r, column=1, value=q).font = BOLD
    for j, v in enumerate((lo, hi, mid, act), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
    if act:
        cell = ws.cell(row=r, column=6, value=act / mid - 1)
        cell.number_format = PCTS
        cell.font = BOLD
        cell.fill = GOOD
    else:
        cell = ws.cell(row=r, column=6, value="not yet reported")
        cell.fill = SEC
    ws.cell(row=r, column=7, value=given)
    r += 1
r += 1
ws.cell(row=r, column=1, value="Four-quarter mean beat").font = BOLD
ws.cell(row=r, column=4, value=BEAT_AVG).number_format = PCTS
r += 1
ws.cell(row=r, column=1, value="Last-two mean beat (used in Base case)").font = BOLD
c = ws.cell(row=r, column=4, value=BEAT_L2)
c.number_format = PCTS
c.fill = GOOD
c.font = BOLD
r += 3
ws.cell(row=r, column=1, value="FULL-YEAR GUIDANCE WALK — revenue keeps rising, operating income stopped").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for j, lab in enumerate(["Revenue guide", "non-GAAP op income", "Given on"], start=2):
    c = ws.cell(row=r, column=j, value=lab)
    c.fill = SUB
    c.font = WHITE
r += 1
for label, rev, op, given in FY_WALK:
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=2, value=rev).number_format = NUMFMT
    if op:
        ws.cell(row=r, column=3, value=op).number_format = NUMFMT
    ws.cell(row=r, column=4, value=given)
    if label == "FY2026 @ Q2 2026":
        for j in (2, 3):
            ws.cell(row=r, column=j).fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="Feb to May: +$55M revenue guide, +$25M op income guide (45% incremental). "
                               "May to Aug: +$40M revenue guide, +$0M op income guide (0% incremental). "
                               "That is what the stock repriced.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Consensus_Market ---------
ws = wb.create_sheet("Consensus_Market")
ws.column_dimensions["A"].width = 48
ws.column_dimensions["B"].width = 74
ws.column_dimensions["C"].width = 14
header(ws, ["Value", "Source"], width0=48, width=74)
ws.column_dimensions["C"].width = 14
r = 2
for k, v, src in STREET:
    ws.cell(row=r, column=1, value=k).font = BOLD
    ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True)
    ws.cell(row=r, column=3, value=src)
    r += 1
r += 1
ws.cell(row=r, column=1, value="Street and market figures only. Nothing on this sheet feeds the reported "
                               "columns in IS_Quarterly — same discipline as the NBIS and MU workbooks.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Forecast_Q3_2026 ---------
ws = wb.create_sheet("Forecast_Q3_2026")
ws.column_dimensions["A"].width = 48
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value=f"Q3 2026 guide midpoint ${Q3_GUIDE_MID/M:,.0f}M implies just "
                               f"{Q3_GUIDE_MID/Q2_REV-1:+.1%} q/q after eight quarters averaging near +10%. "
                               f"The scenarios apply Figma's own history of beating its guide.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k)
    c.fill = HDR
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("Beat applied to guide midpoint", lambda k: SCEN_Q3[k]["beat"], PCTS, False),
        ("Q3 2026E revenue", lambda k: Q3[k]["revenue"], NUMFMT, True),
        ("  vs guide midpoint", lambda k: Q3[k]["vs_guide"], PCTS, False),
        ("  q/q growth", lambda k: Q3[k]["qoq"], PCTS, False),
        ("  y/y growth (guide implies 36.4%)", lambda k: Q3[k]["yoy"], PCTS, True),
        ("Non-GAAP operating margin", lambda k: Q3[k]["ng_margin"], PCT, False),
        ("Non-GAAP operating income", lambda k: Q3[k]["ng_op"], NUMFMT, False),
        ("Q4 2026E revenue", lambda k: FY26[k]["q4"], NUMFMT, False),
        ("FY2026E revenue", lambda k: FY26[k]["fy"], NUMFMT, True),
        ("  vs FY guide $1.465B", lambda k: FY26[k]["vs_guide"], PCTS, False),
        ("  FY2026E growth", lambda k: FY26[k]["growth"], PCTS, True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k))
        c.number_format = fmt
        c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=SCEN_Q3[k]["note"])
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 76

# ---------------------------------------------------------- Valuation -------
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 50
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value=f"Target = average of (EV/Sales on FY2027E revenue, plus ${NET_CASH/B:.2f}B net cash) "
                               f"and (FY2027E non-GAAP EPS x P/E). Non-GAAP net income = "
                               f"(operating income + ${INTEREST/M:.0f}M interest) x (1 - {TAX_NG:.1%}). "
                               f"Figma has no debt.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k)
    c.fill = HDR
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("FY2026E revenue", lambda k: FY26[k]["fy"], NUMFMT, False),
        ("FY2027E revenue growth", lambda k: FY[k]["g27"], PCT, False),
        ("FY2027E revenue", lambda k: VAL[k]["rev27"], NUMFMT, True),
        ("FY2027E non-GAAP operating margin", lambda k: FY[k]["ngm27"], PCT, False),
        ("FY2027E non-GAAP operating income", lambda k: VAL[k]["ng_op"], NUMFMT, False),
        ("FY2027E non-GAAP net income", lambda k: VAL[k]["ng_net"], NUMFMT, False),
        ("FY2027E diluted shares", lambda k: FY[k]["sh"], NUMFMT, False),
        ("FY2027E non-GAAP EPS", lambda k: VAL[k]["eps27"], '$0.00', True),
        ("EV/Sales multiple applied", lambda k: FY[k]["evs"], '0.0"x"', False),
        ("P/E multiple applied", lambda k: FY[k]["pe"], '0.0"x"', False),
        ("Route 1 — EV/Sales", lambda k: VAL[k]["px_evs"], USD, False),
        ("Route 2 — P/E on non-GAAP EPS", lambda k: VAL[k]["px_pe"], USD, False),
        ("12-MONTH TARGET (average)", lambda k: VAL[k]["target"], USD, True),
        (f"Upside vs ${PX_NOW:.2f}", lambda k: VAL[k]["upside"], PCTS, True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k))
        c.number_format = fmt
        c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k]["note"])
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 76
r += 2
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    ws.cell(row=r, column=j, value=PROB[k]).number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET)
c.number_format = USD
c.font = Font(bold=True, size=12)
c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET / PX_NOW - 1)
c.number_format = PCTS
c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
for label, val, fmt in [
        (f"Share price (2026-08-07 close)", PX_NOW, USD),
        ("Diluted shares (non-GAAP, Q2 2026)", SH_DIL, NUMFMT),
        ("Market capitalisation", MCAP, NUMFMT),
        ("Net cash and investments", NET_CASH, NUMFMT),
        ("Enterprise value", EV, NUMFMT),
        ("EV / FY2026 guide revenue ($1.465B)", EV / FY26_GUIDE, '0.0"x"'),
        ("EV / FY2026E revenue (base case)", EV / FY26["Base"]["fy"], '0.0"x"'),
        ("EV / FY2027E revenue (base case)", EV / VAL["Base"]["rev27"], '0.0"x"'),
        ("P/E on FY2027E base non-GAAP EPS", PX_NOW / VAL["Base"]["eps27"], '0.0"x"'),
        ("Net cash as % of market cap", NET_CASH / MCAP, PCT),
        (f"Street mean target ${STREET_PT:.2f} implies EV/FY2027E of",
         (STREET_PT * FY["Base"]["sh"] - NET_CASH) / VAL["Base"]["rev27"], '0.0"x"'),
        (f"Pre-earnings price ${PX_PRE:.2f} implied EV/FY2027E of",
         (PX_PRE * FY["Base"]["sh"] - NET_CASH) / VAL["Base"]["rev27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027E revenue x EV/Sales (per share, net cash added back)").font = BOLD
r += 1
mults = [4.0, 5.5, 7.0, 8.5, 10.0]
ws.cell(row=r, column=1, value="FY27 revenue \\ EV/S").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x")
    c.fill = SUB
    c.font = WHITE
r += 1
for rev_ in (1.85 * B, 1.95 * B, 2.05 * B, 2.15 * B, 2.25 * B):
    c = ws.cell(row=r, column=1, value=rev_)
    c.number_format = NUMFMT
    c.fill = SUB
    c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        px = (m_ * rev_ + NET_CASH) / 555e6
        c = ws.cell(row=r, column=j, value=px)
        c.number_format = USD
        if abs(px - PX_NOW) < 2.0:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/FIG_Financial_Model.xlsx"
wb.save(out)
print(f"wrote {out}\n")

print(f"{'':22}{'Bear':>12}{'Base':>12}{'Bull':>12}")
print(f"{'Q3 2026E revenue':22}" + "".join(f"{Q3[k]['revenue']/M:>12,.0f}" for k in ("Bear", "Base", "Bull")))
print(f"{'  y/y':22}" + "".join(f"{Q3[k]['yoy']:>11.1%} " for k in ("Bear", "Base", "Bull")))
print(f"{'  vs guide':22}" + "".join(f"{Q3[k]['vs_guide']:>11.1%} " for k in ("Bear", "Base", "Bull")))
print(f"{'FY2026E revenue':22}" + "".join(f"{FY26[k]['fy']/M:>12,.0f}" for k in ("Bear", "Base", "Bull")))
print(f"{'FY2027E revenue':22}" + "".join(f"{VAL[k]['rev27']/M:>12,.0f}" for k in ("Bear", "Base", "Bull")))
print(f"{'FY2027E nonGAAP EPS':22}" + "".join(f"{VAL[k]['eps27']:>12.2f}" for k in ("Bear", "Base", "Bull")))
print(f"{'Route 1 EV/Sales':22}" + "".join(f"{VAL[k]['px_evs']:>12.2f}" for k in ("Bear", "Base", "Bull")))
print(f"{'Route 2 P/E':22}" + "".join(f"{VAL[k]['px_pe']:>12.2f}" for k in ("Bear", "Base", "Bull")))
print(f"{'TARGET':22}" + "".join(f"{VAL[k]['target']:>12.2f}" for k in ("Bear", "Base", "Bull")))
print(f"{'upside':22}" + "".join(f"{VAL[k]['upside']:>11.1%} " for k in ("Bear", "Base", "Bull")))
print(f"\nProbability-weighted target: ${PW_TARGET:.2f}  ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW:.2f})")
print(f"Street mean target:          ${STREET_PT:.2f}  ({STREET_PT/PX_NOW-1:+.1%})")
print(f"Market cap ${MCAP/B:.2f}B | net cash ${NET_CASH/B:.2f}B | EV ${EV/B:.2f}B "
      f"| EV/FY26E {EV/FY26['Base']['fy']:.1f}x | EV/FY27E {EV/VAL['Base']['rev27']:.1f}x")
print(f"Guidance beat history: " + ", ".join(f"{b:.2%}" for b in BEATS) + f" | last-2 mean {BEAT_L2:.2%}")
