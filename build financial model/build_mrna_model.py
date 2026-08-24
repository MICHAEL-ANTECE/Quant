#!/usr/bin/env python3
"""
MRNA (Moderna, Inc.) model — same framework as the MU / NBIS / SHOP / APP /
SNDK / AXTI / AAOI / CRDO / FIG / COHR workbooks.

DATA SOURCE: SEC EDGAR. Q2 2026 8-K EX-99.1 filed 2026-07-31 (full three
statements plus the 2026 financial framework) and XBRL companyfacts for the
quarterly history. Street and market data are on their own sheet.

FISCAL CALENDAR: calendar year. Q2 2026 = quarter ended 2026-06-30, reported
2026-07-31 before the open.

WHY THIS WORKBOOK IS BUILT DIFFERENTLY FROM THE OTHERS: on 2026-08-19 Merck and
Moderna announced that the Phase 3 INTerpath-001 trial of intismeran autogene
(mRNA-4157) plus KEYTRUDA met BOTH endpoints -- recurrence-free survival and
distant metastasis-free survival -- in completely resected Stage IIB-IV
melanoma, n=1,137. That is the first successful Phase 3 for any mRNA cancer
therapy, ever. The stock went from $62.96 to an intraday $163.47, the largest
one-day gain in company history (the prior record was 27.81% on 2020-02-25).

So the income statement no longer sets the share price. What the financials
still decide is how long Moderna can fund the oncology programme without
issuing equity -- the cash-runway sheet -- and that is what this model is for.
The Valuation sheet deliberately leads with a REVERSE calculation: what peak
intismeran revenue does today's price require? A forward point target on a
company that just re-rated 117% on a binary readout would be false precision,
and the workbook says so rather than pretending otherwise.

THE QUARTER ITSELF, in one line: revenue $145M versus $142M, but PRODUCT sales
fell 17.5% ($114M to $94M) and the whole increase came from grants, collaboration
and stand-ready manufacturing revenue ($28M to $51M).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

M, B = 1e6, 1e9

COLS = ["2026Q2", "2026Q1", "2025Q4", "2025Q3", "2025Q2", "2025Q1"]
ENDS = ["2026-06-30", "2026-03-31", "2025-12-31", "2025-09-30", "2025-06-30", "2025-03-31"]
n = len(COLS)

PX_NOW = 136.50        # 2026-08-19 intraday, moomoo (session still open)
PX_PREV = 62.96        # 2026-08-18 close, the day before the readout
PX_HIGH = 163.47       # 2026-08-19 intraday high
PX_52L = 22.28
SH = 398 * M           # Q2 2026 weighted average basic/diluted shares


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"{len(v)} > {n}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Net product sales", row(94*M, 352*M, None, None, 114*M, 86*M)),
    ("Other revenue (grant/collab/stand-ready)", row(51*M, 37*M, None, None, 28*M, 22*M)),
    ("Total revenue", row(145*M, 389*M, 678*M, 1016*M, 142*M, 108*M)),
    ("  revenue y/y", row(0.021, 2.602, -0.298, -0.454, -0.411, -0.353)),
    ("Cost of sales", row(93*M, 955*M, 452*M, 207*M, 119*M, 90*M)),
    ("  of which litigation settlement", row(0.0, 950*M, None, None, None, None)),
    ("Research and development", row(651*M, 649*M, 775*M, 801*M, 700*M, 856*M)),
    ("Selling, general and administrative", row(216*M, 173*M, 308*M, 268*M, 230*M, 212*M)),
    ("Total operating expenses", row(960*M, 1777*M, 1535*M, 1276*M, 1049*M, 1158*M)),
    ("Loss from operations", row(-815*M, -1388*M, -857*M, -260*M, -907*M, -1050*M)),
    ("Interest income", row(67*M, 72*M, None, None, 81*M, 90*M)),
    ("Other (expense) income, net", row(-19*M, -18*M, None, None, 8*M, -4*M)),
    ("Loss before income taxes", row(-767*M, -1334*M, None, None, -818*M, -964*M)),
    ("Provision for income taxes", row(15*M, 9*M, None, None, 7*M, 7*M)),
    ("Net loss", row(-782*M, -1343*M, -826*M, -200*M, -825*M, -971*M)),
    ("Diluted EPS", row(-1.97, -3.40, -2.11, -0.51, -2.13, -2.52)),
    ("Weighted average shares", row(398*M, 395*M, None, None, 388*M, 385*M)),
]

# ============================================== BALANCE SHEET / CASH FLOW ====
BS = [
    ("Cash and cash equivalents", 1723*M, 2595*M),
    ("Investments, current", 3415*M, 3204*M),
    ("Investments, non-current", 1772*M, 2336*M),
    ("TOTAL CASH AND INVESTMENTS", 6910*M, 8135*M),
    ("Accounts receivable, net", 17*M, 184*M),
    ("Inventory", 280*M, 153*M),
    ("Property, plant and equipment, net", 2059*M, 2134*M),
    ("Total assets", 10961*M, 12338*M),
    ("Accounts payable", 167*M, 317*M),
    ("ACCRUED LIABILITIES", 1909*M, 1386*M),
    ("Deferred revenue (current + non-current)", 442*M, 252*M),
    ("Long-term debt", 591*M, 590*M),
    ("Total liabilities", 4200*M, 3688*M),
    ("Retained earnings", 5098*M, 7223*M),
    ("Total stockholders' equity", 6761*M, 8650*M),
]

CF = [
    ("Net loss", -2125*M, -1796*M),
    ("Stock-based compensation", 224*M, 245*M),
    ("Depreciation and amortization", 122*M, 96*M),
    ("Change in accounts receivable", 167*M, 310*M),
    ("Change in inventory", -126*M, -122*M),
    ("Change in accounts payable", -128*M, -203*M),
    ("CHANGE IN ACCRUED LIABILITIES", 475*M, -395*M),
    ("Change in deferred revenue", 195*M, 68*M),
    ("NET CASH USED IN OPERATING ACTIVITIES", -1156*M, -1956*M),
    ("Purchases of property, plant and equipment", -99*M, -120*M),
    ("FREE CASH FLOW", -1255*M, -2076*M),
]

# =================================================== CASH RUNWAY =============
CASH_JUN = 6910 * M
SETTLE_PAID = 950 * M            # paid in cash July 2026, i.e. lands in Q3
DEBT = 591 * M
CASH_PRO = CASH_JUN - SETTLE_PAID
NET_CASH = CASH_PRO - DEBT
YE_GUIDE_LO, YE_GUIDE_HI = 4.7 * B, 5.2 * B
CREDIT_UNDRAWN = 0.9 * B

RUNWAY = [
    ("Cash and investments at 2026-06-30", CASH_JUN, "per press release; XBRL current-only figure is $5,138M"),
    ("Less: litigation settlement paid July 2026", -SETTLE_PAID, "accrued in Q1 P&L, cash left the door in Q3"),
    ("Pro-forma cash and investments", CASH_PRO, "the real starting point for the second half"),
    ("Less: long-term debt", -DEBT, ""),
    ("PRO-FORMA NET CASH", NET_CASH, ""),
    ("2026 year-end cash guidance (low)", YE_GUIDE_LO, "raised by ~$0.2B this quarter"),
    ("2026 year-end cash guidance (high)", YE_GUIDE_HI, ""),
    ("Implied second-half burn (low end of guide)", YE_GUIDE_LO - CASH_PRO, "and this is the half with the big vaccine quarter"),
    ("Implied second-half burn (high end of guide)", YE_GUIDE_HI - CASH_PRO, ""),
    ("Undrawn credit facility", CREDIT_UNDRAWN, "excluded from the year-end cash guidance"),
]

# ============================================ 2026 FRAMEWORK AND WHAT IT IMPLIES
FY25_REV = (108 + 142 + 1016 + 678) * M      # $1,944M
H1_26 = 534 * M
H2_25 = (1016 + 678) * M                      # $1,694M

FRAMEWORK = [
    ("2026 revenue target", "up to 10% growth on 2025", FY25_REV * 1.10),
    ("2026 revenue if flat on 2025", "the other end of 'up to'", FY25_REV),
    ("Cost of sales", "~$1.7B, lowered from ~$1.8B, INCLUDING the $0.9B settlement", 1.7 * B),
    ("Research and development", "~$2.9B, lowered from ~$3.0B", 2.9 * B),
    ("Selling, general and administrative", "~$1.0B", 1.0 * B),
    ("Capital expenditures", "$0.2B - $0.3B", 0.25 * B),
    ("Year-end cash and investments", "$4.7B - $5.2B, improved ~$0.2B", 4.95 * B),
]

H2_IMPLIED = [
    ("H1 2026 actual revenue", H1_26),
    ("H2 2026 implied at the 10% ceiling", FY25_REV * 1.10 - H1_26),
    ("H2 2026 implied if full year is flat", FY25_REV - H1_26),
    ("H2 2025 actual revenue", H2_25),
    ("H2 y/y at the ceiling", (FY25_REV * 1.10 - H1_26) / H2_25 - 1),
    ("H2 y/y if flat", (FY25_REV - H1_26) / H2_25 - 1),
    ("Q3 2026 implied (55% of H2, ceiling)", 0.55 * (FY25_REV * 1.10 - H1_26)),
    ("Q3 2025 actual", 1016 * M),
    ("Q3 y/y at the ceiling", 0.55 * (FY25_REV * 1.10 - H1_26) / (1016 * M) - 1),
]

# Full-year operating loss implied by the framework, at the revenue ceiling
FY26_REV_CEIL = FY25_REV * 1.10
FY26_OPEX = 1.7 * B + 2.9 * B + 1.0 * B
FY26_OP_LOSS = FY26_REV_CEIL - FY26_OPEX
FY26_OP_LOSS_EXSETTLE = FY26_OP_LOSS + 0.9 * B

# ==================================================== CONSENSUS / MARKET =====
STREET = [
    ("Q2 2026 consensus revenue", "$126.6M (Zacks) — actual $145M, a beat", "street"),
    ("Q2 2026 consensus EPS", "$(1.97) — actual $(1.97), in line", "street"),
    ("Reaction to the Q2 print (2026-07-31)", "stock FELL about 6-7% on the day despite the revenue beat", "market"),
    ("2026-08-05", "FDA approved mFLUSIVA (mRNA-1010), first mRNA flu vaccine, adults 50+", "regulatory"),
    ("2026-08-19 — THE EVENT", "Phase 3 INTerpath-001 met both RFS and DMFS endpoints in resected "
                               "Stage IIB-IV melanoma (n=1,137); first successful Phase 3 for any "
                               "mRNA cancer therapy", "clinical"),
    ("Price 2026-08-18 close", f"${PX_PREV:.2f}", "market"),
    ("Price 2026-08-19 intraday", f"${PX_NOW:.2f}, high ${PX_HIGH:.2f} — biggest one-day gain in company history", "market"),
    ("52-week low", f"${PX_52L:.2f}", "market"),
    ("Volume 2026-08-19", "100.7M shares versus a typical 4-6M", "market"),
    ("Goldman Sachs price target", "raised to $67 from $49", "street"),
    ("Morgan Stanley price target", "raised to $209 from $170, Equal Weight", "street"),
    ("TARGET DISPERSION WARNING", "$67 versus $209 with the stock at $136.50 — the published targets "
                                  "are a mix of pre- and post-readout vintages and cannot be reconciled "
                                  "into a usable consensus. Treat the street as having no settled view.", "judgement"),
    ("Barclays intismeran estimate", "about $3B in melanoma by 2035 (gross programme, before the Merck split)", "street"),
    ("Negative in the same release", "norovirus mRNA-1403 Phase 3 MISSED statistical criteria for early "
                                     "success at interim; an additional cohort is being enrolled", "clinical"),
    ("Next scheduled catalyst", "Analyst Day, November 12, 2026", "company"),
]

# ======================================== VALUATION — REVERSE FIRST ==========
# Moderna and Merck share intismeran economics 50/50, so a gross programme peak
# of $X supports roughly $X/2 of Moderna revenue.
MCAP = SH * PX_NOW
EV = MCAP - NET_CASH

VACCINE_VALUE = {"Bear": 0.0, "Base": 3.2 * B, "Bull": 5.3 * B}
ONC = {
    "Bear": dict(peak_gross=3.0*B, mult=4.0, pos=0.75,
                 note="Melanoma only, roughly the Barclays number. Adjuvant uptake is slow, "
                      "personalised manufacturing caps throughput, and the other INTerpath "
                      "tumours do not replicate the result."),
    "Base": dict(peak_gross=7.0*B, mult=5.0, pos=0.85,
                 note="Melanoma plus one further adjuvant indication converting from the nine "
                      "ongoing Phase 2/3 trials. Regulatory filings proceed on the strength of "
                      "both RFS and DMFS."),
    "Bull": dict(peak_gross=14.0*B, mult=6.0, pos=0.90,
                 note="The platform generalises across adjuvant solid tumours — NSCLC, bladder, "
                      "renal cell — and intismeran becomes standard of care alongside "
                      "checkpoint inhibition."),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}
DISC, YEARS = 0.11, 6.0   # peak around 2035, discounted from a mid-ramp ~2032


def onc_value(k):
    p = ONC[k]
    mrna_peak = p["peak_gross"] / 2.0          # 50/50 with Merck
    return mrna_peak * p["mult"] * p["pos"] / (1 + DISC) ** YEARS


def equity_value(k):
    onc = onc_value(k)
    tot = NET_CASH + VACCINE_VALUE[k] + onc
    return dict(onc=onc, vaccine=VACCINE_VALUE[k], total=tot, ps=tot / SH,
                upside=tot / SH / PX_NOW - 1, mrna_peak=ONC[k]["peak_gross"] / 2.0)


VAL = {k: equity_value(k) for k in ONC}
PW = sum(PROB[k] * VAL[k]["ps"] for k in PROB)

# Reverse: what MRNA-share peak revenue does today's price require, holding the
# Base multiple, probability of success and discount rate constant?
IMPLIED_ONC = EV - VACCINE_VALUE["Base"]
IMPLIED_PEAK_MRNA = IMPLIED_ONC * (1 + DISC) ** YEARS / (ONC["Base"]["mult"] * ONC["Base"]["pos"])
IMPLIED_PEAK_GROSS = IMPLIED_PEAK_MRNA * 2.0

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
GOOD = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'
PCTS = '+0.0%;-0.0%'
USD = '$#,##0.00'

wb = Workbook()
wb.remove(wb.active)


def header(ws, labels, r=1, width0=46, width=15):
    ws.column_dimensions["A"].width = width0
    for j, lab in enumerate(labels, start=2):
        ws.column_dimensions[chr(64 + j)].width = width
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).fill = HDR


# ---------------------------------------------------------------- README ----
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 108
notes = [
    ("Company", "Moderna, Inc. (NASDAQ: MRNA) — mRNA vaccines and therapeutics."),
    ("Model built", "2026-08-19, after Q2 2026 results (2026-07-31) and the INTerpath-001 readout (2026-08-19)."),
    ("Data source", "SEC EDGAR: 8-K EX-99.1 filed 2026-07-31 plus XBRL companyfacts. Clinical and "
                    "market items are on the Consensus_Market sheet and never touch reported columns."),
    ("Latest reported quarter", "2026Q2: revenue $145M, GAAP net loss $(782)M, EPS $(1.97)."),
    ("", ""),
    ("WHY THIS MODEL IS SHAPED DIFFERENTLY", "On 2026-08-19 the Phase 3 INTerpath-001 trial of intismeran "
                                             "autogene plus KEYTRUDA met BOTH recurrence-free and distant "
                                             "metastasis-free survival endpoints in resected Stage IIB-IV "
                                             "melanoma (n=1,137) — the first successful Phase 3 for any mRNA "
                                             "cancer therapy. The stock went from $62.96 to an intraday "
                                             "$163.47. The income statement no longer sets the price. What "
                                             "the financials still decide is the cash runway that funds the "
                                             "oncology programme — see Cash_Runway."),
    ("Trap 1 — the revenue line", "Total revenue $145M versus $142M looks like growth. PRODUCT sales fell "
                                  "17.5%, from $114M to $94M. The entire increase came from grants, "
                                  "collaboration, licensing and stand-ready manufacturing revenue, $28M to "
                                  "$51M. The commercial base is still shrinking."),
    ("Trap 2 — 'up to 10% growth'", "2025 revenue was $1,944M, so the ceiling is $2,138M. H1 2026 delivered "
                                    "$534M against $250M, up 114%. That forces H2 2026 to be $1,410M-$1,604M "
                                    "against $1,694M in H2 2025, i.e. DOWN 5% to 17% y/y. The headline growth "
                                    "number is entirely an H1 comparison artefact."),
    ("Trap 3 — the settlement timing", "The $950M litigation settlement was charged to Q1 2026 cost of sales "
                                       "but the cash left in July 2026. H1 operating cash flow of $(1,156)M "
                                       "was flattered by a $475M build in accrued liabilities. True H1 "
                                       "operating burn was closer to $(1,631)M, and Q3 absorbs the payment."),
    ("Trap 4 — interest income decays", "Interest income fell from $81M to $67M y/y purely because the cash "
                                        "pile is shrinking. It funded 8% of the quarter's operating loss. "
                                        "Every quarter of burn makes the next quarter's loss slightly worse."),
    ("Trap 5 — a miss in the same release", "The norovirus Phase 3 (mRNA-1403) did NOT meet statistical "
                                            "criteria for early success at interim. It was overshadowed, "
                                            "but it is a real pipeline setback."),
    ("", ""),
    ("Calibration 1 — the cost programme", "R&D $1,300M in H1 2026 against $1,556M, down 16%. SG&A $389M "
                                           "against $442M, down 12%. Full-year R&D guidance cut to ~$2.9B "
                                           "from ~$3.0B and cost of sales to ~$1.7B from ~$1.8B. The cuts "
                                           "are real and are landing roughly on schedule."),
    ("Calibration 2 — burn versus guide", f"Pro-forma cash after the July settlement is ${CASH_PRO/B:.2f}B. "
                                          f"Year-end guidance of $4.7-5.2B implies a second-half burn of "
                                          f"${abs(YE_GUIDE_HI-CASH_PRO)/B:.2f}B to ${abs(YE_GUIDE_LO-CASH_PRO)/B:.2f}B "
                                          f"— in the half that contains the big vaccine quarter."),
    ("Calibration 3 — framework loss", f"At the revenue ceiling of ${FY26_REV_CEIL/B:.2f}B against guided "
                                       f"opex of ${FY26_OPEX/B:.2f}B, the 2026 operating loss is "
                                       f"${abs(FY26_OP_LOSS)/B:.2f}B, or ${abs(FY26_OP_LOSS_EXSETTLE)/B:.2f}B "
                                       f"excluding the settlement."),
    ("", ""),
    ("Valuation method — READ THIS", "The Valuation sheet leads with a REVERSE calculation, not a target. "
                                     "Putting a point price target on a company that re-rated 117% in one "
                                     "session on a binary readout would be false precision. The useful "
                                     "question is what today's price requires intismeran to become, and "
                                     "that is answered explicitly. Scenario values follow, but their "
                                     "multiples and probabilities are judgement, not observation."),
    ("Reverse answer", f"At ${PX_NOW:.2f} the enterprise value is ${EV/B:.1f}B. Net of a ${VACCINE_VALUE['Base']/B:.1f}B "
                       f"vaccine business, the market is paying ${IMPLIED_ONC/B:.1f}B for intismeran. Holding the "
                       f"base-case multiple, probability of success and discount rate, that implies Moderna-share "
                       f"peak revenue of ${IMPLIED_PEAK_MRNA/B:.1f}B, i.e. a gross programme peak near "
                       f"${IMPLIED_PEAK_GROSS/B:.1f}B — against Barclays' ~$3B melanoma-only estimate."),
    ("Street dispersion", "Goldman Sachs $67, Morgan Stanley $209, stock at $136.50. Published targets mix "
                          "pre- and post-readout vintages and do not reconcile. There is no usable consensus "
                          "on this name today."),
    ("How to update", "Watch three things: (1) Q3 2026 cash — does the year-end $4.7-5.2B guide hold after the "
                      "$950M payment; (2) the melanoma filing timetable and any disclosed pricing; (3) whether "
                      "any second INTerpath tumour reads out. The oncology scenario collapses to Bear if "
                      "melanoma stays the only indication."),
]
for i, (k, v) in enumerate(notes, start=1):
    c = ws.cell(row=i, column=1, value=k)
    c.font = BOLD
    c.alignment = Alignment(vertical="top")
    c2 = ws.cell(row=i, column=2, value=v)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    if k.startswith("Trap"):
        c.fill = WARN
        c2.fill = WARN
    if k.startswith("Calibration"):
        c.fill = GOOD
        c2.fill = GOOD
    if "READ THIS" in k or k == "Reverse answer" or "DIFFERENTLY" in k:
        c.fill = SEC
        c2.fill = SEC
    ws.row_dimensions[i].height = 58 if len(v) > 200 else (44 if len(v) > 120 else 30)

# ------------------------------------------------------------ IS_Quarterly --
ws = wb.create_sheet("IS_Quarterly")
header(ws, COLS)
ws.cell(row=2, column=1, value="Period ended").font = BOLD
for j, e in enumerate(ENDS, start=2):
    ws.cell(row=2, column=j, value=e).alignment = Alignment(horizontal="center")
r = 3
for label, vals in IS:
    c = ws.cell(row=r, column=1, value=label)
    if label.isupper() or "Total" in label or "Net loss" in label:
        c.font = BOLD
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = THIN
        if v is None:
            continue
        if "y/y" in label:
            cell.number_format = PCTS
        elif "EPS" in label:
            cell.number_format = '$0.00'
        else:
            cell.number_format = NUMFMT
        if label == "Net product sales":
            cell.fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="Q2 2026 product sales fell 17.5% y/y while total revenue rose 2.1%. The gap "
                               "is grant, collaboration and stand-ready manufacturing revenue, which nearly "
                               "doubled. Q1 2026 cost of sales carries the $950M litigation settlement, which "
                               "is why that quarter's gross margin is -145%.").font = Font(italic=True, size=9, color="7F7F7F")

# --------------------------------------------------------------- BS_CF ------
ws = wb.create_sheet("BS_CF")
header(ws, ["Jun 30, 2026", "Dec 31, 2025"], width0=48, width=18)
r = 2
ws.cell(row=r, column=1, value="BALANCE SHEET").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, a, b_ in BS:
    c = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        c.font = BOLD
    for j, v in enumerate((a, b_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
        if hl:
            cell.font = BOLD
            cell.fill = WARN if "ACCRUED" in label else SEC
    r += 1
r += 2
ws.cell(row=r, column=1, value="CASH FLOW — six months ended June 30").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for j, lab in enumerate(["2026", "2025"], start=2):
    c = ws.cell(row=r, column=j, value=lab)
    c.fill = SUB
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, a, b_ in CF:
    c = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        c.font = BOLD
    for j, v in enumerate((a, b_), start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = NUMFMT
        cell.border = THIN
        if hl:
            cell.font = BOLD
            cell.fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="The $475M build in accrued liabilities is the litigation settlement sitting "
                               "on the balance sheet at quarter end. It flatters H1 operating cash flow and "
                               "reverses in Q3 when the $950M was paid.").font = Font(italic=True, size=9, color="7F7F7F")

# ---------------------------------------------------------- Cash_Runway ----
ws = wb.create_sheet("Cash_Runway")
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 62
ws.cell(row=1, column=1, value="The only part of the financials that still matters to the equity: how long "
                               "Moderna can fund the oncology programme without issuing shares.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for label, val, note in RUNWAY:
    c = ws.cell(row=r, column=1, value=label)
    hl = label.isupper()
    if hl:
        c.font = BOLD
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = NUMFMT
    cell.border = THIN
    if hl:
        cell.font = BOLD
        cell.fill = SEC
    if "Implied second-half burn" in label:
        cell.fill = WARN
        cell.font = BOLD
    ws.cell(row=r, column=3, value=note).font = Font(italic=True, size=9, color="7F7F7F")
    r += 1
r += 2
ws.cell(row=r, column=1, value="2026 FINANCIAL FRAMEWORK (company guidance)").font = BOLD
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, note, val in FRAMEWORK:
    ws.cell(row=r, column=1, value=label).font = BOLD
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = NUMFMT
    cell.border = THIN
    ws.cell(row=r, column=3, value=note)
    r += 1
r += 1
c = ws.cell(row=r, column=1, value="Implied 2026 operating loss at the revenue ceiling")
c.font = BOLD
cell = ws.cell(row=r, column=2, value=FY26_OP_LOSS)
cell.number_format = NUMFMT
cell.fill = WARN
cell.font = BOLD
ws.cell(row=r, column=3, value="revenue ceiling less guided cost of sales, R&D and SG&A")
r += 1
c = ws.cell(row=r, column=1, value="  excluding the $0.9B settlement")
cell = ws.cell(row=r, column=2, value=FY26_OP_LOSS_EXSETTLE)
cell.number_format = NUMFMT
r += 3
ws.cell(row=r, column=1, value="WHAT 'UP TO 10% GROWTH' ACTUALLY IMPLIES FOR THE SECOND HALF").font = BOLD
ws.cell(row=r, column=1).fill = WARN
r += 1
for label, val in H2_IMPLIED:
    ws.cell(row=r, column=1, value=label).font = BOLD if "y/y" in label else Font()
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = PCTS if "y/y" in label else NUMFMT
    cell.border = THIN
    if "y/y" in label:
        cell.fill = WARN
        cell.font = BOLD
    r += 1
r += 1
ws.cell(row=r, column=1, value="H1 2026 revenue more than doubled. The full-year guide therefore requires the "
                               "second half to DECLINE year over year, and the third quarter — the vaccine "
                               "quarter — to fall well short of last year's $1,016M.").font = Font(italic=True, size=9, color="7F7F7F")

# ------------------------------------------------- Consensus_Market ---------
ws = wb.create_sheet("Consensus_Market")
header(ws, ["Value", "Source"], width0=44, width=80)
ws.column_dimensions["C"].width = 14
r = 2
for k, v, src in STREET:
    c = ws.cell(row=r, column=1, value=k)
    c.font = BOLD
    c2 = ws.cell(row=r, column=2, value=v)
    c2.alignment = Alignment(wrap_text=True)
    ws.cell(row=r, column=3, value=src)
    if "EVENT" in k or "WARNING" in k:
        c.fill = SEC
        c2.fill = SEC
    ws.row_dimensions[r].height = 32 if len(v) > 90 else 16
    r += 1

# ---------------------------------------------------------- Valuation -------
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 54
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="This sheet leads with the REVERSE question. A point target on a stock that "
                               "re-rated 117% in one session on a binary readout would be false precision.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="WHAT TODAY'S PRICE ALREADY REQUIRES").font = Font(bold=True, size=12)
ws.cell(row=r, column=1).fill = SEC
r += 1
for label, val, fmt in [
        (f"Share price (2026-08-19 intraday)", PX_NOW, USD),
        ("Shares outstanding", SH, NUMFMT),
        ("Market capitalisation", MCAP, NUMFMT),
        ("Pro-forma net cash (after July settlement, less debt)", NET_CASH, NUMFMT),
        ("Enterprise value", EV, NUMFMT),
        ("Less: assumed value of the vaccine business", VACCINE_VALUE["Base"], NUMFMT),
        ("IMPLIED VALUE OF INTISMERAN", IMPLIED_ONC, NUMFMT),
        ("Implied Moderna-share peak revenue", IMPLIED_PEAK_MRNA, NUMFMT),
        ("IMPLIED GROSS PROGRAMME PEAK (50/50 with Merck)", IMPLIED_PEAK_GROSS, NUMFMT),
        ("For reference: Barclays melanoma-only estimate", 3.0 * B, NUMFMT)]:
    c = ws.cell(row=r, column=1, value=label)
    if label.isupper():
        c.font = BOLD
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = fmt
    cell.border = THIN
    if label.isupper():
        cell.font = BOLD
        cell.fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value=f"Holding the base-case multiple ({ONC['Base']['mult']:.0f}x peak), probability "
                               f"of success ({ONC['Base']['pos']:.0%}) and discount rate ({DISC:.0%} over "
                               f"{YEARS:.0f} years) constant.").font = Font(italic=True, size=9, color="7F7F7F")
r += 3
ws.cell(row=r, column=1, value="SCENARIO ARITHMETIC (judgement, not observation)").font = Font(bold=True, size=12)
ws.cell(row=r, column=1).fill = SEC
r += 1
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k)
    c.fill = HDR
    c.font = WHITE
    c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("Gross programme peak revenue", lambda k: ONC[k]["peak_gross"], NUMFMT, False),
        ("Moderna share (50/50 with Merck)", lambda k: VAL[k]["mrna_peak"], NUMFMT, False),
        ("Multiple of peak revenue", lambda k: ONC[k]["mult"], '0.0"x"', False),
        ("Probability of success from here", lambda k: ONC[k]["pos"], PCT, False),
        ("Risk-adjusted, discounted oncology value", lambda k: VAL[k]["onc"], NUMFMT, True),
        ("Value of the vaccine business", lambda k: VAL[k]["vaccine"], NUMFMT, False),
        ("Pro-forma net cash", lambda k: NET_CASH, NUMFMT, False),
        ("TOTAL EQUITY VALUE", lambda k: VAL[k]["total"], NUMFMT, True),
        ("Per share", lambda k: VAL[k]["ps"], USD, True),
        (f"Versus ${PX_NOW:.2f}", lambda k: VAL[k]["upside"], PCTS, True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k))
        c.number_format = fmt
        c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=ONC[k]["note"])
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 92
r += 2
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    ws.cell(row=r, column=j, value=PROB[k]).number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED VALUE PER SHARE").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW)
c.number_format = USD
c.font = Font(bold=True, size=12)
c.fill = SEC
c = ws.cell(row=r, column=3, value=PW / PX_NOW - 1)
c.number_format = PCTS
c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Read this as a statement about how much the single session has already "
                               "priced, not as a target. The honest summary is that the readout is "
                               "genuinely transformative AND the stock moved further in one day than the "
                               "scenario arithmetic supports on melanoma alone.").font = Font(italic=True, size=9, color="7F7F7F")
r += 3
ws.cell(row=r, column=1, value="SENSITIVITY: gross programme peak x multiple, per share").font = BOLD
r += 1
mults = [3.0, 4.5, 6.0, 7.5, 9.0]
ws.cell(row=r, column=1, value="Gross peak \\ multiple").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x")
    c.fill = SUB
    c.font = WHITE
r += 1
for gross in (3.0 * B, 7.0 * B, 11.0 * B, 15.0 * B, 20.0 * B):
    c = ws.cell(row=r, column=1, value=gross)
    c.number_format = NUMFMT
    c.fill = SUB
    c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        ps = (NET_CASH + VACCINE_VALUE["Base"]
              + (gross / 2.0) * m_ * ONC["Base"]["pos"] / (1 + DISC) ** YEARS) / SH
        c = ws.cell(row=r, column=j, value=ps)
        c.number_format = USD
        if abs(ps - PX_NOW) < 12:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/MRNA_Financial_Model.xlsx"
wb.save(out)
print(f"wrote {out}\n")

ks = ("Bear", "Base", "Bull")
print("=== THE QUARTER ===")
print(f"Q2 2026 revenue $145M vs $142M (+2.1%), but PRODUCT sales $94M vs $114M ({94/114-1:+.1%})")
print(f"Net loss $(782)M, EPS $(1.97) — in line with consensus; revenue beat $126.6M consensus")
print(f"FY2025 revenue ${FY25_REV/B:.3f}B; 'up to 10% growth' ceiling ${FY26_REV_CEIL/B:.3f}B")
print(f"H1 2026 ${H1_26/B:.3f}B (+114% y/y) -> H2 implied ${(FY26_REV_CEIL-H1_26)/B:.3f}B "
      f"vs H2 2025 ${H2_25/B:.3f}B = {(FY26_REV_CEIL-H1_26)/H2_25-1:+.1%} y/y")
print(f"2026 implied operating loss ${FY26_OP_LOSS/B:.2f}B (${FY26_OP_LOSS_EXSETTLE/B:.2f}B ex-settlement)")
print("\n=== CASH RUNWAY ===")
print(f"Cash+investments 6/30: ${CASH_JUN/B:.2f}B | less July settlement ${SETTLE_PAID/B:.2f}B "
      f"= ${CASH_PRO/B:.2f}B | net of debt ${NET_CASH/B:.2f}B")
print(f"Year-end guide $4.7-5.2B -> H2 burn ${abs(YE_GUIDE_HI-CASH_PRO)/B:.2f}B to ${abs(YE_GUIDE_LO-CASH_PRO)/B:.2f}B")
print("\n=== WHAT $136.50 ALREADY REQUIRES ===")
print(f"Market cap ${MCAP/B:.1f}B | net cash ${NET_CASH/B:.2f}B | EV ${EV/B:.1f}B")
print(f"Implied value of intismeran ${IMPLIED_ONC/B:.1f}B")
print(f"-> Moderna-share peak ${IMPLIED_PEAK_MRNA/B:.1f}B, gross programme peak ${IMPLIED_PEAK_GROSS/B:.1f}B")
print(f"   (Barclays melanoma-only estimate: ~$3.0B gross)")
print("\n=== SCENARIO ARITHMETIC ===")
print(f"{'':34}{'Bear':>12}{'Base':>12}{'Bull':>12}")
print(f"{'Gross programme peak ($B)':34}" + "".join(f"{ONC[k]['peak_gross']/B:>12.1f}" for k in ks))
print(f"{'Oncology value ($B)':34}" + "".join(f"{VAL[k]['onc']/B:>12.1f}" for k in ks))
print(f"{'Total equity value ($B)':34}" + "".join(f"{VAL[k]['total']/B:>12.1f}" for k in ks))
print(f"{'Per share':34}" + "".join(f"{VAL[k]['ps']:>12.2f}" for k in ks))
print(f"{'vs spot':34}" + "".join(f"{VAL[k]['upside']:>11.1%} " for k in ks))
print(f"\nProbability-weighted: ${PW:.2f} ({PW/PX_NOW-1:+.1%} vs ${PX_NOW:.2f})")
print(f"Street: Goldman $67, Morgan Stanley $209 — no usable consensus")
