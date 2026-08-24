#!/usr/bin/env python3
"""
NBIS (Nebius Group N.V.) financial model.

Data entry = 100% from the user's moomoo terminal screenshots (Aug 3 2026, post-market):
  1) Financials > Balance Sheet   (Quarterly, All, USD)
  2) Financials > Income Statement(Quarterly, All, USD)
  3) Financials > Cash Flow       (Quarterly, All, USD)
  4) Financials > Revenue Breakdown (2025/FY)
  5) Financials > Key Indicators  (Quarterly TTM)
  6) Quote panel (price / mkt cap / shares / P-E / P-B / 52wk ...)

Consensus + guidance layer (web, clearly tagged as non-screenshot) is kept in a separate sheet.

Caveat carried through the whole model: Nebius Group is the former Yandex N.V.  The Russian
business was divested in Jul-2024, so every column at/older than 2024/Q2 mixes continuing +
discontinued/restated Yandex figures and is NOT comparable to today's AI-cloud business.
Those columns are entered faithfully but flagged LEGACY.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M = 1e6
B = 1e9

# ---------------------------------------------------------------- periods ----
# moomoo column order, left -> right (newest first)
COLS = ["2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2", "2025/Q1",
        "2024/Q4", "2024/Q3", "2024/Q2", "2024/Q1",
        "2023/Q4", "2023/Q3", "2023/Q2", "2023/Q1",
        "2022/Q4", "2022/Q3"]
LEGACY_FROM = "2024/Q2"          # this column and older = Yandex-era / restated

n = len(COLS)


def row(*vals):
    """Pad a transcribed row to the full column count."""
    v = list(vals)
    assert len(v) <= n, f"too many values: {len(v)}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
# screenshot 2 -- every visible line, in screen order.  None == "–" (n/a)
IS = [
    ("Total Revenue as Reported", row(399.00*M, 227.70*M, 146.10*M, 105.10*M, 50.90*M,
                                      35.20*M, 32.10*M, 14.50*M, 11.40*M,
                                      -4.40*M, 5.00*M, 4.70*M, 4.87*M, -6.10*B, 2.28*B)),
    ("Total Operating Revenue", row(399.00*M, 227.70*M, 146.10*M, 105.10*M, 50.90*M,
                                    35.20*M, 32.10*M, 14.50*M, 11.40*M,
                                    -4.40*M, 5.00*M, 4.70*M, 4.87*M, -6.10*B, 2.28*B)),
    ("Cost of Revenue", row(103.80*M, 68.50*M, 42.90*M, 30.10*M, 24.70*M,
                            21.10*M, 9.90*M, 7.70*M, 8.90*M,
                            -2.50*M, 7.50*M, 5.90*M, 8.17*M, -2.63*B, 952.82*M)),
    ("Gross Profit", row(295.20*M, 159.20*M, 103.20*M, 75.00*M, 26.20*M,
                         14.10*M, 22.20*M, 6.80*M, 2.50*M,
                         -1.90*M, -2.50*M, -1.20*M, -3.30*M, -3.47*B, 1.33*B)),
    ("Operating Expense", row(423.20*M, 409.20*M, 233.40*M, 186.20*M, 146.50*M,
                              150.60*M, 102.80*M, 119.00*M, 85.40*M,
                              53.80*M, 75.10*M, 77.20*M, 67.34*M, -3.19*B, 1.14*B)),
    ("  Selling and Admin Expenses", row(143.80*M, 161.50*M, 89.50*M, 68.20*M, 60.90*M,
                                         85.40*M, 47.90*M, 75.60*M, 51.30*M,
                                         41.10*M, 43.50*M, 42.00*M, 32.57*M, -1.98*B, 722.24*M)),
    ("  Research & Development", row(67.40*M, 53.10*M, 44.90*M, 42.80*M, 36.50*M,
                                     31.90*M, 31.40*M, 32.00*M, 25.20*M,
                                     2.90*M, 25.50*M, 28.70*M, 28.26*M, -849.85*M, 292.04*M)),
    ("  Depreciation & Amortization & Depletion", row(212.00*M, 194.60*M, 99.00*M, 75.20*M, 49.10*M,
                                                      33.30*M, 23.50*M, 11.40*M, 8.90*M,
                                                      9.80*M, 6.10*M, 6.50*M, 6.51*M, -360.24*M, 127.86*M)),
    ("    – Depreciation & Amortization", row(212.00*M, 194.60*M, 99.00*M, 75.20*M, 49.10*M,
                                              33.30*M, 23.50*M, 11.40*M, 8.90*M,
                                              9.80*M, 6.10*M, 6.50*M, 6.51*M, -360.24*M, 127.86*M)),
    ("Operating Profit", row(-128.00*M, -250.00*M, -130.20*M, -111.20*M, -120.30*M,
                             -136.50*M, -80.60*M, -112.20*M, -82.90*M,
                             -55.70*M, -77.60*M, -78.40*M, -70.64*M, -279.71*M, 184.85*M)),
    ("Net Non-Operating Interest Income (Expense)", row(-49.50*M, -28.50*M, -8.50*M, -1.20*M, 8.50*M,
                                                        21.90*M, 28.60*M, 12.70*M, 400.00e3,
                                                        600.00e3, 500.00e3, 1.00*M, 543.48e3, -16.43*M, 5.96*M)),
    ("  Non-Operating Interest Income", row(14.20*M, 13.50*M, 6.20*M, 3.60*M, 8.50*M,
                                            21.90*M, 28.60*M, 12.70*M, 400.00e3,
                                            600.00e3, 500.00e3, 1.00*M, 15.72*M, -59.37*M, 19.29*M)),
    ("  Non-Operating Interest Expense", row(63.70*M, 42.00*M, 14.70*M, 4.80*M, 0.0,
                                             0.0, 0.0, 0.0, 72.89*M,
                                             None, None, None, 15.18*M, 2.65*M, 13.34*M)),
    ("Other Income (Expense)", row(792.90*M, 12.30*M, 18.80*M, 615.70*M, 8.40*M,
                                   -8.40*M, 7.40*M, -14.80*M, -1.00*M,
                                   2.80*M, -17.50*M, 5.00*M, 74.08*M, -800.89*M, 705.60*M)),
    ("  Gain on Sale of Security", row(780.60*M, 25.00*M, 26.20*M, 597.40*M, None,
                                       -8.70*M, 5.60*M)),
    ("  Earnings from Equity Interest", row(-7.60*M, -10.60*M, -7.50*M, -6.30*M, 100.00e3,
                                            0.0, 400.00e3, 0.0, 0.0,
                                            1.10*M, -9.80*M, -1.40*M, -1.71*M, 8.96*M, -15.24*M)),
    ("  Special Income (Charges)", row(*[None]*13, -175.10*M, 651.45*M)),
    ("    – Less: Impairment of Capital Assets", row(*[None]*13, 0.0, 0.0)),
    ("    – Less: Other Special Charges", row(*[None]*13, 34.41*M, None)),
    ("    – Gain on Sale Of Business", row(*[None]*13, -140.70*M, 651.45*M)),
    ("  Other Non-Operating Income (Expenses)", row(19.90*M, -2.10*M, 100.00e3, -8.10*M, 8.30*M,
                                                    300.00e3, 1.40*M, -14.80*M, -1.00*M,
                                                    4.10*M, -7.70*M, 6.40*M, 75.79*M, 2.30*M, 69.39*M)),
]

# ========================================================= BALANCE SHEET =====
BS = [
    ("Total Assets", row(22.30*B, 12.43*B, 10.10*B, 5.10*B, 3.44*B,
                         3.55*B, 3.01*B, 4.41*B, 8.65*B,
                         8.76*B, 7.36*B, 7.50*B, 8.58*B, 8.28*B, 9.64*B)),
    ("Total Current Assets", row(11.24*B, 4.71*B, 5.22*B, 2.03*B, 1.68*B,
                                 2.53*B, 2.38*B, 3.97*B, 3.21*B,
                                 3.45*B, 2.76*B, 2.74*B, 3.56*B, 3.22*B, 3.57*B)),
    ("  Cash & Equivalents + Short-Term Investments", row(9.30*B, 3.68*B, 4.79*B, 1.68*B, 1.45*B,
                                                          2.43*B, 2.29*B, 3.89*B, 157.54*M,
                                                          121.20*M, 924.88*M, 866.98*M, 1.58*B, 1.13*B, 1.73*B)),
    ("    – Cash and Cash Equivalents", row(9.30*B, 3.68*B, 4.79*B, 1.68*B, 1.45*B,
                                            2.43*B, 2.29*B, 2.33*B, 152.44*M,
                                            116.10*M, 883.39*M, 827.40*M, 1.55*B, 1.12*B, 1.71*B)),
    ("    – Short Term Investments", row(None, None, None, None, None,
                                         0.0, 0.0, 1.57*B, 5.10*M,
                                         5.10*M, 41.49*M, 39.58*M, 33.19*M, 10.68*M, 21.09*M)),
    ("  Receivables", row(1.53*B, 853.70*M, 237.10*M, 215.40*M, 108.70*M,
                          40.50*M, 52.20*M, 18.30*M, 16.21*M,
                          11.50*M, 1.29*B, 1.28*B, 1.31*B, 1.42*B, 1.23*B)),
    ("    – Accounts Receivable", row(1.48*B, 720.30*M, 91.20*M, 54.70*M, 24.30*M,
                                      11.20*M, 19.00*M, 12.40*M, 3.46*M,
                                      4.10*M, 717.23*M, 747.32*M, 780.31*M, 778.71*M, 801.90*M)),
    ("    – Loans Receivable", row(None, None, None, None, None,
                                   0.0, None, None, None,
                                   700.00e3, 40.89*M, 24.02*M, 25.57*M, 31.17*M, 24.93*M)),
    ("    – Accrued Interest Receivable", row(None, 400.00e3, 1.20*M, 2.40*M, None,
                                              21.60*M, 27.50*M, None, None,
                                              0.0, 879.14e3, 1.25*M, 983.44e3, None, 3.32*M)),
    ("    – Taxes Receivable", row(46.90*M, 131.40*M, 133.40*M, 158.30*M, 84.40*M,
                                   6.20*M, 5.70*M, 5.90*M, 12.75*M,
                                   5.40*M, 257.85*M, 285.78*M, 283.61*M, 303.38*M, 257.35*M)),
    ("    – Other Receivables", row(None, 1.60*M, 11.30*M, None, None,
                                    1.50*M, None, None, None,
                                    1.30*M, 268.72*M, 225.26*M, 217.40*M, 309.69*M, 143.97*M)),
    ("  Prepaid Assets", row(53.50*M, 152.00*M, 72.60*M, 52.40*M, 22.40*M,
                             33.20*M, 22.10*M, 33.10*M, 27.30*M,
                             299.46*M, 303.81*M, 305.91*M, 272.43*M, 277.47*M, None)),
    ("  Inventory", row(*[None]*9, 240.38*M, 235.25*M, 270.70*M, 361.12*M, 378.79*M, 296.47*M)),
    ("  Restricted Cash", row(None, 6.80*M, 107.10*M, 74.50*M, 80.60*M,
                              600.00e3, None, None, None,
                              2.30*M, 5.27*M, 3.82*M, 2.98*M, 8.63*M, 13.68*M)),
    ("  Current Deferred Assets", row(None, 8.40*M, 5.40*M, 5.60*M, None,
                                      2.10*M, None, None, 0.0)),
    ("  Other Current Assets", row(360.50*M, 12.40*M, 61.40*M, 2.40*M, 24.60*M,
                                   800.00e3, 19.40*M, 19.70*M, 18.30*M,
                                   1.70*M, 8.09*M, 11.90*M, 7.61*M, 11.25*M, 22.15*M)),
    ("Total Non-Current Assets", row(11.07*B, 7.72*B, 4.89*B, 3.07*B, 1.75*B,
                                     1.02*B, 623.30*M, 448.30*M, 5.43*B,
                                     5.30*B, 4.61*B, 4.76*B, 5.02*B, 5.06*B, 6.07*B)),
    ("  Net PPE", row(8.40*B, 6.47*B, 3.82*B, 2.07*B, 1.58*B,
                      891.50*M, 507.00*M, 338.20*M, 194.19*M,
                      146.90*M, 2.07*B, 2.03*B, 2.10*B, 2.10*B, 2.38*B)),
    ("    – Gross PPE", row(8.40*B, 7.11*B, 3.82*B, 2.43*B, 1.58*B,
                            1.13*B, 507.00*M, 338.20*M, 194.19*M,
                            451.70*M, 3.43*B, 3.44*B, 3.55*B, 3.50*B, 3.99*B)),
    ("    – Accumulated Depreciation", row(None, -637.00*M, None, -362.80*M, None,
                                           -236.30*M, None, None, None,
                                           -304.80*M, -1.36*B, -1.41*B, -1.45*B, -1.40*B, -1.61*B)),
    ("  Investments and Advances", row(1.56*B, 847.70*M, 856.80*M, 867.40*M, 97.10*M,
                                       97.10*M, 97.10*M, 97.10*M, 98.36*M,
                                       97.10*M, 205.67*M, 207.26*M, 164.61*M, 132.12*M, 116.13*M)),
]

# ============================================================= CASH FLOW =====
CF = [
    ("Operating Cash Flow", row(2.26*B, 834.30*M, -80.40*M, -171.60*M, -197.50*M,
                                -74.60*M, -55.80*M, 225.47*M, 189.18*M,
                                123.76*M, 291.40*M, -44.09*M, 453.01*M, 397.37*M, -48.47*M)),
    ("  Net cash flow from continuing operations", row(2.26*B, 834.30*M, -80.40*M, -167.90*M, -184.10*M,
                                                       -73.10*M, -34.80*M, -826.33*M, 189.18*M,
                                                       123.76*M, 291.40*M, -880.79*M, 453.01*M, 397.37*M, -48.47*M)),
    ("    – Net Income from Continuing Operations", row(621.20*M, -268.80*M, -119.60*M, 502.50*M, -104.30*M,
                                                        -122.90*M, -43.60*M, -589.60*M, 55.62*M,
                                                        160.13*M, 74.86*M, 779.68*M, 1.55*B, 339.29*M, -156.07*M)),
    ("    – Gain/Loss from Continuing Operations", row(5.90*M, 8.40*M, -6.70*M, -7.90*M, 3.30*M,
                                                       9.50*M, -6.00*M, 227.61*M, -45.84*M,
                                                       -93.52*M, -74.95*M, 800.58*M, -708.84*M, -63.12*M, -15.65*M)),
    ("    – Depreciation & Depletion & Amortization", row(241.80*M, 215.90*M, 111.80*M, 84.70*M, 56.10*M,
                                                          36.80*M, 25.30*M, -435.42*M, 138.53*M,
                                                          152.56*M, 180.73*M, -662.37*M, 220.77*M, 257.39*M, 167.53*M)),
    ("    – Deferred Tax", row(-7.80*M, -2.80*M, -200.00e3, 400.00e3, -800.00e3,
                               -500.00e3, -3.50*M, -762.17e3, 3.80*M,
                               6.72*M, -11.06*M, -286.32*M, 18.56*M, 6.72*M, -454.90e3)),
    ("    – Other Non-Cash Items", row(19.70*M, 67.00*M, 5.70*M, 2.50*M, 1.50*M,
                                       None, -34.98*M, 26.60*M, 13.07*M,
                                       -3.79*M, -17.11*M, 5.48*M, 14.69*M, 3.52*M)),
    ("    – Change in Working Capital", row(2.12*B, 787.50*M, -98.50*M, -168.00*M, -157.60*M,
                                            -36.10*M, -13.20*M, 93.63*M, -30.79*M,
                                            -157.80*M, 94.06*M, 60.66*M, 112.55*M, 20.20*M, -140.80*M)),
    ("        Change in Receivables", row(-673.50*M, -625.60*M, -9.20*M, -96.40*M, -85.20*M,
                                          -15.10*M, -5.80*M, 209.11*M, -124.98*M,
                                          -92.54*M, 17.51*M, 108.77*M, -193.89*M, -39.29*M, 83.59*M)),
    ("        Change in Inventory", row(*[None]*7, 17.70*M, -23.66*M,
                                        41.67*M, 7.36*M, -116.92*M, -25.01*M, -45.70*M, -43.86*M)),
    ("        Change in Prepaid Assets", row(-19.00*M, -11.90*M, -700.00e3, -6.50*M, 1.30*M,
                                             -10.10*M, 4.50*M, 56.53*M, -42.31*M,
                                             -7.33*M, -12.49*M, 69.91*M, -46.81*M, -32.63*M, -8.15*M)),
    ("        Change in Payables and Accrued Expense", row(-64.90*M, 25.40*M, -4.30*M, 6.90*M, -57.00*M,
                                                           -16.90*M, 3.40*M, -187.64*M, 129.02*M,
                                                           -56.06*M, 116.18*M, -378.54*M, 410.51*M, 659.02*M, -125.24*M)),
    ("        Change in Other Current Assets", row(-318.90*M, -166.90*M, -80.60*M, -72.60*M, -19.10*M,
                                                   6.40*M, -19.00*M, 90.40*M, 8.58*M,
                                                   -61.04*M, -45.47*M, 134.53*M, -30.08*M, -47.59*M, -46.60*M)),
    ("        Change in Other Current Liabilities", row(*[None]*7, 7.71*M, 11.19*M,
                                                        -3.32*M, 2.67*M, 10.41*M, -18.61*M, -9.97*M, 2.05*M)),
    ("        Change in Other Working Capital", row(3.20*B, 1.57*B, -3.70*M, 600.00e3, 2.40*M,
                                                    -400.00e3, 3.70*M, -38.86*M, 11.36*M,
                                                    20.82*M, 8.58*M, -27.52*M, 16.44*M, 17.32*M, -2.59*M)),
    ("  Cash Flow from Discontinued Operating", row(0.0, 0.0, 0.0, -3.70*M, -13.40*M,
                                                    -1.50*M, -21.00*M)),
    ("Net cash flow from investing", row(-2.64*B, -2.13*B, -952.00*M, -602.30*M, -544.00*M,
                                         -416.00*M, 12.60*M, -518.21*M, -311.51*M,
                                         -264.66*M, -117.73*M, -386.19*M, -76.32*M, -107.43*M, 109.10*M)),
    ("  Net Cash Flow from Continuing Investing", row(-2.64*B, -2.09*B, -994.70*M, -559.60*M, -543.90*M,
                                                      -1.70*B, 12.60*M, 600.99*M, -311.51*M,
                                                      -264.66*M, -117.73*M, 4.31*B, -76.32*M, -107.43*M, 109.10*M)),
    ("    – Net PPE Purchase and Sale (capex)", row(-2.47*B, -2.06*B, -955.50*M, -510.60*M, -543.90*M,
                                                    -415.90*M, -172.10*M, 587.70*M, -282.05*M,
                                                    -264.93*M, -123.63*M, 506.08*M, -78.91*M, -146.64*M, -215.28*M)),
    ("    – Net Business Purchase and Sale", row(-170.20*M, 42.70*M, 0.0, None, 0.0,
                                                 -1.28*B, 184.20*M, 0.0, 0.0,
                                                 0.0, 0.0, -16.69*M, 30.73*M, 0.0, -9.82*M)),
    ("    – Net Investment Purchase and Sale", row(None, -75.00*M, 0.0, None, 0.0,
                                                   0.0, -11.59*M, -179.74e3, 1.99*M,
                                                   -219.98e3, -486.10*M, -21.14*M, 38.07*M, 327.73*M)),
]

# ======================================================= KEY INDICATORS ======
# TTM ratios, in % (or x / days where noted)
KI = [
    ("— Profitability (TTM) —", row()),
    ("Gross Margin %", row(72.06, 68.63, 59.12, 53.67, 41.76, 37.53, 55.06, 55.00, 55.03,
                           55.00, 54.41, 54.86, 55.10, 55.30, 55.73)),
    ("Operating Margin %", row(-70.55, -115.46, -148.44, -196.43, -301.92, -375.06, 3.00, 3.11, 3.58,
                               3.70, 5.04, 6.19, 4.49, 2.54, 1.20)),
    ("EBIT Margin %", row(97.71, 14.21, 24.44, 59.93, -301.92, -375.06, 6.16, 6.05, 6.75,
                          8.22, 15.60, 14.14, 12.86, None, None)),
    ("Net Margin %", row(93.09, 15.57, 60.03, 97.67, -271.69, -545.87, -6.62, -4.68, 3.32,
                         2.48, 3.19, 13.50, 9.66, 7.56, 7.12)),
    ("EBITDA Margin %", row(172.23, 102.64, 104.29, 141.52, -229.06, -301.62, 11.25, 11.21, 11.07,
                            14.64, 15.93, 27.07, 25.04, 24.53, 23.94)),
    ("Tax Rate %", row(None, 28.99, 1.55, 0.58, None, 50.44, 46.35, 59.94, 49.53,
                       49.65, 21.49, 25.66, 32.32, 27.42, None)),
    ("Interest Coverage (x)", row(6.85, 1.22, None, None, None, None, None, 3.04, 4.97,
                                  7.29, 26.25, 22.94, 21.72, 16.91, None)),
    ("R&D Expense Ratio %", row(23.72, 33.47, 47.07, 63.18, 89.53, 110.38, 12.89, 12.84, 12.82,
                                12.87, 12.32, 12.15, 13.12, 13.85, 14.44)),
    ("— Solvency —", row()),
    ("Long-Term Debt to Equity %", row(130.89, 105.87, 93.46, 31.33, 5.74, 0.93, 0.75, 0.09, 0.16,
                                       0.29, 25.65, 22.52, 22.11, 21.71, 21.81)),
    ("Total Assets to Common Equity %", row(308.02, 270.58, 209.99, 134.99, 108.71, 109.06, 104.58, 101.86, 259.34,
                                            265.81, 235.81, 220.90, 207.14, 194.97, 183.43)),
    ("Equity Ratio %", row(32.47, 36.96, 47.62, 74.08, 91.99, 91.69, 95.62, 98.17, 38.56,
                           37.62, 42.41, 45.27, 48.28, 51.29, 54.52)),
    ("Debt to Asset Ratio % (moomoo defn.)", row(131.15, 108.25, 94.97, 32.43, 5.94, 1.52, 0.97, 0.24, 0.53,
                                                 0.80, 57.80, 50.88, 39.81, 32.80, 32.86)),
    ("Current Ratio (x)", row(8.33, 3.08, 6.57, 14.70, 18.00, 9.60, 22.24, 51.77, 0.86,
                              0.89, 0.89, 0.92, 1.23, 1.28, 1.34)),
    ("Quick Ratio (x)", row(8.02, 2.97, 6.33, 13.72, 16.64, 9.38, 21.85, 51.08, 0.05,
                            0.03, 0.72, 0.72, 0.99, 1.01, 1.11)),
    ("— Operating Capacity (TTM) —", row()),
    ("Cash Conversion Cycle (days)", row(*[None]*9, -37.33, -36.78, -28.60, -24.20, -36.98, -33.27)),
    ("Receivable Turnover (x)", row(1.17, 1.44, 6.59, 7.43, 11.63, 0.24, 24.74, 23.87, 23.08,
                                    10.37, 8.67, 9.18, 12.53, 10.28, 12.10)),
    ("Inventory Turnover (x)", row(*[None]*9, 13.14, 11.29, 10.61, 13.48, 12.34, 16.92)),
    ("Account Payable Turnover (x)", row(0.72, 0.30, 0.41, 2.30, 1.65, 0.11, 8.32, 8.88, 8.35,
                                         3.64, 3.28, 3.55, 4.53, 3.58, 4.29)),
    ("Fixed Assets Turnover (x)", row(0.18, 0.14, 0.17, 0.21, 0.18, 0.07, 7.07, 7.66, 7.89,
                                      3.85, 2.96, 2.88, 4.08, 3.59, 3.66)),
]

# =================================================== REVENUE BREAKDOWN =======
REV_BUSINESS = [("Nebius (AI cloud)", 480.30*M, 90.66), ("TripleTen", 54.10*M, 10.21),
                ("Avride", 1.30*M, 0.25), ("Eliminations", -5.90*M, -1.11)]
REV_REGION = [("United States", 340.10*M, 64.19), ("United Kingdom", 137.50*M, 25.95),
              ("Rest of the world", 52.20*M, 9.85)]

# ========================================================== MARKET DATA ======
MKT = [
    ("Last price (Aug 3 2026 close)", 212.580, "USD"),
    ("Change", 22.170, "+11.64%"),
    ("Post-market", 215.730, "+1.48% (19:25 ET)"),
    ("Previous close", 190.410, "USD"),
    ("Open", 184.085, "USD"),
    ("Day High", 221.670, "USD"),
    ("Day Low", 176.250, "USD"),
    ("Day Range %", 23.85, "%"),
    ("Volume", 28.58e6, "shares"),
    ("Turnover", 5.97*B, "USD"),
    ("Average Price", 208.877, "USD"),
    ("Turnover Rate %", 13.71, "%"),
    ("Market Cap", 53.97*B, "USD"),
    ("Float Market Cap", 44.32*B, "USD"),
    ("Total Shares", 253.9e6, "shares"),
    ("Free Float", 208.47e6, "shares"),
    ("P/E (TTM)", 72.80, "x"),
    ("P/E (LFY)", 644.18, "x"),
    ("P/B", 7.453, "x"),
    ("Beta", 2.915, ""),
    ("52-week High", 299.860, "USD"),
    ("52-week Low", 62.010, "USD"),
    ("Historical High", 299.860, "USD"),
    ("Historical Low", 9.940, "USD"),
    ("Bid/Ask imbalance %", -80.24, "%"),
    ("Volume Ratio", 0.91, "x"),
    ("Dividend TTM", None, "none"),
]

# -------- derived from screenshots (book value, implied balance-sheet items) --
DERIVED = [
    ("Book value of equity  = Mkt cap / P/B", 53.97*B/7.453, "≈ Total assets x Equity Ratio 32.47%"),
    ("Check: Total assets x Equity ratio", 22.30*B*0.3247, "22.30B x 32.47%"),
    ("Implied total liabilities = TA - Equity", 22.30*B - 22.30*B*0.3247, ""),
    ("Implied current liabilities = TCA / Current ratio", 11.24*B/8.33, "11.24B / 8.33"),
    ("Implied non-current liabilities", (22.30*B-22.30*B*0.3247) - 11.24*B/8.33, ""),
    ("Implied long-term debt = Equity x LTD/E 130.89%", 22.30*B*0.3247*1.3089, ""),
    ("Implied net debt = LT debt - cash 9.30B", 22.30*B*0.3247*1.3089 - 9.30*B, ""),
    ("TTM revenue (sum of last 4 reported quarters)", (399.00+227.70+146.10+105.10)*M, "2025/Q2-2026/Q1"),
    ("TTM net income implied by P/E 72.80", 53.97*B/72.80, "one-off-gain inflated"),
    ("Enterprise value = Mkt cap + LT debt - cash", 53.97*B + 22.30*B*0.3247*1.3089 - 9.30*B, ""),
]

# ================================== CONSENSUS / GUIDANCE (non-screenshot) ====
CONSENSUS = [
    ("Next report date", "2026-08-12 (Q2 2026)", "company / press"),
    ("Q2 2026 consensus revenue", "$576.7M (range $576-595M)", "Benzinga / street"),
    ("Q2 2026 consensus EPS (adj.)", "-$0.73", "Benzinga (same as Q1 est., looks stale)"),
    ("Q1 2026 actual vs est.", "rev $399.0M vs $374-392M (beat); adj EPS -$0.33 vs -$0.73", "reported 2026-05-13"),
    ("Q4 2025 actual vs est.", "rev $227.7M vs $238.6M (miss); adj EPS -$0.68 vs -$0.54", ""),
    ("Q3 2025 actual vs est.", "rev $146.1M vs $153.7M (miss); adj EPS -$0.40 vs -$0.44", ""),
    ("Q2 2025 actual vs est.", "rev $105.1M vs $98.7M (beat); adj EPS -$0.38 vs -$0.42", ""),
    ("ARR end-Q1 2026 (Nebius AI)", "$1.90B (vs $1.25B end-2025, +52% q/q)", "company"),
    ("Q1 2026 adj. EBITDA", "$129.5M, group margin 32%; Nebius AI margin 45%", "company"),
    ("FY2026 revenue guidance", "$3.0-3.4B (reiterated at Q1)", "company"),
    ("FY2026 adj. EBITDA margin guidance", "~40%", "company"),
    ("FY2026 capex guidance", "$20-25B (raised from $16-20B)", "company"),
    ("Exit-2026 ARR target", "$7-9B", "company"),
    ("FY2026 street revenue", "$3.39B", "stockanalysis.com, 17 analysts"),
    ("FY2026 street EPS", "-$2.95", "stockanalysis.com"),
    ("FY2027 street revenue", "~$10B (+~200% y/y)", "street composite"),
    ("Analyst mean price target", "$238-258 (23 / 17 analysts)", "marketbeat / stockanalysis"),
    ("Analyst PT range", "$120 low / $410 high", ""),
    ("Consensus rating", "Buy / Overweight", ""),
    ("Key contracts", "Microsoft (~$17-19B), Meta 5-yr ~$27B, Nvidia $2B equity, >4GW contracted power", ""),
]

# ============================================ Q2-2026 FORECAST ENGINE ========
# Calibration: quarterly revenue ~= k * average(entry ARR, exit ARR) / 4
#   Q1 2026: entry ARR 1.25B, exit 1.90B -> avg 1.575B -> /4 = 393.75M vs actual 399.0M
#   => k = 1.013  (the ARR-average bridge is essentially unbiased for this company)
K_ARR = 399.0*M / ((1.25*B + 1.90*B)/2/4)

# SBC add-back calibrated on Q1 2026: reported adj EBITDA 129.5M = op profit (-128.0) + D&A (212.0)
# + X  ->  X = 45.5M.  D&A calibrated as ~2.85% of average net PPE per quarter (212.0 / 7.44B).
SBC_Q1 = 129.5*M - (-128.0*M + 212.0*M)
DA_RATE_Q = 212.0*M / ((6.47*B + 8.40*B)/2)

SCEN = {
    #                exit-Q2 ARR, gross margin, S&A, R&D, D&A, net interest, SBC addback
    "Bear":  dict(arr_exit=2.35*B, gm=0.715, sga=152*M, rnd=72*M, da=290*M, nint=-70*M, sbc=50*M),
    "Base":  dict(arr_exit=2.80*B, gm=0.735, sga=157*M, rnd=76*M, da=280*M, nint=-62*M, sbc=55*M),
    "Bull":  dict(arr_exit=3.20*B, gm=0.750, sga=162*M, rnd=80*M, da=272*M, nint=-55*M, sbc=60*M),
}
ARR_ENTRY_Q2 = 1.90*B
SHARES = 253.9e6


def forecast_q2(p):
    rev = K_ARR * (ARR_ENTRY_Q2 + p["arr_exit"]) / 2 / 4
    gp = rev * p["gm"]
    opex = p["sga"] + p["rnd"] + p["da"]
    op = gp - opex
    adj_ebitda = op + p["da"] + p["sbc"]          # adj EBITDA = op profit + D&A + SBC
    net_gaap = op + p["nint"]                      # excl. one-off investment gains
    adj_net = net_gaap + p["sbc"]
    return dict(revenue=rev, gross_profit=gp, gross_margin=p["gm"], opex=opex,
                op_profit=op, adj_ebitda=adj_ebitda, adj_ebitda_margin=adj_ebitda/rev,
                net_gaap=net_gaap, adj_net=adj_net, adj_eps=adj_net/SHARES,
                gaap_eps=net_gaap/SHARES, arr_exit=p["arr_exit"],
                arr_qoq=p["arr_exit"]/ARR_ENTRY_Q2 - 1)


Q2 = {k: forecast_q2(v) for k, v in SCEN.items()}

# ================================= FY2026-2028 MODEL + VALUATION ============
FY = {
    # revenue path (FY26 anchored on guidance 3.0-3.4B; FY27 on street ~10B; FY28 off contracted backlog)
    # netdebt / shares from the funding bridge below: capex 20-25B in FY26 alone cannot be funded
    # from the 9.3B cash + customer prepayments, so both leverage AND dilution are modelled.
    "Bear": dict(rev26=2.90*B, rev27=6.0*B,  rev28=9.0*B,  ebitda_m=0.32,
                 exit_arr26=5.0*B, ev_sales=4.0, ev_ebitda=12.0,
                 netdebt27=11*B, netdebt28=13*B, shares27=310e6, shares28=320e6),
    "Base": dict(rev26=3.25*B, rev27=10.0*B, rev28=17.0*B, ebitda_m=0.40,
                 exit_arr26=7.5*B, ev_sales=5.4, ev_ebitda=14.0,
                 netdebt27=17*B, netdebt28=22*B, shares27=290e6, shares28=305e6),
    "Bull": dict(rev26=3.55*B, rev27=13.0*B, rev28=24.0*B, ebitda_m=0.44,
                 exit_arr26=9.0*B, ev_sales=7.0, ev_ebitda=16.0,
                 netdebt27=19*B, netdebt28=26*B, shares27=285e6, shares28=300e6),
}
DISCOUNT = 0.12      # cost of equity used to bring FY2028 value back ~1.5y to a 12-mo target
YEARS_BACK = 1.5
PX_NOW = 212.58


def valuation(p):
    ebitda28 = p["rev28"] * p["ebitda_m"]
    # --- method 1: fair value on FY2028, discounted back to a 12-month horizon
    ev_from_sales = p["rev28"] * p["ev_sales"]
    ev_from_ebitda = ebitda28 * p["ev_ebitda"]
    ev = (ev_from_sales + ev_from_ebitda) / 2
    eq = ev - p["netdebt28"]
    px28 = eq / p["shares28"]
    px_m1 = px28 / (1 + DISCOUNT) ** YEARS_BACK
    # --- method 2: multiple roll-forward. In Aug-2027 the tape prices the NEXT year (FY2028E)
    #     at the same EV/forward-sales the stock carries today (5.4x on FY2027E).
    ev_m2 = p["rev28"] * p["ev_sales"]
    px_m2 = (ev_m2 - p["netdebt27"]) / p["shares27"]
    px_target = (px_m1 + px_m2) / 2
    return dict(ebitda28=ebitda28, ev_from_sales=ev_from_sales, ev_from_ebitda=ev_from_ebitda,
                ev=ev, equity=eq, px_2028=px28, px_m1=px_m1, px_m2=px_m2,
                px_target_12m=px_target, upside=px_target/PX_NOW - 1)


VAL = {k: valuation(v) for k, v in FY.items()}

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
LEG = PatternFill("solid", fgColor="4A4A4A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'   # millions
PCT = '0.0%'


def style_header(ws, r, labels, width0=46, width=13):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR if j == 1 else (LEG if lab in COLS[COLS.index(LEGACY_FROM):] else SUB)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = width0
    for j in range(2, len(labels)+1):
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=r+1, column=2)


def write_stmt(wb, title, data, note, fmt=NUMFMT):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=9, color="7F7F7F")
    style_header(ws, 2, ["USD, millions"] + COLS)
    r = 3
    for name, vals in data:
        c = ws.cell(row=r, column=1, value=name)
        if not name.startswith(" "):
            c.font = BOLD
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = fmt
            cell.border = THIN
        r += 1
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---- 0. cover
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 110
lines = [
    ("NBIS — Nebius Group N.V. financial model", ""),
    ("Built", "2026-08-03 (post-market)"),
    ("Primary data source", "moomoo terminal screenshots supplied by the user (IS / BS / CF / Key Indicators / Revenue breakdown / quote panel)"),
    ("Secondary layer", "street consensus + company guidance — kept isolated on the 'Consensus_Guidance' sheet, never mixed into reported figures"),
    ("Latest REPORTED quarter", "2026/Q1 (reported 2026-05-13): revenue $399.0M, +684% y/y"),
    ("Next report", "2026-08-12, Q2 2026 — this is the quarter the 'Forecast_Q2_2026' sheet projects"),
    ("LEGACY WARNING", "Nebius Group = former Yandex N.V.; the Russian business was sold in Jul-2024. Columns 2024/Q2 and older mix continuing + discontinued/restated Yandex figures (grey headers) and are NOT comparable."),
    ("Known data oddity", "Income statement 2022/Q4 shows revenue -$6.10B and matching negative costs — a moomoo restatement artifact, entered as shown but not meaningful."),
    ("Units", "All statement sheets are in USD; number format displays millions."),
]
for i, (a, b) in enumerate(lines, start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

# ---- 1-3. statements
write_stmt(wb, "IS_Quarterly", IS,
           "moomoo > Financials > Income Statement > Quarterly · All · USD  (screenshot 2). '–' entered as blank.")
write_stmt(wb, "BS_Quarterly", BS,
           "moomoo > Financials > Balance Sheet > Quarterly · All · USD (screenshot 1). Screenshot is cut off after 'Investments and Advances'; liability/equity side is reconstructed on 'Market_Derived'.")
write_stmt(wb, "CF_Quarterly", CF,
           "moomoo > Financials > Cash Flow > Quarterly · All · USD (screenshot 3).")

# ---- 4. key indicators (raw %, not millions)
ws = wb.create_sheet("Key_Indicators_TTM")
ws.cell(row=1, column=1, value="moomoo > Financials > Key Indicators (screenshot 5). Values are TTM; % as shown, x = times, days = days.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["TTM indicator"] + COLS)
r = 3
for name, vals in KI:
    c = ws.cell(row=r, column=1, value=name)
    if name.startswith("—"):
        c.font = BOLD
        for j in range(1, n+2):
            ws.cell(row=r, column=j).fill = SEC
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = '0.00'
        cell.border = THIN
    r += 1

# ---- 5. derived quarterly analytics (computed off the entered data)
ws = wb.create_sheet("Derived_Quarterly")
ws.cell(row=1, column=1, value="Computed from the entered statements — not from the screenshots' own ratio tab.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["Derived metric"] + COLS)
rev = dict(IS)["Total Revenue as Reported"]
cogs = dict(IS)["Cost of Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Profit"]
da = dict(IS)["  Depreciation & Amortization & Depletion"]
sga = dict(IS)["  Selling and Admin Expenses"]
rnd = dict(IS)["  Research & Development"]
ocf = dict(CF)["Operating Cash Flow"]
capex = dict(CF)["    – Net PPE Purchase and Sale (capex)"]
ppe = dict(BS)["  Net PPE"]
cash = dict(BS)["    – Cash and Cash Equivalents"]
ar = dict(BS)["    – Accounts Receivable"]


def safe(f, *a):
    try:
        return f(*a)
    except (TypeError, ZeroDivisionError):
        return None


drv = [
    ("Revenue q/q %", [safe(lambda i: rev[i]/rev[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Gross margin % (quarterly)", [safe(lambda i: gp[i]/rev[i], i) for i in range(n)]),
    ("Operating margin % (quarterly)", [safe(lambda i: op[i]/rev[i], i) for i in range(n)]),
    ("EBITDA (op profit + D&A)", [safe(lambda i: op[i]+da[i], i) for i in range(n)]),
    ("EBITDA margin % (quarterly)", [safe(lambda i: (op[i]+da[i])/rev[i], i) for i in range(n)]),
    ("D&A as % of revenue", [safe(lambda i: da[i]/rev[i], i) for i in range(n)]),
    ("S&A as % of revenue", [safe(lambda i: sga[i]/rev[i], i) for i in range(n)]),
    ("R&D as % of revenue", [safe(lambda i: rnd[i]/rev[i], i) for i in range(n)]),
    ("Capex (abs)", [safe(lambda i: -capex[i], i) for i in range(n)]),
    ("Capex / revenue (x)", [safe(lambda i: -capex[i]/rev[i], i) for i in range(n)]),
    ("Free cash flow (OCF + capex)", [safe(lambda i: ocf[i]+capex[i], i) for i in range(n)]),
    ("Net PPE q/q growth %", [safe(lambda i: ppe[i]/ppe[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Annualised revenue (qtr x4)", [safe(lambda i: rev[i]*4, i) for i in range(n)]),
    ("Cash (end of quarter)", cash),
    ("AR / quarterly revenue (x)", [safe(lambda i: ar[i]/rev[i], i) for i in range(n)]),
]
r = 3
for name, vals in drv:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if ("%" in name) else ('0.00"x"' if "(x)" in name else NUMFMT)
        cell.border = THIN
    r += 1

# ---- 6. revenue breakdown
ws = wb.create_sheet("Rev_Breakdown_FY2025")
ws.cell(row=1, column=1, value="moomoo > Financials > Revenue Breakdown, 2025/FY, USD (screenshot 4).").font = Font(italic=True, size=9, color="7F7F7F")
ws.column_dimensions["A"].width = 26
for col, w in zip("BCD", (16, 12, 16)):
    ws.column_dimensions[col].width = w
r = 3
for title, rows in (("By business", REV_BUSINESS), ("By region", REV_REGION)):
    ws.cell(row=r, column=1, value=title).font = WHITE
    for j in range(1, 4):
        ws.cell(row=r, column=j).fill = HDR
    ws.cell(row=r, column=2, value="Revenue").font = WHITE
    ws.cell(row=r, column=3, value="Ratio").font = WHITE
    r += 1
    tot = 0
    for name, val, pct in rows:
        ws.cell(row=r, column=1, value=name)
        c = ws.cell(row=r, column=2, value=val); c.number_format = NUMFMT
        c = ws.cell(row=r, column=3, value=pct/100); c.number_format = PCT
        tot += val
        r += 1
    ws.cell(row=r, column=1, value="Total").font = BOLD
    c = ws.cell(row=r, column=2, value=tot); c.number_format = NUMFMT; c.font = BOLD
    r += 2

# ---- 7. market data + derived balance-sheet reconstruction
ws = wb.create_sheet("Market_Derived")
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 46
ws.cell(row=1, column=1, value="Quote panel, moomoo, 2026-08-03 post-market (screenshot right rail) + items derived from it.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="QUOTE / MARKET DATA (as displayed)").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.000' if (val is not None and abs(val) < 1000) else '#,##0'
    ws.cell(row=r, column=3, value=unit)
    r += 1
r += 1
ws.cell(row=r, column=1, value="RECONSTRUCTED (screenshot cut off the liability side)").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, note in DERIVED:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val); c.number_format = NUMFMT
    ws.cell(row=r, column=3, value=note)
    r += 1

# ---- 8. consensus / guidance
ws = wb.create_sheet("Consensus_Guidance")
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 42
ws.cell(row=1, column=1, value="NOT from the screenshots — company guidance and sell-side consensus gathered from the web on 2026-08-03.").font = Font(italic=True, size=9, color="C00000")
r = 3
for a, b, c_ in CONSENSUS:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    r += 1

# ---- 9. Q2 2026 forecast
ws = wb.create_sheet("Forecast_Q2_2026")
ws.column_dimensions["A"].width = 44
for col in "BCDE":
    ws.column_dimensions[col].width = 16
ws.cell(row=1, column=1, value=f"ARR bridge calibrated on Q1 2026: revenue = k x avg(entry ARR, exit ARR)/4, k = {K_ARR:.3f}").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value="Entry ARR for Q2 2026 = $1.90B (reported end-March). Street: revenue $576.7M, adj EPS -$0.73.").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="Q2 2026 (reports 2026-08-12)").font = WHITE
ws.cell(row=r, column=1).fill = HDR
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Exit ARR (Jun-26)", "arr_exit", NUMFMT, None),
          ("ARR q/q growth", "arr_qoq", PCT, None),
          ("Revenue", "revenue", NUMFMT, 576.67*M),
          ("Gross profit", "gross_profit", NUMFMT, None),
          ("Gross margin", "gross_margin", PCT, None),
          ("Operating expense", "opex", NUMFMT, None),
          ("GAAP operating profit", "op_profit", NUMFMT, None),
          ("Adj. EBITDA", "adj_ebitda", NUMFMT, None),
          ("Adj. EBITDA margin", "adj_ebitda_margin", PCT, None),
          ("GAAP net income (ex one-offs)", "net_gaap", NUMFMT, None),
          ("Adj. net income", "adj_net", NUMFMT, None),
          ("Adj. EPS", "adj_eps", '0.00', -0.73),
          ("GAAP EPS (ex one-offs)", "gaap_eps", '0.00', None)]
for label, key, fmt, street in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k][key]); c.number_format = fmt; c.border = THIN
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    r += 1
r += 1
ws.cell(row=r, column=1, value="Beat/miss vs street revenue").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=Q2[k]["revenue"]/(576.67*M)-1); c.number_format = '+0.0%;-0.0%'

# ---- 10. FY model + valuation
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 46
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="FY2026 anchored on company guidance ($3.0-3.4B, ~40% adj EBITDA margin); FY2027 on street (~$10B); FY2028 extrapolated from exit-2026 ARR and contracted backlog (Meta ~$27B/5y, Microsoft ~$17-19B).").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value=f"Target = blend of EV/Sales and EV/EBITDA on FY2028E, less net debt, / diluted shares, discounted back {YEARS_BACK} years at {DISCOUNT:.0%}.").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
rows = [("FY2026E revenue", lambda k: FY[k]["rev26"], NUMFMT),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT),
        ("FY2028E revenue", lambda k: FY[k]["rev28"], NUMFMT),
        ("Exit-2026 ARR", lambda k: FY[k]["exit_arr26"], NUMFMT),
        ("Adj. EBITDA margin (FY28)", lambda k: FY[k]["ebitda_m"], PCT),
        ("FY2028E adj. EBITDA", lambda k: VAL[k]["ebitda28"], NUMFMT),
        ("EV/Sales multiple applied", lambda k: FY[k]["ev_sales"], '0.0"x"'),
        ("EV/EBITDA multiple applied", lambda k: FY[k]["ev_ebitda"], '0.0"x"'),
        ("EV — sales method", lambda k: VAL[k]["ev_from_sales"], NUMFMT),
        ("EV — EBITDA method", lambda k: VAL[k]["ev_from_ebitda"], NUMFMT),
        ("EV — blended", lambda k: VAL[k]["ev"], NUMFMT),
        ("Net debt FY2028E", lambda k: FY[k]["netdebt28"], NUMFMT),
        ("Equity value", lambda k: VAL[k]["equity"], NUMFMT),
        ("Diluted shares FY2028E", lambda k: FY[k]["shares28"], '#,##0'),
        ("Implied price at FY2028", lambda k: VAL[k]["px_2028"], '$#,##0.00'),
        ("Method 1 — FY28 value discounted 1.5y", lambda k: VAL[k]["px_m1"], '$#,##0.00'),
        ("Method 2 — multiple roll-forward (Aug-27)", lambda k: VAL[k]["px_m2"], '$#,##0.00'),
        ("12-month target (average of methods)", lambda k: VAL[k]["px_target_12m"], '$#,##0.00'),
        ("Upside vs $212.58", lambda k: VAL[k]["upside"], '+0.0%;-0.0%')]
for label, fn, fmt in rows:
    c0 = ws.cell(row=r, column=1, value=label)
    if "target" in label or "Upside" in label:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if "target" in label or "Upside" in label:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
PROB = {"Bear": 0.25, "Base": 0.55, "Bull": 0.20}
pw = sum(PROB[k]*VAL[k]["px_target_12m"] for k in PROB)
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=pw); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=pw/PX_NOW-1); c.number_format = '+0.0%;-0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Sanity checks on today's price ($212.58, mkt cap $53.97B)").font = BOLD
r += 1
ev_now = 53.97*B + 22.30*B*0.3247*1.3089 - 9.30*B
checks = [("Current EV (mkt cap + LT debt - cash)", ev_now, NUMFMT),
          ("EV / TTM revenue", ev_now/((399.00+227.70+146.10+105.10)*M), '0.0"x"'),
          ("EV / FY2026E revenue (street $3.39B)", ev_now/(3.39*B), '0.0"x"'),
          ("EV / FY2027E revenue (street ~$10B)", ev_now/(10.0*B), '0.0"x"'),
          ("EV / exit-2026 ARR (mid $8B)", ev_now/(8.0*B), '0.0"x"'),
          ("EV / FY2027E adj EBITDA (40%)", ev_now/(10.0*B*0.40), '0.0"x"'),
          ("Price / book (moomoo)", 7.453, '0.00"x"')]
for label, val, fmt in checks:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

out = "/Users/antaiwei/Downloads/NBIS_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

# ------------------------------------------------------------ console dump --
print(f"\nARR bridge k = {K_ARR:.4f}")
print("\n== Q2 2026 forecast (reports Aug 12) ==")
print(f"{'':34s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}")
for label, key, _, street in fields:
    vals = "".join(f"{Q2[k][key]/M:>12,.0f}" if abs(Q2[k][key]) > 100 else f"{Q2[k][key]:>12.2f}"
                   for k in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.0f}" if abs(street) > 100 else f"{street:>12.2f}")
    print(f"{label:34s}{vals}{s}")
print(f"\nQ1-2026 calibration: implied SBC add-back {SBC_Q1/M:.1f}M, D&A rate {DA_RATE_Q:.2%} of avg net PPE/qtr")
print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY28 rev {FY[k]['rev28']/B:5.1f}B  EBITDA {v['ebitda28']/B:5.1f}B  EV {v['ev']/B:6.1f}B "
          f"-> M1 ${v['px_m1']:7.2f}  M2 ${v['px_m2']:7.2f}  12m target ${v['px_target_12m']:7.2f}  ({v['upside']:+.1%})")
print(f"\nCurrent EV {ev_now/B:.1f}B | EV/TTM rev {ev_now/((399.00+227.70+146.10+105.10)*M):.1f}x | "
      f"EV/FY27E rev {ev_now/(10.0*B):.1f}x | EV/exit-ARR(8B) {ev_now/(8.0*B):.1f}x")
