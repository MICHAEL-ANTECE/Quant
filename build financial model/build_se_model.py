#!/usr/bin/env python3
"""
SE (Sea Limited) model — same framework as the MU / NBIS / SHOP / APP / SNDK / AXTI /
AAOI / CRDO workbooks.

DATA CAVEAT: built from filings/aggregators, not the user's moomoo screenshots. Income
statement, guidance and market data are exact; balance-sheet / cash-flow / TTM-ratio and
the three-segment split (Shopee / Monee / Garena) are absent at quarterly granularity.

TIMING: Q2 2026 reports 2026-08-11 — five days out.

THE ONE LINE THAT EXPLAINS THE 42% DRAWDOWN:
    Q1 2026 revenue grew +46.6% y/y. Q1 2026 net income grew +6.2% y/y.
Sea is still compounding revenue at 40%+, but essentially none of it is reaching the
bottom line. Incremental operating margin has collapsed to 7-11% — versus CRDO's 40-55%.
Two causes: Shopee's competitive spend against TikTok Shop, and Monee's credit-loss
provisions ballooning as the loan book matures. Management's FY2026 guide that adjusted
EBITDA will be "no lower than" 2025 in absolute dollars is the formal acknowledgement.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2",
        "2025/Q1", "2024/Q4", "2024/Q3", "2024/Q2"]
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Revenue", row(7097*M, 6852*M, 5986*M, 5259*M, 4841*M, 4950*M, 4328*M, 3807*M)),
    ("Gross Profit", row(3146*M, 2998*M, 2599*M, 2410*M, 2236*M, 2205*M, 1861*M, 1585*M)),
    ("Operating Income", row(592.99*M, 565.24*M, 475.95*M, 487.72*M,
                             456.40*M, 305.75*M, 202.42*M, 82.89*M)),
    ("Net Income", row(427.94*M, 397.10*M, 374.99*M, 414.20*M,
                       403.05*M, 237.31*M, 153.32*M, 79.91*M)),
    ("Diluted EPS", row(0.67, 0.63, 0.59, 0.65, 0.65, 0.39, 0.24, 0.14)),
]

# ============================================================== GUIDANCE =====
GUIDE = [
    ("Q2 2026 report date", "2026-08-11", "company"),
    ("FY2026 Shopee GMV guidance", "approximately +25% y/y", "management"),
    ("FY2026 adjusted EBITDA guidance", "'NO LOWER THAN' 2025 levels in absolute dollars",
     "management — the phrase that de-rated the stock"),
    ("Why that language matters", "for a stock on a premium growth multiple, 'no lower than' is a de facto admission that margin expansion stalls", "market interpretation"),
    ("What broke in Q4 2025 (reported Mar-2026)", "record revenue but a disappointing EPS line; the stock fell ~23-25% on the print", "market"),
    ("Shopee pressure", "aggressive competitive spending against TikTok Shop; margin compression", "company/street"),
    ("Monee (fintech) growth", "revenue +54.3% y/y to $1.1B", "company"),
    ("Monee risk", "provisions for credit losses ballooning as the loan book expands; default normalisation as the book matures", "street"),
    ("Third narrative", "perceived AI disintermediation risk on e-commerce discovery", "street"),
    ("Analyst consensus", "Strong Buy, average target $141.97 (+23.6%)", "stockanalysis"),
]

# ========================================================== MARKET DATA ======
PX_NOW = 114.86
SHARES_OUT = 612.48e6
DILUTED_Q1 = 427.94 / 0.67 * 1e6          # implied diluted share count
MKT = [
    ("Last price (2026-08-06)", 114.86, "USD, +1.26%"),
    ("Market cap", 70.35*B, "USD"),
    ("Shares outstanding", 612.48e6, "shares"),
    ("Implied diluted shares (Q1 net / EPS)", DILUTED_Q1, "shares"),
    ("52-week high", 199.30, "USD"),
    ("52-week low", 77.05, "USD"),
    ("Drawdown from 52wk high", 114.86/199.30-1, "PCT"),
    ("P/E (TTM)", 44.66, "x"),
    ("Forward P/E", 27.59, "x -> implies street forward EPS of ~$4.16"),
    ("TTM revenue", 25.19*B, "+40.5%"),
    ("TTM EPS", 2.54, "+78.6%"),
    ("Analyst mean target", 141.97, "USD (+23.6%)"),
    ("Next earnings", None, "2026-08-11"),
]

# ============================================ Q2-2026 FORECAST ENGINE ========
Q1_REV = 7097*M
Q2_25_REV = 5259*M
Q1_GM = 3146/7097
Q1_OPEX = (3146 - 592.99) * M              # $2,553M
Q1_NET_TO_OP = 427.94 / 592.99             # 72.2% — tax + minority + FX drag
DILUTED = 638.7e6

SCEN = {
    #             revenue, gross margin, opex % of revenue, net/op conversion
    "Bear": dict(rev=7350*M, gm=0.430, opex_pct=0.370, conv=0.700,
                 note="competitive spend keeps rising; Monee provisions step up again"),
    "Base": dict(rev=7550*M, gm=0.440, opex_pct=0.362, conv=0.722,
                 note="revenue compounds, profit stays flat — the Q1 pattern repeats"),
    "Bull": dict(rev=7800*M, gm=0.450, opex_pct=0.350, conv=0.740,
                 note="Shopee take-rate gains outrun the spend; provisions stabilise"),
}


def forecast_q2(p):
    gp = p["rev"] * p["gm"]
    opex = p["rev"] * p["opex_pct"]
    op = gp - opex
    net = op * p["conv"]
    return dict(revenue=p["rev"], yoy=p["rev"]/Q2_25_REV-1, qoq=p["rev"]/Q1_REV-1,
                gross_profit=gp, gross_margin=p["gm"], opex=opex,
                op_income=op, op_margin=op/p["rev"], net=net, net_margin=net/p["rev"],
                eps=net/DILUTED, eps_yoy=(net/DILUTED)/0.65-1,
                op_yoy=op/(487.72*M)-1, net_yoy=net/(414.20*M)-1)


Q2 = {k: forecast_q2(v) for k, v in SCEN.items()}

# ================================= FY2026-2028 MODEL + VALUATION ============
FY25_REV = (4841 + 5259 + 5986 + 6852) * M       # $22,938M
FY25_NET = (403.05 + 414.20 + 374.99 + 397.10) * M

FY = {
    # The single swing variable is NET MARGIN — i.e. whether the incremental revenue ever
    # reaches the bottom line. Revenue growth is not really in dispute.
    "Bear": dict(q3=8150*M, q4=9250*M, rev27=38.0*B, nm27=0.055, rev28=44.0*B, nm28=0.060,
                 pe27=16.0, pe28=15.0, sh=643e6,
                 note="TikTok Shop war grinds on; Monee credit cycle turns"),
    "Base": dict(q3=8456*M, q4=9640*M, rev27=42.5*B, nm27=0.075, rev28=53.0*B, nm28=0.090,
                 pe27=24.0, pe28=22.0, sh=643e6,
                 note="spend peaks in 2026; margin expansion resumes in 2027"),
    "Bull": dict(q3=8800*M, q4=10200*M, rev27=46.0*B, nm27=0.095, rev28=60.0*B, nm28=0.110,
                 pe27=32.0, pe28=28.0, sh=640e6,
                 note="Shopee wins the market, Monee scales cleanly, operating leverage returns"),
}
PROB = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}
DISCOUNT, YEARS_BACK = 0.12, 1.5


def valuation(k):
    p = FY[k]
    rev26 = Q1_REV + Q2[k]["revenue"] + p["q3"] + p["q4"]
    eps27 = p["rev27"] * p["nm27"] / p["sh"]
    eps28 = p["rev28"] * p["nm28"] / p["sh"]
    px_m1 = eps27 * p["pe27"]
    px_m2 = eps28 * p["pe28"] / (1 + DISCOUNT) ** YEARS_BACK
    target = (px_m1 + px_m2) / 2
    return dict(rev26=rev26, growth26=rev26/FY25_REV-1, eps27=eps27, eps28=eps28,
                px_m1=px_m1, px_m2=px_m2, target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(k) for k in FY}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WARN = PatternFill("solid", fgColor="FCE4D6")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'

wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("SE — Sea Limited model", ""),
    ("Built", "2026-08-06, price $114.86"),
    ("DATA CAVEAT", "Built from filings/aggregators, NOT the user's moomoo screenshots. Income statement, guidance and market data are exact; balance-sheet / cash-flow / TTM-ratio sheets and the quarterly Shopee / Monee / Garena segment split are absent. Send the moomoo Financials + Revenue Breakdown tabs to fill them in."),
    ("TIMING", "Q2 2026 reports 2026-08-11 — five days out."),
    ("Latest REPORTED quarter", "Q1 2026: revenue $7,097M (+46.6% y/y), gross margin 44.32%, operating income $592.99M (8.4%), net income $427.94M, EPS $0.67"),
    ("THE ONE LINE", "Q1 2026 revenue grew +46.6% y/y while net income grew only +6.2% y/y. That gap is the entire investment debate and the entire reason the stock is 42% below its 52-week high while revenue compounds at 40%+."),
    ("Incremental operating margin", "7.5% / -1.6% / 10.3% / 11.3% over the last four quarters. Every incremental dollar of revenue is dropping roughly a dime to operating profit. For contrast, CRDO runs 40-55%."),
    ("The two causes", "(1) Shopee's competitive spending against TikTok Shop compresses margin; (2) Monee's credit-loss provisions balloon as the loan book expands and defaults normalise. Monee revenue is +54.3% y/y to $1.1B — the provisions scale with it."),
    ("The guidance tell", "FY2026 adjusted EBITDA guided 'no lower than' 2025 in absolute dollars. Not 'grows' — 'no lower than'. Management is telling you margin expansion stalls this year."),
    ("Units", "USD; statement sheet displays millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("IS_Quarterly")
ws.cell(row=1, column=1, value="Reported quarterly income statement (calendar quarters).").font = Font(italic=True, size=9, color="7F7F7F")
for j, lab in enumerate(["USD, millions"] + COLS, start=1):
    c = ws.cell(row=2, column=j, value=lab); c.fill = HDR if j == 1 else SUB; c.font = WHITE
    c.alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 44
for j in range(2, n+2):
    ws.column_dimensions[get_column_letter(j)].width = 14
ws.freeze_panes = ws.cell(row=3, column=2)
r = 3
for name, vals in IS:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v)
        c.number_format = '0.00' if "EPS" in name else NUMFMT
        c.border = THIN
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED — the growth-vs-profit gap").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
rev = dict(IS)["Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Income"]
ni = dict(IS)["Net Income"]
for name, vals, fmt, hl in [
        ("Revenue y/y %", [rev[i]/rev[i+4]-1 if i+4 < n else None for i in range(n)], PCT, True),
        ("Operating income y/y %", [op[i]/op[i+4]-1 if i+4 < n else None for i in range(n)], PCT, False),
        ("NET INCOME y/y %", [ni[i]/ni[i+4]-1 if i+4 < n else None for i in range(n)], PCT, True),
        ("Revenue q/q %", [rev[i]/rev[i+1]-1 if i+1 < n else None for i in range(n)], PCT, False),
        ("Gross margin %", [gp[i]/rev[i] for i in range(n)], PCT, False),
        ("Operating margin %", [op[i]/rev[i] for i in range(n)], PCT, False),
        ("Net margin %", [ni[i]/rev[i] for i in range(n)], PCT, False),
        ("Implied opex (GP - op income)", [gp[i]-op[i] for i in range(n)], NUMFMT, False),
        ("Opex as % of revenue", [(gp[i]-op[i])/rev[i] for i in range(n)], PCT, False),
        ("INCREMENTAL operating margin", [(op[i]-op[i+1])/(rev[i]-rev[i+1]) if i+1 < n else None for i in range(n)], PCT, True),
        ("Net / operating income conversion", [ni[i]/op[i] for i in range(n)], PCT, False)]:
    c0 = ws.cell(row=r, column=1, value=name)
    if hl:
        c0.font = BOLD
    for j, v in enumerate(vals, start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = fmt; c.border = THIN
        if hl:
            c.fill = WARN
    r += 1

ws = wb.create_sheet("Guidance_Market")
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 34
r = 1
ws.cell(row=r, column=1, value="GUIDANCE & THE DE-RATING STORY").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for a, b, c_ in GUIDE:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    if "NO LOWER" in str(b):
        for j in range(1, 4):
            ws.cell(row=r, column=j).fill = WARN
    r += 1
r += 1
ws.cell(row=r, column=1, value="MARKET DATA").font = WHITE
for j in range(1, 4):
    ws.cell(row=r, column=j).fill = HDR
r += 1
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = PCT if unit == "PCT" else ('#,##0.00' if (val is not None and abs(val) < 10000) else NUMFMT)
    ws.cell(row=r, column=3, value="" if unit == "PCT" else unit)
    r += 1

ws = wb.create_sheet("Forecast_Q2_2026")
ws.column_dimensions["A"].width = 42
for col in "BCDE":
    ws.column_dimensions[col].width = 16
ws.cell(row=1, column=1, value=f"Q1 2026 base: revenue ${Q1_REV/M:,.0f}M, gross margin {Q1_GM:.2%}, opex ${Q1_OPEX/M:,.0f}M ({Q1_OPEX/Q1_REV:.1%} of revenue), net/operating conversion {Q1_NET_TO_OP:.1%}.").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value="Year-ago comparison: Q2 2025 revenue $5,259M, operating income $487.72M, net income $414.20M, EPS $0.65.").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="Q2 2026 (reports 2026-08-11)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Q2 2025 actual"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, 5259*M),
          ("Revenue y/y", "yoy", PCT, None),
          ("Revenue q/q", "qoq", PCT, None),
          ("Gross profit", "gross_profit", NUMFMT, 2410*M),
          ("Gross margin", "gross_margin", PCT, 2410/5259),
          ("Operating expense", "opex", NUMFMT, (2410-487.72)*M),
          ("Operating income", "op_income", NUMFMT, 487.72*M),
          ("Operating margin", "op_margin", PCT, 487.72/5259),
          ("Operating income y/y", "op_yoy", PCT, None),
          ("Net income", "net", NUMFMT, 414.20*M),
          ("Net margin", "net_margin", PCT, 414.20/5259),
          ("NET INCOME y/y", "net_yoy", PCT, None),
          ("Diluted EPS", "eps", '0.00', 0.65),
          ("EPS y/y", "eps_yoy", PCT, None)]
for label, key, fmt, prior in fields:
    c0 = ws.cell(row=r, column=1, value=label)
    hl = "NET INCOME" in label
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k][key]); c.number_format = fmt; c.border = THIN
        if hl:
            c.fill = WARN
            c.font = BOLD
    if prior is not None:
        c = ws.cell(row=r, column=5, value=prior); c.number_format = fmt
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=SCEN[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 50
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="The swing variable is NET MARGIN, not revenue. Revenue growth is not seriously in dispute; whether it ever reaches the bottom line is.").font = Font(italic=True, size=9, color="C00000")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt, hl in [
        ("FY2025 actual revenue", lambda k: FY25_REV, NUMFMT, False),
        ("FY2025 actual net income", lambda k: FY25_NET, NUMFMT, False),
        ("Q3 2026E revenue", lambda k: FY[k]["q3"], NUMFMT, False),
        ("Q4 2026E revenue", lambda k: FY[k]["q4"], NUMFMT, False),
        ("FY2026E revenue", lambda k: VAL[k]["rev26"], NUMFMT, False),
        ("FY2026E growth", lambda k: VAL[k]["growth26"], PCT, False),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT, False),
        ("FY2027E NET MARGIN", lambda k: FY[k]["nm27"], PCT, True),
        ("FY2027E EPS", lambda k: VAL[k]["eps27"], '$0.00', False),
        ("FY2028E revenue", lambda k: FY[k]["rev28"], NUMFMT, False),
        ("FY2028E NET MARGIN", lambda k: FY[k]["nm28"], PCT, True),
        ("FY2028E EPS", lambda k: VAL[k]["eps28"], '$0.00', False),
        ("P/E applied — FY2027", lambda k: FY[k]["pe27"], '0.0"x"', False),
        ("P/E applied — FY2028", lambda k: FY[k]["pe28"], '0.0"x"', False),
        ("Route 1 — FY2027 earnings", lambda k: VAL[k]["px_m1"], '$#,##0.00', False),
        ("Route 2 — FY2028 discounted", lambda k: VAL[k]["px_m2"], '$#,##0.00', False),
        ("12-month target (avg)", lambda k: VAL[k]["target"], '$#,##0.00', True),
        ("Upside vs $114.86", lambda k: VAL[k]["upside"], '+0.0%;-0.0%', True)]:
    c0 = ws.cell(row=r, column=1, value=label)
    if hl:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if hl:
            c.font = BOLD
            c.fill = WARN if "MARGIN" in label else SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario thesis").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")
r += 3
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
mc = 70.35*B
for label, val, fmt in [
        ("Market cap / TTM revenue ($25.19B)", mc/(25.19*B), '0.00"x"'),
        ("Market cap / FY2026E revenue (base)", mc/VAL["Base"]["rev26"], '0.00"x"'),
        ("Market cap / FY2027E revenue (base)", mc/FY["Base"]["rev27"], '0.00"x"'),
        ("P/E on FY2027E base EPS", PX_NOW/VAL["Base"]["eps27"], '0.0"x"'),
        ("Street mean PT $141.97 implies FY27 P/E of", 141.97/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk high $199.30 implies FY27 P/E of", 199.30/VAL["Base"]["eps27"], '0.0"x"'),
        ("52wk low $77.05 implies FY27 P/E of", 77.05/VAL["Base"]["eps27"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027 net margin is worth this much (revenue $42.5B, 643M shares, 24x P/E)").font = BOLD
r += 1
for j, h in enumerate(["Net margin", "Net income", "EPS", "Implied price"], start=1):
    c = ws.cell(row=r, column=j, value=h); c.fill = SUB; c.font = WHITE
r += 1
for nm_ in (0.050, 0.060, 0.075, 0.090, 0.105, 0.120):
    net_ = 42.5*B*nm_
    eps_ = net_/643e6
    px_ = eps_*24
    c = ws.cell(row=r, column=1, value=nm_); c.number_format = PCT; c.fill = SUB; c.font = WHITE
    for j, (v, f) in enumerate([(net_, NUMFMT), (eps_, '$0.00'), (px_, '$#,##0.00')], start=2):
        c = ws.cell(row=r, column=j, value=v); c.number_format = f
        if j == 4 and abs(px_ - PX_NOW) < 12:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/SE_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print("\nThe growth-vs-profit gap, by quarter:")
print(f"{'quarter':10s}{'revenue':>10s}{'rev y/y':>10s}{'op inc':>10s}{'net inc':>10s}{'net y/y':>10s}{'incr op mgn':>13s}")
for i in range(n):
    ry = f"{rev[i]/rev[i+4]-1:>9.1%}" if i+4 < n else " " * 10
    ny = f"{ni[i]/ni[i+4]-1:>9.1%}" if i+4 < n else " " * 10
    im = f"{(op[i]-op[i+1])/(rev[i]-rev[i+1]):>12.1%}" if i+1 < n else " " * 13
    print(f"{COLS[i]:10s}{rev[i]/M:>10,.0f}{ry}{op[i]/M:>10,.0f}{ni[i]/M:>10,.0f}{ny}{im}")

print("\n== Q2 2026 forecast (reports 2026-08-11) ==")
print(f"{'':32s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Q2-25 act':>12s}")
for label, key, _, prior in fields:
    vals = "".join(f"{Q2[k][key]/M:>12,.0f}" if abs(Q2[k][key]) > 100 else f"{Q2[k][key]:>12.3f}"
                   for k in ("Bear", "Base", "Bull"))
    p_ = "" if prior is None else (f"{prior/M:>12,.0f}" if abs(prior) > 100 else f"{prior:>12.3f}")
    print(f"{label:32s}{vals}{p_}")

print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY26 ${v['rev26']/B:5.1f}B (+{v['growth26']:.0%}) | FY27 ${FY[k]['rev27']/B:5.1f}B nm {FY[k]['nm27']:.1%} "
          f"EPS ${v['eps27']:5.2f} | FY28 EPS ${v['eps28']:5.2f} | M1 ${v['px_m1']:7.2f} M2 ${v['px_m2']:7.2f} "
          f"-> ${v['target']:7.2f} ({v['upside']:+.0%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW}) | street $141.97")
print(f"Today = {PX_NOW/VAL['Base']['eps27']:.1f}x FY27E base EPS | street PT = {141.97/VAL['Base']['eps27']:.1f}x")
print(f"Mkt cap/TTM rev {mc/(25.19*B):.2f}x | /FY26E {mc/VAL['Base']['rev26']:.2f}x | /FY27E {mc/FY['Base']['rev27']:.2f}x")
