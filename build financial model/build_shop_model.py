#!/usr/bin/env python3
"""
SHOP (Shopify Inc.) financial model — same framework as the MU / NBIS models.

Data entry = 100% from the user's moomoo terminal screenshots (Aug 4 2026, 11:54 ET intraday):
  1) Valuation > Analyst Ratings (consensus rating, target price, 15 institution rows)
  2) Financials > Balance Sheet    (Quarterly, All, USD)
  3) Financials > Cash Flow        (Quarterly, All, USD)
  4) Financials > Income Statement (Quarterly, All, USD)
  5) Financials > Key Indicators   (Quarterly TTM)
  6) Quote panel + Trade Overview / money flow

Street consensus gathered from the web is tagged separately and never mixed into reported data.

NEXT EVENT: Q2 2026 results, 2026-08-05 PRE-MARKET (tomorrow).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2", "2025/Q1",
        "2024/Q4", "2024/Q3", "2024/Q2", "2024/Q1",
        "2023/Q4", "2023/Q3", "2023/Q2", "2023/Q1",
        "2022/Q4", "2022/Q3"]
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"too many values: {len(v)}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Total Revenue as Reported", row(3.17*B, 3.67*B, 2.84*B, 2.68*B, 2.36*B,
                                      2.81*B, 2.16*B, 2.05*B, 1.86*B,
                                      2.14*B, 1.71*B, 1.69*B, 1.51*B, 1.74*B, 1.37*B)),
    ("Total Operating Revenue", row(3.17*B, 3.67*B, 2.84*B, 2.68*B, 2.36*B,
                                    2.81*B, 2.16*B, 2.05*B, 1.86*B,
                                    2.14*B, 1.71*B, 1.69*B, 1.51*B, 1.74*B, 1.37*B)),
    ("Cost of Revenue", row(1.62*B, 1.98*B, 1.45*B, 1.38*B, 1.19*B,
                            1.46*B, 1.04*B, 1.00*B, 904.00*M,
                            1.08*B, 813.00*M, 859.00*M, 791.00*M, 937.00*M, 704.00*M)),
    ("Gross Profit", row(1.55*B, 1.69*B, 1.39*B, 1.30*B, 1.17*B,
                         1.35*B, 1.12*B, 1.05*B, 957.00*M,
                         1.06*B, 901.00*M, 835.00*M, 717.00*M, 798.00*M, 662.00*M)),
    ("Operating Expense", row(1.05*B, 948.00*M, 900.00*M, 931.00*M, 891.00*M,
                              811.00*M, 777.00*M, 762.00*M, 820.00*M,
                              728.00*M, 745.00*M, 1.10*B, 868.00*M, 952.00*M, 969.00*M)),
    ("  Selling and Admin Expenses", row(611.00*M, 558.00*M, 525.00*M, 537.00*M, 514.00*M,
                                         460.00*M, 445.00*M, 413.00*M, 485.00*M,
                                         417.00*M, 432.00*M, 452.00*M, 410.00*M, 512.00*M, 557.00*M)),
    ("    – Selling & Marketing Expense", row(496.00*M, 433.00*M, 410.00*M, 415.00*M, 405.00*M,
                                              348.00*M, 331.00*M, 353.00*M, 361.00*M,
                                              317.00*M, 295.00*M, 321.00*M, 287.00*M, 298.00*M, 302.00*M)),
    ("    – General & Admin Expense", row(115.00*M, 125.00*M, 115.00*M, 122.00*M, 109.00*M,
                                          112.00*M, 114.00*M, 60.00*M, 124.00*M,
                                          100.00*M, 137.00*M, 131.00*M, 123.00*M, 214.00*M, 255.00*M)),
    ("  Research & Development", row(437.00*M, 390.00*M, 375.00*M, 394.00*M, 377.00*M,
                                     351.00*M, 332.00*M, 349.00*M, 335.00*M,
                                     311.00*M, 313.00*M, 648.00*M, 458.00*M, 440.00*M, 412.00*M)),
    # moomoo shows this line as a duplicate of the restructuring charge; oldest-column
    # alignment is ambiguous in the screenshot, taken from the restructuring row.
    ("  Other Operating Expenses", row(None, None, None, 80.00*M, 75.00*M,
                                       76.00*M, 58.00*M, 42.00*M, 51.00*M,
                                       45.00*M, 34.00*M, 31.00*M, 42.00*M, 34.00*M, 39.00*M)),
    ("Operating Profit", row(498.00*M, 745.00*M, 491.00*M, 371.00*M, 278.00*M,
                             541.00*M, 341.00*M, 283.00*M, 137.00*M,
                             334.00*M, 156.00*M, -265.00*M, -151.00*M, -154.00*M, -307.00*M)),
    ("Net Non-Operating Interest Income (Expense)", row(75.00*M, 79.00*M, 81.00*M, 106.00*M, 65.00*M,
                                                        72.00*M, 77.00*M, 80.00*M, 79.00*M,
                                                        68.00*M, 63.00*M, 58.00*M, None, 38.00*M, 20.00*M)),
    ("  Non-Operating Interest Income", row(75.00*M, 79.00*M, 81.00*M, 106.00*M, 65.00*M,
                                            72.00*M, 77.00*M, 80.00*M, 79.00*M,
                                            68.00*M, 63.00*M, 58.00*M, None, 38.00*M, 20.00*M)),
    ("  Non-Operating Interest Expense", row()),
    ("  Total Other Finance Costs", row()),
    ("Other Income (Expense)", row(-1.21*B, 68.00*M, -264.00*M, 602.00*M, -1.11*B,
                                   808.00*M, 442.00*M, -160.00*M, -472.00*M,
                                   280.00*M, 562.00*M, -1.09*B, 227.00*M, -498.00*M, 129.00*M)),
    ("  Gain on Sale of Security", row(-1.07*B, 154.00*M, -95.00*M, 706.00*M, -1.02*B,
                                       906.00*M, 528.00*M, -74.00*M, -377.00*M,
                                       373.00*M, 553.00*M, 277.00*M, None, -464.00*M, 168.00*M)),
    ("  Earnings from Equity Interest", row(-21.00*M, 28.00*M, -21.00*M, -24.00*M, -23.00*M,
                                            -22.00*M, -28.00*M, -44.00*M, -44.00*M,
                                            -48.00*M, -10.00*M, None, None, 0.0, 0.0)),
    ("  Special Income (Charges)", row(-116.00*M, -114.00*M, -148.00*M, -80.00*M, -75.00*M,
                                       -76.00*M, -58.00*M, -42.00*M, -51.00*M,
                                       -45.00*M, -34.00*M, -1.37*B, -42.00*M, -34.00*M, -39.00*M)),
    ("    – Less: Restructuring and M&A", row(116.00*M, 114.00*M, 148.00*M, 80.00*M, 75.00*M,
                                              76.00*M, 58.00*M, 42.00*M, 51.00*M,
                                              45.00*M, 34.00*M, 31.00*M, 42.00*M, 34.00*M, 39.00*M)),
    ("    – Less: Write Off", row(None, None, None, None, 0.0,
                                  0.0, 0.0, None, 0.0,
                                  0.0, None, 1.34*B, 0.0, 0.0, 0.0)),
]

# ========================================================= BALANCE SHEET =====
BS = [
    ("Total Assets", row(14.12*B, 15.19*B, 15.04*B, 14.56*B, 13.40*B,
                         13.92*B, 12.33*B, 11.35*B, 11.11*B,
                         11.30*B, 10.46*B, 9.70*B, 10.96*B, 10.76*B, 11.20*B)),
    ("Total Current Assets", row(8.50*B, 8.30*B, 8.71*B, 8.00*B, 7.45*B,
                                 7.25*B, 6.54*B, 6.40*B, 6.44*B,
                                 6.28*B, 6.16*B, 5.91*B, 5.93*B, 6.05*B, 6.01*B)),
    ("  Cash & Equivalents + Short-Term Investments", row(5.74*B, 5.77*B, 6.35*B, 5.82*B, 5.51*B,
                                                          5.47*B, 4.90*B, 5.02*B, 5.18*B,
                                                          5.00*B, 4.92*B, 4.78*B, 4.86*B, 5.05*B, 4.94*B)),
    ("    – Cash and Cash Equivalents", row(1.85*B, 1.53*B, 2.41*B, 1.54*B, 1.31*B,
                                            1.49*B, 1.51*B, 1.54*B, 1.62*B,
                                            1.41*B, 1.29*B, 1.61*B, 1.74*B, 1.65*B, 1.38*B)),
    ("    – Short Term Investments", row(3.90*B, 4.23*B, 3.94*B, 4.28*B, 4.21*B,
                                         3.98*B, 3.39*B, 3.48*B, 3.55*B,
                                         3.60*B, 3.63*B, 3.17*B, 3.13*B, 3.40*B, 3.56*B)),
    ("  Receivables", row(2.55*B, 2.37*B, 2.13*B, 1.95*B, 1.69*B,
                          1.66*B, 1.44*B, 1.19*B, 1.18*B,
                          1.16*B, 1.08*B, 981.00*M, 905.00*M, 885.00*M, 919.23*M)),
    ("    – Accounts Receivable", row(99.00*M, 421.00*M, 79.00*M, 69.00*M, 71.00*M,
                                      280.00*M, 58.00*M, 275.00*M, 265.00*M,
                                      206.00*M, 48.00*M, 65.00*M, 325.00*M, 500.00*M, 516.41*M)),
    ("    – Loans Receivable (Shopify Capital)", row(2.10*B, 1.46*B, 1.73*B, 1.60*B, 1.39*B,
                                                     1.02*B, 961.00*M, 918.00*M, 815.00*M,
                                                     672.00*M, 679.00*M, 547.00*M, 408.00*M, 209.00*M, 207.60*M)),
    ("    – Accrued Interest Receivable", row(34.00*M, 39.00*M, 36.00*M, 39.00*M, 10.00*M,
                                              19.00*M, 21.00*M, None, None,
                                              15.00*M, 18.00*M, 20.00*M, 19.00*M, 16.00*M, 9.88*M)),
    ("    – Taxes Receivable", row(71.00*M, 143.00*M, 74.00*M, 41.00*M, 35.00*M,
                                   81.00*M, 39.00*M, None, None,
                                   56.00*M, 35.00*M, 37.00*M, 40.00*M, 36.00*M, 66.38*M)),
    ("    – Other Receivables", row(245.00*M, 306.00*M, 207.00*M, 201.00*M, 181.00*M,
                                    254.00*M, 356.00*M, None, None,
                                    206.00*M, 301.00*M, 312.00*M, 138.00*M, 124.00*M, 118.97*M)),
    ("    – Receivables Adjustments", row(*[None]*12, -25.00*M)),
    ("  Prepaid Assets", row(None, 124.00*M, None, None, None,
                             94.00*M, None, None, None,
                             86.00*M, None, None, 82.00*M)),
    ("  Inventory", row(None, 21.00*M, None, None, None,
                        26.00*M, None, None, None, 19.00*M)),
    ("  Restricted Cash", row(None, 11.00*M, None, None, None,
                              9.00*M, None, None, None, 8.00*M)),
    ("  Other Current Assets", row(206.00*M, None, 227.00*M, 234.00*M, 242.00*M,
                                   None, 211.00*M, 188.00*M, 183.00*M,
                                   1.00*M, 162.00*M, 148.00*M, 159.00*M, 29.00*M, 152.32*M)),
    ("Total Non-Current Assets", row(5.63*B, 6.89*B, 6.33*B, 6.56*B, 5.96*B,
                                     6.67*B, 5.79*B, 4.94*B, 4.67*B,
                                     5.02*B, 4.30*B, 3.79*B, 5.04*B, 4.71*B, 5.19*B)),
    ("  Net PPE", row(138.00*M, 141.00*M, 145.00*M, 143.00*M, 151.00*M,
                      140.00*M, 143.00*M, 144.00*M, 148.00*M,
                      147.00*M, 124.00*M, 170.00*M, 482.00*M, 486.00*M, 478.20*M)),
    ("    – Gross PPE", row(138.00*M, 316.00*M, 145.00*M, 143.00*M, 151.00*M,
                            310.00*M, 143.00*M, 144.00*M, 148.00*M,
                            303.00*M, 124.00*M, 170.00*M, 482.00*M, 639.00*M, 478.20*M)),
    ("    – Accumulated Depreciation", row(None, -175.00*M, None, None, None,
                                           -170.00*M, None, None, None,
                                           -156.00*M, None, None, None, -153.00*M)),
    ("  Investments and Advances (equity stakes)", row(4.82*B, 6.16*B, 5.61*B, 5.83*B, 5.15*B,
                                                       6.00*B, 5.13*B, 4.28*B, 4.03*B,
                                                       4.38*B, 3.68*B, 3.13*B, 2.28*B, 1.95*B, 2.42*B)),
]

# ============================================================= CASH FLOW =====
CF = [
    ("Operating Cash Flow", row(481.00*M, 725.00*M, 513.00*M, 428.00*M, 367.00*M,
                                615.00*M, 423.00*M, 340.00*M, 238.00*M,
                                448.00*M, 278.00*M, 118.00*M, 100.00*M, 98.00*M, -134.00*M)),
    ("  Net cash flow from continuing operations", row(481.00*M, 725.00*M, 513.00*M, 428.00*M, 367.00*M,
                                                       615.00*M, 423.00*M, 340.00*M, 238.00*M,
                                                       448.00*M, 278.00*M, 118.00*M, 100.00*M, 98.00*M, -134.00*M)),
    ("    – Net Income from Continuing Operations", row(-581.00*M, 743.00*M, 264.00*M, 906.00*M, -682.00*M,
                                                        1.29*B, 828.00*M, 171.00*M, -273.00*M,
                                                        657.00*M, 718.00*M, -1.31*B, 68.00*M, -623.00*M, -159.00*M)),
    ("    – Gain/Loss from Continuing Operations", row(1.10*B, -185.00*M, 116.00*M, -704.00*M, 1.03*B,
                                                       -873.00*M, -503.00*M, 118.00*M, 423.00*M,
                                                       -332.00*M, -537.00*M, -280.00*M, -218.00*M, 462.00*M, -159.00*M)),
    ("    – Depreciation & Depletion & Amortization", row(7.00*M, 7.00*M, 8.00*M, 8.00*M, 8.00*M,
                                                          8.00*M, 8.00*M, 10.00*M, 10.00*M,
                                                          10.00*M, 13.00*M, 17.00*M, 30.00*M, 30.00*M, 29.00*M)),
    ("    – Deferred Tax", row(-139.00*M, 22.00*M, -3.00*M, 98.00*M, 1.00*M,
                               72.00*M, 2.00*M, 2.00*M, 1.00*M,
                               -3.00*M, -1.00*M, 2.00*M, 1.00*M, 2.00*M, -1.00*M)),
    ("    – Other Non-Cash Items", row(-12.00*M, -12.00*M, -12.00*M, -12.00*M, -13.00*M,
                                       -19.00*M, -19.00*M, -21.00*M, -35.00*M,
                                       -35.00*M, -36.00*M, -39.00*M, -48.00*M, -34.00*M, -33.00*M)),
    ("    – Change in Working Capital", row(-70.00*M, -6.00*M, -70.00*M, -35.00*M, -6.00*M,
                                            -27.00*M, -46.00*M, -72.00*M, -21.00*M,
                                            24.00*M, -27.00*M, 94.00*M, 104.00*M, 23.00*M, 17.00*M)),
    ("        Change in Other Working Capital", row(-70.00*M, 130.00*M, -70.00*M, -35.00*M, -6.00*M,
                                                    165.00*M, -46.00*M, -72.00*M, -21.00*M,
                                                    -111.00*M, -27.00*M, 94.00*M, 154.00*M, 17.00*M)),
    ("Net cash flow from investing", row(310.00*M, -654.00*M, 335.00*M, -252.00*M, -619.00*M,
                                         -661.00*M, -474.00*M, -424.00*M, -27.00*M,
                                         -346.00*M, -606.00*M, -273.00*M, -19.00*M, 159.00*M, -1.83*B)),
    ("  Net Cash Flow from Continuing Investing", row(310.00*M, -654.00*M, 335.00*M, -252.00*M, -619.00*M,
                                                      -661.00*M, -474.00*M, -424.00*M, -27.00*M,
                                                      -346.00*M, -606.00*M, -273.00*M, -19.00*M, 159.00*M, -1.83*B)),
    ("    – Net PPE Purchase and Sale (capex)", row(-5.00*M, -10.00*M, -6.00*M, -6.00*M, -4.00*M,
                                                    -4.00*M, -2.00*M, -7.00*M, -6.00*M,
                                                    -2.00*M, -2.00*M, -21.00*M, -14.00*M, -8.00*M, -14.00*M)),
    ("    – Net Business Purchase and Sale", row(-1.00*M, -16.00*M, -8.00*M, -71.00*M, -6.00*M,
                                                 -26.00*M, -8.00*M, -132.00*M, -1.00*M,
                                                 -260.00*M, 0.0, -14.00*M, -121.00*M, -5.00*M, -1.78*B)),
    ("    – Net Investment Purchase and Sale", row(620.00*M, -597.00*M, 496.00*M, 0.0, -387.00*M,
                                                   -529.00*M, -249.00*M, -171.00*M, 10.00*M,
                                                   -71.00*M, -445.00*M, -34.00*M, 299.00*M, 173.00*M, 49.00*M)),
    ("    – Net Other Investing Changes", row(-304.00*M, -31.00*M, -147.00*M, -175.00*M, -168.00*M,
                                              -102.00*M, -215.00*M, -114.00*M, -30.00*M,
                                              -13.00*M, -159.00*M, -204.00*M, -183.00*M, -1.00*M, -80.00*M)),
    ("Financing Cash Flow", row(-485.00*M, -938.00*M, 24.00*M, 44.00*M, 59.00*M,
                                49.00*M, 6.00*M, 3.00*M, 3.00*M,
                                17.00*M, 11.00*M, 26.00*M, 6.00*M, 7.00*M, 3.00*M)),
]

# ======================================================= KEY INDICATORS ======
KI = [
    ("— Profitability (TTM) —", row()),
    ("Gross Margin %", row(47.97, 48.07, 48.75, 49.34, 49.94, 50.36, 50.93, 51.07, 50.65,
                           49.79, 48.88, 47.79, 47.99, 49.18, 50.49)),
    ("Operating Margin %", row(17.02, 16.31, 15.72, 15.29, 15.39, 12.11, 13.33, 11.72, 4.88,
                               -1.10, -6.23, -13.92, -12.88, -14.68, -9.36)),
    ("EBIT Margin %", row(17.02, 16.31, 15.72, 15.29, 15.39, 12.11, 13.33, 11.72, 4.88,
                          -1.10, -50.35, -60.48, -62.58, -64.67, -66.65)),
    ("Net Margin %", row(10.77, 10.65, 16.65, 23.42, 17.17, 22.74, 16.84, 16.40, -2.82,
                         1.87, -17.27, -32.13, -32.49, -61.79, -61.16)),
    ("EBITDA Margin %", row(17.27, 16.58, 16.02, 15.61, 15.75, 12.51, 13.80, 12.27, 5.56,
                            -0.11, -49.03, -58.83, -60.84, -63.06, -65.07)),
    ("Tax Rate %", row(19.03, 18.42, 12.61, 9.46, 6.07, 9.38, 7.12, 6.19, None, 28.65)),
    ("Interest Coverage (x)", row(*[None]*13, -3.15e3, -3.04e3)),
    ("R&D Expense Ratio %", row(12.91, 13.29, 14.00, 14.52, 15.02, 15.39, 16.16, 16.85, 21.68,
                                24.50, 27.95, 31.07, 28.07, 26.84, 25.48)),
    ("Sales Expense Ratio %", row(14.18, 14.39, 14.71, 14.97, 15.32, 15.69, 16.59, 17.08, 17.46,
                                  17.28, 18.07, 19.17, 20.57, 21.97, 23.04)),
    ("Administrative Expense Rate %", row(3.86, 4.08, 4.28, 4.56, 4.21, 4.62, 4.85, 5.42, 6.64,
                                          6.95, 9.09, 11.47, 12.23, 12.64, 11.33)),
    ("— Solvency —", row()),
    ("Long-Term Debt to Equity %", row(1.30, 1.27, 1.41, 1.65, 1.81, 1.64, 11.10, 12.21, 12.68,
                                       12.50, 13.36, 14.98, 16.26, 16.73, 14.75)),
    ("Total Assets to Common Equity %", row(112.96, 112.74, 120.21, 120.10, 121.26, 120.47, 121.85, 123.70, 124.95,
                                            124.63, 126.32, 130.01, 129.64, 130.56, 128.88)),
    ("Equity Ratio %", row(88.53, 88.70, 83.19, 83.26, 82.47, 83.01, 82.07, 80.84, 80.03,
                           80.24, 79.17, 76.91, 77.13, 76.59, 77.59)),
    ("Debt to Asset Ratio %", row(1.43, 1.40, 8.91, 9.39, 10.30, 9.74, 11.28, 12.40, 12.88,
                                  12.68, 13.54, 15.20, 16.54, 16.94, 14.94)),
    ("Current Ratio (x)", row(6.20, 5.96, 3.93, 3.85, 3.71, 3.71, 7.10, 7.32, 7.14,
                              6.99, 7.23, 6.71, 6.71, 7.07, 6.64)),
    ("Quick Ratio (x)", row(6.05, 5.85, 3.83, 3.74, 3.59, 3.65, 6.87, 7.11, 6.94,
                            6.85, 7.04, 6.55, 6.53, 6.94, 6.47)),
    ("— Operating Capacity (TTM) —", row()),
    ("Cash Conversion Cycle (days)", row(None, -15.78, None, None, -17.09)),
    ("Receivable Turnover (x)", row(145.48, 32.97, 156.15, 58.22, 55.83, 34.02, 154.94, 45.67, 25.13,
                                    20.38, 23.57, 24.63, 15.73, 12.55, 10.22)),
    ("Inventory Turnover (x)", row(None, 255.36, None, None, 169.54)),
]

# ================================================ ANALYST RATINGS (screenshot 1)
CONSENSUS_RATING = dict(rating="Strong Buy", analysts=26, updated="2026-08-04",
                        buy=92.31, hold=7.69, sell=0.00,
                        pt_high=200.00, pt_avg=158.41, pt_low=130.00, current=120.23)
RATINGS = [  # broker, rating, prior PT, new PT, action, date
    ("BMO Capital", "Buy", 145, 145, "Maintained", "2026-08-03"),
    ("Citi", "Buy", 156, 150, "Downgrade (PT cut)", "2026-07-24"),
    ("J.P. Morgan", "Buy", 150, 155, "Upgrade", "2026-07-23"),
    ("Piper Sandler", "Buy", 150, 150, "Maintained", "2026-07-22"),
    ("Morgan Stanley", "Buy", 192, 192, "Maintained", "2026-07-21"),
    ("Rothschild & Co Redburn", "Hold", 220, 130, "Downgrade", "2026-07-21"),
    ("RBC Capital", "Buy", 170, 170, "Maintained", "2026-07-20"),
    ("Jefferies", "Buy", 160, 160, "Maintained", "2026-07-16"),
    ("Wedbush", "Buy", 160, 155, "Downgrade (PT cut)", "2026-07-15"),
    ("Stifel", "Buy", 175, 150, "Upgrade", "2026-07-10"),
    ("CIBC", "Buy", 185, 185, "Maintained", "2026-07-08"),
    ("BofA Securities", "Buy", 190, 150, "Downgrade (PT cut)", "2026-07-07"),
    ("William Blair", "Buy", None, None, "Maintained", "2026-06-03"),
    ("National Bank", "Buy", None, None, "New", "2026-05-28"),
    ("Truist Financial", "Buy", 150, 150, "Maintained", "2026-05-08"),
]

# ========================================================== MARKET DATA ======
PX_NOW = 120.230
SHARES = 1.30e9
MKT = [
    ("Last price (Aug 4 2026, 11:54 ET)", 120.230, "USD, +2.75%"),
    ("Previous close", 117.010, "USD"),
    ("Open", 117.600, "USD"), ("Day High", 120.910, "USD"), ("Day Low", 117.600, "USD"),
    ("Day Range %", 2.83, "%"),
    ("Volume", 7.45e6, "shares"), ("Turnover", 891.89*M, "USD"),
    ("Average Price", 119.687, "USD"), ("Turnover Rate %", 0.61, "%"),
    ("Vol Ratio", 2.01, "x"), ("Bid/Ask %", 89.09, "% (bid-heavy book)"),
    ("Market Cap", 156.02*B, "USD"), ("Float Market Cap", 146.27*B, "USD"),
    ("Total Shares", 1.30e9, "shares"), ("Free Float", 1.22e9, "shares"),
    ("P/E (TTM)", 117.87, "x"), ("P/E (LFY)", 127.90, "x"), ("P/B", 12.479, "x"),
    ("Beta", 2.082, ""),
    ("52-week High", 182.190, "USD"), ("52-week Low", 94.000, "USD"),
    ("Historical High", 182.190, "USD"), ("Historical Low", 1.848, "USD"),
    ("Dividend TTM", None, "none"),
    ("Next earnings", None, "2026-08-05 PRE-MARKET"),
    ("Money flow — Net Inflow", 7.83*M, "USD (intraday)"),
    ("  Extra Large net", (5.88-6.83)*M, "in 5.88M / out 6.83M"),
    ("  Large net", (44.34-34.81)*M, "in 44.34M / out 34.81M"),
    ("  Medium net", (17.66-17.72)*M, "in 17.66M / out 17.72M"),
    ("  Small net", (64.09-64.78)*M, "in 64.09M / out 64.78M"),
]

# ========================================== STREET CONSENSUS (non-screenshot) =
STREET = [
    ("Q2 2026 report date", "2026-08-05, pre-market", "moomoo quote panel"),
    ("Q2 2026 consensus revenue", "$3.43B (+28.1% y/y)", "Zacks consensus"),
    ("Q2 2026 consensus EPS (adj.)", "$0.39 (+11.4% y/y)", "Zacks, unchanged over 30 days"),
    ("Company guidance for Q2", "revenue growth in the high-20% range y/y", "given at Q1 print"),
    ("Analyst mean PT (moomoo)", "$158.41 (26 analysts, Strong Buy)", "screenshot"),
    ("Analyst mean PT (other trackers)", "~$148-149 (33 analysts, Buy)", "web, post-July cuts"),
    ("July PT cuts", "avg PT trimmed ~$31 to ~$148: softer growth assumptions, higher discount rate, lower terminal P/E", "web"),
    ("Key bear catalyst", "Rothschild & Co Redburn cut to Hold, PT 220->130, citing Meta competition", "2026-07-21"),
    ("Key bull moves", "J.P. Morgan upgrade (PT 155), Jefferies Buy (160), Stifel upgrade (150)", "July 2026"),
]

# ============================================ Q2-2026 FORECAST ENGINE ========
# Two independent bridges, both anchored on the entered quarterly history:
#   (a) seasonal step  Q2 = Q1 x (Q2/Q1 ratio of 2023/2024/2025)
#   (b) YoY            Q2 = Q2-2025 x (1 + y/y growth trend)
Q1_26 = 3.17*B
Q2_25 = 2.68*B
QOQ_STEPS = [2.68/2.36, 2.05/1.86, 1.69/1.51]          # 2025, 2024, 2023 Q2/Q1
YOY_RECENT = [3.17/2.36, 3.67/2.81, 2.84/2.16, 2.68/2.05]  # last 4 quarters y/y

SCEN = {
    #            y/y growth, gross margin, S&M, G&A, R&D, other opex, net interest, tax
    "Bear": dict(yoy=0.280, gm=0.478, sm=515*M, ga=125*M, rnd=450*M, oth=0, nint=72*M, tax=0.19),
    "Base": dict(yoy=0.325, gm=0.483, sm=510*M, ga=120*M, rnd=455*M, oth=0, nint=75*M, tax=0.19),
    "Bull": dict(yoy=0.360, gm=0.490, sm=505*M, ga=118*M, rnd=458*M, oth=0, nint=78*M, tax=0.19),
}


def forecast_q2(p):
    rev = Q2_25 * (1 + p["yoy"])
    gp = rev * p["gm"]
    opex = p["sm"] + p["ga"] + p["rnd"] + p["oth"]
    op = gp - opex
    pretax = op + p["nint"]
    net = pretax * (1 - p["tax"])
    return dict(revenue=rev, yoy=p["yoy"], qoq=rev/Q1_26-1, gross_profit=gp,
                gross_margin=p["gm"], opex=opex, op_profit=op, op_margin=op/rev,
                pretax=pretax, adj_net=net, adj_eps=net/SHARES,
                fcf=rev*0.16, fcf_margin=0.16)


Q2 = {k: forecast_q2(v) for k, v in SCEN.items()}
ST_REV, ST_EPS = 3.43*B, 0.39

# ================================= FY2026-2028 MODEL + VALUATION ============
# FY2025 actual = 2.36 + 2.68 + 2.84 + 3.67 = 11.55B
FY25 = (2.36 + 2.68 + 2.84 + 3.67) * B
NET_CASH = 5.74*B - 0.163*B      # cash+STI less implied LT debt (equity 12.50B x 1.30%)

FY = {
    "Bear": dict(rev26=14.85*B, rev27=17.5*B, rev28=20.4*B, opm27=0.185, fcfm27=0.175,
                 ev_sales=7.0, ev_fcf=36.0, pe=45.0),
    "Base": dict(rev26=15.25*B, rev27=19.0*B, rev28=23.2*B, opm27=0.200, fcfm27=0.200,
                 ev_sales=9.5, ev_fcf=45.0, pe=58.0),
    "Bull": dict(rev26=15.60*B, rev27=20.5*B, rev28=26.0*B, opm27=0.215, fcfm27=0.215,
                 ev_sales=12.0, ev_fcf=55.0, pe=70.0),
}
PROB = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}


def valuation(p):
    fcf27 = p["rev27"] * p["fcfm27"]
    op27 = p["rev27"] * p["opm27"]
    eps27 = (op27 + 0.30*B) * (1 - 0.19) / SHARES
    px_sales = (p["rev27"] * p["ev_sales"] + NET_CASH) / SHARES
    px_fcf = (fcf27 * p["ev_fcf"] + NET_CASH) / SHARES
    px_pe = eps27 * p["pe"]
    target = (px_sales + px_fcf + px_pe) / 3
    return dict(fcf27=fcf27, op27=op27, eps27=eps27, px_sales=px_sales, px_fcf=px_fcf,
                px_pe=px_pe, target=target, upside=target/PX_NOW - 1)


VAL = {k: valuation(v) for k, v in FY.items()}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'


def style_header(ws, r, labels, width0=46, width=13):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR if j == 1 else SUB
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = width0
    for j in range(2, len(labels)+1):
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=r+1, column=2)


def write_stmt(wb, title, data, note):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=9, color="7F7F7F")
    style_header(ws, 2, ["USD, millions"] + COLS)
    r = 3
    for name, vals in data:
        c = ws.cell(row=r, column=1, value=name)
        if not name.startswith(" "):
            c.font = BOLD
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = NUMFMT
            cell.border = THIN
        r += 1
    return ws


wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 110
for i, (a, b) in enumerate([
    ("SHOP — Shopify Inc. financial model", ""),
    ("Built", "2026-08-04 (intraday, price $120.23)"),
    ("Primary data source", "moomoo terminal screenshots supplied by the user: Analyst Ratings / Balance Sheet / Cash Flow / Income Statement / Key Indicators / quote panel + money flow"),
    ("Secondary layer", "street consensus, kept isolated on the 'Street_Consensus' sheet"),
    ("Latest REPORTED quarter", "2026/Q1: revenue $3.17B, +34.3% y/y, operating profit $498M; GAAP net LOSS $581M driven by -$1.07B of equity-stake marks"),
    ("NEXT REPORT", "2026-08-05 PRE-MARKET (Q2 2026) — the quarter 'Forecast_Q2_2026' projects"),
    ("Data note", "moomoo leaves several balance-sheet lines blank in non-Q4 quarters (prepaid, inventory, restricted cash, accumulated depreciation) — these are annual-only disclosures, entered as blanks rather than zeros."),
    ("Data note 2", "'Other Operating Expenses' duplicates the restructuring charge; the oldest column alignment is ambiguous in the screenshot and was taken from the restructuring row."),
    ("Units", "USD; number format displays millions."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

write_stmt(wb, "IS_Quarterly", IS, "moomoo > Financials > Income Statement > Quarterly · All · USD.")
write_stmt(wb, "BS_Quarterly", BS, "moomoo > Financials > Balance Sheet > Quarterly · All · USD.")
write_stmt(wb, "CF_Quarterly", CF, "moomoo > Financials > Cash Flow > Quarterly · All · USD.")

ws = wb.create_sheet("Key_Indicators_TTM")
ws.cell(row=1, column=1, value="moomoo > Financials > Key Indicators. TTM basis.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["TTM indicator"] + COLS)
r = 3
for name, vals in KI:
    c = ws.cell(row=r, column=1, value=name)
    if name.startswith("—"):
        c.font = BOLD
        for j in range(1, n+2):
            ws.cell(row=r, column=j).fill = SEC
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = '0.00'
        cell.border = THIN
    r += 1

# derived quarterly
ws = wb.create_sheet("Derived_Quarterly")
ws.cell(row=1, column=1, value="Computed from the entered statements.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["Derived metric"] + COLS)
rev = dict(IS)["Total Revenue as Reported"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Profit"]
sm = dict(IS)["    – Selling & Marketing Expense"]
rnd = dict(IS)["  Research & Development"]
oth = dict(IS)["Other Income (Expense)"]
ocf = dict(CF)["Operating Cash Flow"]
capex = dict(CF)["    – Net PPE Purchase and Sale (capex)"]
fin = dict(CF)["Financing Cash Flow"]
loans = dict(BS)["    – Loans Receivable (Shopify Capital)"]
inv = dict(BS)["  Investments and Advances (equity stakes)"]
ni = dict(CF)["    – Net Income from Continuing Operations"]


def safe(f, i):
    try:
        return f(i)
    except (TypeError, ZeroDivisionError):
        return None


drv = [
    ("Revenue y/y %", [safe(lambda i: rev[i]/rev[i+4]-1, i) if i+4 < n else None for i in range(n)]),
    ("Revenue q/q %", [safe(lambda i: rev[i]/rev[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Gross margin % (quarterly)", [safe(lambda i: gp[i]/rev[i], i) for i in range(n)]),
    ("Operating margin % (quarterly)", [safe(lambda i: op[i]/rev[i], i) for i in range(n)]),
    ("S&M as % of revenue", [safe(lambda i: sm[i]/rev[i], i) for i in range(n)]),
    ("R&D as % of revenue", [safe(lambda i: rnd[i]/rev[i], i) for i in range(n)]),
    ("Operating cash flow margin %", [safe(lambda i: ocf[i]/rev[i], i) for i in range(n)]),
    ("Free cash flow (OCF + capex)", [safe(lambda i: ocf[i]+capex[i], i) for i in range(n)]),
    ("FCF margin %", [safe(lambda i: (ocf[i]+capex[i])/rev[i], i) for i in range(n)]),
    ("GAAP net income (from CF stmt)", ni),
    ("Equity-stake P&L (Other income)", oth),
    ("GAAP net income EX other income", [safe(lambda i: ni[i]-oth[i], i) for i in range(n)]),
    ("Shopify Capital loan book", loans),
    ("Loan book y/y %", [safe(lambda i: loans[i]/loans[i+4]-1, i) if i+4 < n else None for i in range(n)]),
    ("Equity stakes carrying value", inv),
    ("Financing CF (negative = buyback)", fin),
]
r = 3
for name, vals in drv:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if "%" in name else NUMFMT
        cell.border = THIN
    r += 1

# analyst ratings
ws = wb.create_sheet("Analyst_Ratings")
ws.column_dimensions["A"].width = 30
for col, w in zip("BCDEF", (12, 12, 12, 24, 14)):
    ws.column_dimensions[col].width = w
ws.cell(row=1, column=1, value="moomoo > Valuation > Analyst Ratings (screenshot, updated 2026-08-04).").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for k, v in [("Consensus rating", CONSENSUS_RATING["rating"]),
             ("Analysts covering", CONSENSUS_RATING["analysts"]),
             ("Buy / Hold / Sell %", f"{CONSENSUS_RATING['buy']} / {CONSENSUS_RATING['hold']} / {CONSENSUS_RATING['sell']}"),
             ("Target price — High", CONSENSUS_RATING["pt_high"]),
             ("Target price — Average", CONSENSUS_RATING["pt_avg"]),
             ("Target price — Low", CONSENSUS_RATING["pt_low"]),
             ("Current price", CONSENSUS_RATING["current"]),
             ("Implied upside to mean PT", CONSENSUS_RATING["pt_avg"]/CONSENSUS_RATING["current"]-1)]:
    ws.cell(row=r, column=1, value=k).font = BOLD
    c = ws.cell(row=r, column=2, value=v)
    if k.startswith("Implied"):
        c.number_format = '+0.0%'
    r += 1
r += 1
for j, h in enumerate(["Institution", "Rating", "Prior PT", "New PT", "Action", "Date"], start=1):
    c = ws.cell(row=r, column=j, value=h); c.fill = HDR; c.font = WHITE
r += 1
for name, rating, old, new, action, date in RATINGS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=rating)
    ws.cell(row=r, column=3, value=old)
    ws.cell(row=r, column=4, value=new)
    c = ws.cell(row=r, column=5, value=action)
    if old and new and new < old:
        c.font = Font(color="C00000")
    ws.cell(row=r, column=6, value=date)
    r += 1

# market data
ws = wb.create_sheet("Market_Data")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 40
ws.cell(row=1, column=1, value="moomoo quote panel + Trade Overview, 2026-08-04 11:54 ET.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.000' if (val is not None and abs(val) < 1000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
ttm_rev = (3.17+3.67+2.84+2.68)*B
ttm_gp = (1.55+1.69+1.39+1.30)*B
ttm_ocf = (481+725+513+428)*M
ttm_fcf = ttm_ocf - (5+10+6+6)*M
ev = 156.02*B - NET_CASH
for name, val, fmt in [
        ("Book value = Mkt cap / P/B", 156.02*B/12.479, NUMFMT),
        ("Check: Total assets x Equity ratio 88.53%", 14.12*B*0.8853, NUMFMT),
        ("Implied LT debt = Equity x LTD/E 1.30%", 14.12*B*0.8853*0.0130, NUMFMT),
        ("Net cash (cash+STI - LT debt)", NET_CASH, NUMFMT),
        ("Enterprise value", ev, NUMFMT),
        ("TTM revenue (2025/Q2-2026/Q1)", ttm_rev, NUMFMT),
        ("TTM gross profit", ttm_gp, NUMFMT),
        ("TTM operating cash flow", ttm_ocf, NUMFMT),
        ("TTM free cash flow", ttm_fcf, NUMFMT),
        ("TTM FCF margin", ttm_fcf/ttm_rev, PCT),
        ("EV / TTM revenue", ev/ttm_rev, '0.0"x"'),
        ("EV / TTM gross profit", ev/ttm_gp, '0.0"x"'),
        ("EV / TTM FCF", ev/ttm_fcf, '0.0"x"'),
        ("Rule of 40 (rev growth + FCF margin)", 0.343 + ttm_fcf/ttm_rev, PCT)]:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

# street consensus
ws = wb.create_sheet("Street_Consensus")
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 68
ws.column_dimensions["C"].width = 34
ws.cell(row=1, column=1, value="NOT from the screenshots — gathered from the web 2026-08-04.").font = Font(italic=True, size=9, color="C00000")
r = 3
for a, b, c_ in STREET:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    r += 1

# Q2 forecast
ws = wb.create_sheet("Forecast_Q2_2026")
ws.column_dimensions["A"].width = 42
for col in "BCDE":
    ws.column_dimensions[col].width = 16
ws.cell(row=1, column=1, value=f"Seasonal Q2/Q1 steps 2023-25: {', '.join(f'{x:.1%}' for x in QOQ_STEPS)} -> implies ${Q1_26*sum(QOQ_STEPS)/3/B:.2f}B").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value=f"Trailing 4 quarters y/y: {', '.join(f'{x-1:.1%}' for x in YOY_RECENT)} -> guidance is 'high-20%', street $3.43B / $0.39").font = Font(italic=True, size=9, color="7F7F7F")
r = 4
ws.cell(row=r, column=1, value="Q2 2026 (reports 2026-08-05 pre-market)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, ST_REV),
          ("Revenue y/y", "yoy", PCT, 0.281),
          ("Revenue q/q", "qoq", PCT, None),
          ("Gross profit", "gross_profit", NUMFMT, None),
          ("Gross margin", "gross_margin", PCT, None),
          ("Operating expense", "opex", NUMFMT, None),
          ("Operating profit", "op_profit", NUMFMT, None),
          ("Operating margin", "op_margin", PCT, None),
          ("Pre-tax (ex equity marks)", "pretax", NUMFMT, None),
          ("Adj. net income", "adj_net", NUMFMT, None),
          ("Adj. EPS", "adj_eps", '0.00', ST_EPS),
          ("Free cash flow (16% margin)", "fcf", NUMFMT, None)]
for label, key, fmt, street in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k][key]); c.number_format = fmt; c.border = THIN
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    r += 1
r += 1
for label, key, base in [("Beat/miss vs street revenue", "revenue", ST_REV),
                         ("Beat/miss vs street EPS", "adj_eps", ST_EPS)]:
    ws.cell(row=r, column=1, value=label).font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q2[k][key]/base-1); c.number_format = '+0.0%;-0.0%'
    r += 1
r += 1
ws.cell(row=r, column=1, value="GAAP wildcard: equity-stake marks ran -$1.07B in 2026/Q1 and +$906M in 2024/Q4. Carrying value $4.82B. GAAP EPS is not the number to trade.").font = Font(italic=True, size=9, color="C00000")

# valuation
ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 44
for col in "BCD":
    ws.column_dimensions[col].width = 18
ws.cell(row=1, column=1, value="FY2026E built from the quarterly seasonal model; FY2027/28 extrapolated. Target = average of three routes on FY2027E: EV/Sales, EV/FCF and P/E.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt in [
        ("FY2025 actual revenue", lambda k: FY25, NUMFMT),
        ("FY2026E revenue", lambda k: FY[k]["rev26"], NUMFMT),
        ("FY2026E growth", lambda k: FY[k]["rev26"]/FY25-1, PCT),
        ("FY2027E revenue", lambda k: FY[k]["rev27"], NUMFMT),
        ("FY2027E growth", lambda k: FY[k]["rev27"]/FY[k]["rev26"]-1, PCT),
        ("FY2028E revenue", lambda k: FY[k]["rev28"], NUMFMT),
        ("FY2027E operating margin", lambda k: FY[k]["opm27"], PCT),
        ("FY2027E operating profit", lambda k: VAL[k]["op27"], NUMFMT),
        ("FY2027E FCF margin", lambda k: FY[k]["fcfm27"], PCT),
        ("FY2027E free cash flow", lambda k: VAL[k]["fcf27"], NUMFMT),
        ("FY2027E EPS", lambda k: VAL[k]["eps27"], '$0.00'),
        ("EV/Sales multiple", lambda k: FY[k]["ev_sales"], '0.0"x"'),
        ("EV/FCF multiple", lambda k: FY[k]["ev_fcf"], '0.0"x"'),
        ("P/E multiple", lambda k: FY[k]["pe"], '0.0"x"'),
        ("Route 1 — EV/Sales price", lambda k: VAL[k]["px_sales"], '$#,##0.00'),
        ("Route 2 — EV/FCF price", lambda k: VAL[k]["px_fcf"], '$#,##0.00'),
        ("Route 3 — P/E price", lambda k: VAL[k]["px_pe"], '$#,##0.00'),
        ("12-month target (avg of 3)", lambda k: VAL[k]["target"], '$#,##0.00'),
        ("Upside vs $120.23", lambda k: VAL[k]["upside"], '+0.0%;-0.0%')]:
    c0 = ws.cell(row=r, column=1, value=label)
    hi = "target" in label or "Upside" in label
    if hi:
        c0.font = BOLD
    for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k)); c.number_format = fmt; c.border = THIN
        if hi:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="Where the stock trades today").font = BOLD
r += 1
for label, val, fmt in [("EV / FY2026E revenue (base)", ev/FY["Base"]["rev26"], '0.0"x"'),
                        ("EV / FY2027E revenue (base)", ev/FY["Base"]["rev27"], '0.0"x"'),
                        ("Implied multiple at 52wk high $182.19", (182.19*SHARES-NET_CASH)/FY["Base"]["rev26"], '0.0"x"'),
                        ("Implied multiple at street PT $158.41", (158.41*SHARES-NET_CASH)/FY["Base"]["rev26"], '0.0"x"'),
                        ("Implied multiple at 52wk low $94.00", (94.00*SHARES-NET_CASH)/FY["Base"]["rev26"], '0.0"x"')]:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

out = "/Users/antaiwei/Desktop/stock/SHOP_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

# ------------------------------------------------------------ console dump --
print(f"\nSeasonal Q2/Q1 steps: {[f'{x:.1%}' for x in QOQ_STEPS]} -> Q2 = ${Q1_26*sum(QOQ_STEPS)/3/B:.2f}B")
print(f"Trailing y/y: {[f'{x-1:.1%}' for x in YOY_RECENT]}")
print("\n== Q2 2026 forecast (reports Aug 5 pre-market) ==")
print(f"{'':30s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}")
for label, key, _, street in fields:
    vals = "".join(f"{Q2[k][key]/M:>12,.0f}" if abs(Q2[k][key]) > 10 else f"{Q2[k][key]:>12.3f}"
                   for k in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.0f}" if abs(street) > 10 else f"{street:>12.3f}")
    print(f"{label:30s}{vals}{s}")
print("\n== Valuation ==")
for k in ("Bear", "Base", "Bull"):
    v = VAL[k]
    print(f"{k:5s} FY27 rev {FY[k]['rev27']/B:5.1f}B FCF {v['fcf27']/B:4.1f}B EPS ${v['eps27']:4.2f} | "
          f"sales ${v['px_sales']:6.2f} fcf ${v['px_fcf']:6.2f} pe ${v['px_pe']:6.2f} -> ${v['target']:6.2f} ({v['upside']:+.1%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW})")
print(f"EV {ev/B:.1f}B | EV/TTM rev {ev/ttm_rev:.1f}x | EV/TTM FCF {ev/ttm_fcf:.0f}x | "
      f"EV/FY26E rev {ev/FY['Base']['rev26']:.1f}x | EV/FY27E rev {ev/FY['Base']['rev27']:.1f}x")
print(f"TTM FCF margin {ttm_fcf/ttm_rev:.1%} | Rule of 40 = {0.343 + ttm_fcf/ttm_rev:.1%}")
