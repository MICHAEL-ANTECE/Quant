#!/usr/bin/env python3
"""
SNDK (SanDisk Corporation) model — same framework as MU / NBIS / SHOP / APP.

Data = user's moomoo screenshots (2026-08-05, 15:41 ET) + company guidance / street consensus.

FISCAL CALENDAR: SanDisk inherited Western Digital's fiscal year (ends late June / early July).
  2026/Q3 = quarter ended 2026-04-03 (latest reported)
  2026/Q4 = quarter ended 2026-07-03 -> REPORTS TODAY, 2026-08-05, after the close.

TWO LEGACY BREAKS in the moomoo columns:
  1) Columns jump from 2024/Q2 straight back to 2016/Q1 — SanDisk was acquired by Western
     Digital in 2016 and re-spun in Feb 2025, so 2015/Q1-2016/Q1 is the OLD SanDisk, a
     different company. Never trend across the gap.
  2) 2025/Q3 (quarter ended 2025-03-28) carries a -$1.83B goodwill impairment, which is what
     produced the -$13.33 EPS quarter. It is a spin-off artifact, not operations.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M, B = 1e6, 1e9

COLS = ["2026/Q3", "2026/Q2", "2026/Q1", "2025/Q4", "2025/Q3", "2025/Q2", "2025/Q1",
        "2024/Q4", "2024/Q3", "2024/Q2",
        "2016/Q1", "2015/Q4", "2015/Q3", "2015/Q2", "2015/Q1"]
OLD_CO_IDX = 10          # this column and older = pre-WDC-acquisition SanDisk
n = len(COLS)


def row(*vals):
    v = list(vals)
    assert len(v) <= n, f"too many values: {len(v)}"
    return v + [None] * (n - len(v))


# ====================================================== INCOME STATEMENT =====
IS = [
    ("Total Revenue as Reported", row(5.95*B, 3.03*B, 2.31*B, 1.90*B, 1.70*B, 1.88*B, 1.88*B,
                                      1.76*B, 1.71*B, 1.67*B,
                                      1.37*B, 1.54*B, 1.45*B, 1.24*B, 1.33*B)),
    ("Cost of Revenue", row(1.29*B, 1.48*B, 1.62*B, 1.40*B, 1.31*B, 1.27*B, 1.16*B,
                            1.12*B, 1.24*B, 1.50*B,
                            822.41*M, 918.29*M, 849.46*M, 752.82*M, 787.24*M)),
    ("Gross Profit", row(4.66*B, 1.54*B, 687.00*M, 498.00*M, 382.00*M, 606.00*M, 726.00*M,
                         636.00*M, 463.00*M, 161.00*M,
                         543.33*M, 624.86*M, 602.83*M, 484.38*M, 545.00*M)),
    ("Operating Expense", row(498.00*M, 466.00*M, 495.00*M, 447.00*M, 424.00*M, 421.00*M, 413.00*M,
                              415.00*M, 384.00*M, 359.00*M,
                              387.20*M, 404.66*M, 364.55*M, 370.71*M, 386.27*M)),
    ("  Selling and Admin Expenses", row(161.00*M, 139.00*M, 179.00*M, 162.00*M, 139.00*M, 142.00*M, 130.00*M,
                                         117.00*M, 107.00*M, 113.00*M,
                                         136.62*M, 161.47*M, 139.66*M, 138.61*M, 149.87*M)),
    ("  Research & Development", row(337.00*M, 327.00*M, 316.00*M, 285.00*M, 285.00*M, 279.00*M, 283.00*M,
                                     298.00*M, 277.00*M, 246.00*M,
                                     244.19*M, 230.46*M, 211.64*M, 218.42*M, 222.73*M)),
    ("Operating Profit", row(4.16*B, 1.08*B, 192.00*M, 51.00*M, -42.00*M, 185.00*M, 313.00*M,
                             221.00*M, 79.00*M, -198.00*M,
                             156.12*M, 220.20*M, 238.28*M, 113.67*M, 158.73*M)),
    ("Net Non-Operating Interest Income (Expense)", row(6.00*M, -13.00*M, -24.00*M, -30.00*M, -10.00*M, -2.00*M, 1.00*M,
                                                        -6.00*M, -6.00*M, -8.00*M,
                                                        -15.35*M, -27.91*M, -21.43*M, -12.78*M, -23.57*M)),
    ("  Non-Operating Interest Income", row(12.00*M, 12.00*M, 16.00*M, 11.00*M, 6.00*M, 2.00*M, 3.00*M,
                                            3.00*M, 3.00*M, 3.00*M,
                                            8.85*M, 8.64*M, 8.17*M, 8.61*M, 11.03*M)),
    ("  Non-Operating Interest Expense", row(6.00*M, 25.00*M, 40.00*M, 41.00*M, 16.00*M, 4.00*M, 2.00*M,
                                             9.00*M, 9.00*M, 11.00*M,
                                             24.20*M, 36.55*M, 29.60*M, 21.39*M, 34.60*M)),
    ("Other Income (Expense)", row(-63.00*M, -125.00*M, -44.00*M, -39.00*M, -1.85*B, -10.00*M, -47.00*M,
                                   -18.00*M, -19.00*M, -54.00*M,
                                   -19.01*M, -1.64*M, -2.63*M, -9.75*M, -101.54*M)),
    ("  Special Income (Charges)", row(-53.00*M, -10.00*M, -16.00*M, -33.00*M, -1.84*B, 10.00*M, -22.00*M,
                                       -22.00*M, -14.00*M, -47.00*M,
                                       -19.01*M, -2.85*M, -2.63*M, -9.75*M, -101.54*M)),
    ("    – Less: Restructuring and M&A", row(0.0, None, 1.00*M, -3.00*M, 16.00*M, 0.0, 3.00*M,
                                              4.00*M, 1.00*M, 13.00*M,
                                              19.01*M, 2.09*M, 875.00e3, 9.75*M, 40.54*M)),
    ("    – Less: Impairment of Capital Assets", row(0.0, None, 0.0, 0.0, 1.83*B, 0.0, 0.0,
                                                     0.0, 0.0, 0.0,
                                                     0.0, 0.0, 0.0, 0.0, 61.00*M)),
]

# ========================================================= BALANCE SHEET =====
# NOTE: moomoo's balance-sheet tab uses a slightly different column set than the income
# statement (2024/Q4 then jumps to 2016/Q1). Entered on the shared header; the two oldest
# groups are the pre-acquisition company either way.
BS = [
    ("Total Assets", row(17.08*B, 13.00*B, 12.75*B, 12.99*B, 12.96*B, 14.23*B, None,
                         13.51*B, None, None,
                         9.51*B, 9.23*B, 9.07*B, 9.15*B, 9.46*B)),
    ("Total Current Assets", row(9.17*B, 5.15*B, 4.98*B, 5.09*B, 5.09*B, 4.46*B, None,
                                 3.55*B, None, None,
                                 6.15*B, 5.66*B, 3.61*B, 3.66*B, 3.75*B)),
    ("  Cash and Cash Equivalents", row(3.74*B, 1.54*B, 1.44*B, 1.48*B, 1.51*B, 804.00*M, None,
                                        328.00*M, None, None,
                                        3.27*B, 1.48*B, 698.07*M, 685.79*M, 649.94*M)),
    ("  Receivables", row(2.81*B, 1.28*B, 1.27*B, 1.13*B, 1.03*B, 912.00*M, None,
                          1.04*B, None, None,
                          675.75*M, 783.99*M, 906.30*M, 894.63*M, 826.94*M)),
    ("    – Accounts Receivable", row(2.73*B, 1.24*B, 1.19*B, 1.07*B, 979.00*M, 904.00*M, None,
                                      935.00*M, None, None,
                                      497.18*M, 618.19*M, 740.64*M, 640.84*M, 589.25*M)),
    ("    – Taxes Receivable", row(81.00*M, 45.00*M, 72.00*M, 66.00*M, 53.00*M, 8.00*M, None,
                                   7.00*M, None, None,
                                   155.92*M, 142.69*M, 142.61*M, 208.79*M, 168.75*M)),
    ("  Inventory", row(2.24*B, 1.97*B, 1.91*B, 2.08*B, 2.16*B, 2.17*B, None,
                        1.96*B, None, None,
                        881.06*M, 809.40*M, 785.29*M, 780.77*M, 713.05*M)),
    ("  Other Current Assets", row(388.00*M, 357.00*M, 370.00*M, 392.00*M, 391.00*M, 568.00*M, None,
                                   221.00*M, None, None,
                                   44.45*M, 19.64*M, 25.33*M, 34.27*M, 32.34*M)),
    ("Total Non-Current Assets", row(7.91*B, 7.85*B, 7.77*B, 7.90*B, 7.87*B, 9.78*B, None,
                                     9.96*B, None, None,
                                     3.36*B, 3.57*B, 5.46*B, 5.49*B, 5.71*B)),
    ("  Net PPE", row(649.00*M, 631.00*M, 630.00*M, 619.00*M, 603.00*M, 579.00*M, None,
                      791.00*M, None, None,
                      790.40*M, 817.13*M, 831.06*M, 804.76*M, 769.93*M)),
    ("    – Gross PPE", row(2.17*B, 2.13*B, 2.16*B, 2.13*B, 2.10*B, 2.04*B, None,
                            2.99*B, None, None, 2.26*B)),
    ("    – Accumulated Depreciation", row(-1.52*B, -1.50*B, -1.53*B, -1.51*B, -1.49*B, -1.46*B, None,
                                           -2.20*B, None, None, -1.45*B)),
]

# ============================================================= CASH FLOW =====
CF = [
    ("Operating Cash Flow", row(3.04*B, 1.02*B, 488.00*M, 94.00*M, 26.00*M, 95.00*M, -131.00*M,
                                -130.00*M, -12.00*M, None,
                                355.14*M, 433.98*M, 274.97*M, 28.86*M, 308.87*M)),
    ("  Net Income from Continuing Operations", row(3.62*B, 803.00*M, 112.00*M, -23.00*M, -1.93*B, 104.00*M, 211.00*M,
                                                    120.00*M, 27.00*M, None,
                                                    78.35*M, 135.47*M, 133.01*M, 80.97*M, 39.03*M)),
    ("  Depreciation & Depletion & Amortization", row(38.00*M, 38.00*M, 36.00*M, 36.00*M, 37.00*M, 36.00*M, 54.00*M,
                                                      54.00*M, 56.00*M, None,
                                                      139.35*M, 155.37*M, 156.84*M, 156.43*M, 152.46*M)),
    ("  Change in Working Capital", row(-709.00*M, 63.00*M, 250.00*M, 56.00*M, 51.00*M, -143.00*M, -344.00*M,
                                        -353.00*M, -182.00*M, None,
                                        108.47*M, 70.27*M, -26.42*M, -246.66*M, 25.95*M)),
    ("      Change in Receivables", row(-1.49*B, -46.00*M, -125.00*M, -89.00*M, -42.00*M, 133.00*M, -102.00*M,
                                        -120.00*M, -130.00*M, None,
                                        121.45*M, 122.95*M, -99.74*M, -51.19*M, 262.99*M)),
    ("      Change in Inventory", row(-268.00*M, -63.00*M, 172.00*M, 81.00*M, 11.00*M, -103.00*M, -149.00*M,
                                      -225.00*M, -118.00*M, None,
                                      -71.80*M, -23.49*M, -5.11*M, -67.71*M, -13.95*M)),
    ("      Change in Payables and Accrued Expense", row(676.00*M, 62.00*M, 108.00*M, 68.00*M, 14.00*M, 135.00*M, -113.00*M,
                                                         56.00*M, 70.00*M, None,
                                                         861.00e3, 7.40*M, 5.01*M, -4.37*M, -14.27*M)),
    ("      Change in Other Working Capital", row(370.00*M, 110.00*M, 95.00*M, -4.00*M, 68.00*M, -308.00*M, 20.00*M,
                                                  -64.00*M, -4.00*M, None,
                                                  57.96*M, -36.59*M, 73.43*M, -123.39*M, -198.73*M)),
    ("Net cash flow from investing", row(-83.00*M, -165.00*M, -15.00*M, -17.00*M, 404.00*M, 188.00*M, -19.00*M,
                                         -3.00*M, 100.00*M, None,
                                         1.42*B, 342.05*M, 26.76*M, 317.13*M, 341.73*M)),
    ("  Net PPE Purchase and Sale (capex)", row(-45.00*M, -39.00*M, -50.00*M, -45.00*M, -44.00*M, -48.00*M, -67.00*M,
                                                -35.00*M, -29.00*M, None,
                                                -59.46*M, -131.39*M, -88.59*M, -95.56*M, -98.29*M)),
    ("Financing Cash Flow", row(-752.00*M, -758.00*M, -515.00*M, -102.00*M, 276.00*M, 130.00*M, 214.00*M,
                                85.00*M, -135.00*M, None,
                                11.99*M, 5.69*M, -290.28*M, -309.84*M, -808.76*M)),
]

# ======================================================= KEY INDICATORS ======
KI = [
    ("— Profitability (TTM) —", row()),
    ("Gross Margin %", row(56.04, 34.81, 27.93, 30.07, 32.58, 33.65, None, 16.09, None, None,
                           40.29, 40.56, 41.20, 42.74, 44.38)),
    ("Operating Margin %", row(41.58, 14.29, 4.96, 6.89, 9.38, 11.05, None, -6.66, None, None,
                               13.01, 13.13, 14.71, 16.88, 20.54)),
    ("EBIT Margin %", row(39.91, -8.24, -19.55, -19.25, -17.09, 9.90, None, -6.95, None, None,
                          13.08, 11.76, 13.40, 15.28, 19.30)),
    ("Net Margin %", row(34.19, -11.66, -22.37, -22.31, -20.77, 6.40, None, -10.09, None, None,
                         7.64, 6.98, 7.90, 9.66, 12.06)),
    ("EBITDA Margin %", row(41.03, -6.60, -17.69, -17.04, -14.58, 12.67, None, -3.59, None, None,
                            23.94, 22.92, 24.27, 25.53, 28.58)),
    ("Tax Rate %", row(12.49, None, None, None, None, 33.14, None, None, None, None,
                       30.75, 26.65, 29.20, 27.04, 30.15)),
    ("Interest Coverage (x)", row(46.98, -6.03, -15.06, -22.48, -39.77, 29.79, None, -11.58, None, None,
                                  6.41, 5.25, 5.97, 7.50, 9.48)),
    ("R&D Expense Ratio %", row(9.59, 13.58, 14.97, 15.39, 15.87, 15.74, None, 15.92, None, None,
                                16.16, 15.87, 15.27, 14.72, 13.59)),
    ("— Solvency —", row()),
    ("Long-Term Debt to Equity %", row(1.32, 7.53, 16.19, 21.94, 23.17, 1.49, None, 1.54, None, None,
                                       None, 21.57, 37.94, 21.75, 20.75)),
    ("Equity Ratio %", row(80.69, 78.57, 73.58, 70.97, 70.69, 84.31, None, 82.05, None, None,
                           56.04, 62.17, 62.10, 61.45, 61.70)),
    ("Debt to Asset Ratio %", row(1.32, 7.73, 16.41, 22.16, 23.39, 6.10, None, 8.89, None, None,
                                  40.82, 37.48, 37.94, 37.61, 35.83)),
    ("Current Ratio (x)", row(4.78, 3.11, 3.29, 3.56, 3.70, 2.38, None, 1.67, None, None,
                              1.67, 2.82, 3.17, 1.80, 1.80)),
    ("Quick Ratio (x)", row(3.41, 1.71, 1.79, 1.83, 1.85, 0.92, None, 0.65, None, None,
                            1.41, 2.39, 2.29, 1.31, 1.35)),
    ("— Operating Capacity (TTM) —", row()),
    ("Cash Conversion Cycle (days)", row(165.24, 149.96, 154.20, 167.19, 184.38, 186.71, None, 155.54, None, None,
                                         84.75, 90.93, 90.35, 83.87, 77.13)),
    ("Receivable Turnover (x)", row(7.12, 8.33, 6.52, 7.34, 7.37, 7.99, None, 7.13, None, None,
                                    10.31, 7.62, 7.15, 8.64, 10.87)),
    ("Inventory Turnover (x)", row(2.64, 2.81, 2.94, 2.55, 2.25, 2.21, None, 2.86, None, None,
                                   4.19, 4.39, 4.32, 4.52, 4.74)),
]

# ================================================ ANALYST RATINGS (screenshot)
CONSENSUS_RATING = dict(rating="Strong Buy", analysts=12, updated="2026-08-01",
                        buy=83.33, hold=16.67, sell=0.00,
                        pt_high=3050.00, pt_avg=2397.27, pt_low=1620.00, current=1386.00)
RATINGS = [
    ("SIG", "Buy", 3250, 3050, "Downgrade (PT cut)", "2026-07-22"),
    ("Wells Fargo", "Hold", 1250, 1620, "Upgrade", "2026-07-21"),
    ("Bernstein", "Buy", 3000, 3000, "Maintained", "2026-07-20"),
    ("Argus Research", "Hold", None, None, "New", "2026-07-14"),
    ("Goldman Sachs", "Buy", 1200, 2200, "Upgrade", "2026-07-05"),
    ("BofA Securities", "Buy", 2100, 2500, "Upgrade", "2026-07-01"),
    ("Citi", "Buy", 2025, 2500, "Upgrade", "2026-06-25"),
    ("Cantor Fitzgerald", "Buy", 1800, 2900, "Upgrade", "2026-06-08"),
    ("Mizuho Securities", "Buy", 1825, 2200, "Upgrade", "2026-06-07"),
    ("Morgan Stanley", "Buy", 1100, 1750, "Upgrade", "2026-06-03"),
    ("Barclays", "Buy", 1200, 2300, "Upgrade", "2026-05-26"),
    ("Melius Research", "Buy", 1350, 2350, "Upgrade", "2026-05-18"),
    ("Jefferies", "Buy", 1000, 1400, "Upgrade", "2026-05-01"),
    ("Evercore", "Buy", 1200, 1400, "Upgrade", "2026-05-01"),
    ("Raymond James", "Buy", 725, 1470, "Upgrade", "2026-05-01"),
]

# INSIDER ACTIVITY (screenshot 3): 1 bullish vs 8 bearish insiders, last 6 months
INSIDERS = [
    ("Bernard Shek", "Chief Legal Officer", "2026-06-30", -600, 2088.00),
    ("Bernard Shek", "Chief Legal Officer", "2026-06-19", -117, 2184.75),
    ("Bernard Shek", "Chief Legal Officer", "2026-06-02", -600, 1736.00),
    ("Alper Ilkbahar", "Executive Vice President", "2026-06-01", -2694, None),
    ("Alper Ilkbahar", "Executive Vice President", "2026-05-31", -2000, 1756.86),
    ("David V. Goeckeler", "Chairman & CEO", "2026-05-24", -1569, 1478.69),
    ("Alper Ilkbahar", "Executive Vice President", "2026-05-24", -653, 1478.69),
    ("Bernard Shek", "Chief Legal Officer", "2026-05-24", -211, 1478.69),
    ("Luis Visoso", "Chief Financial Officer", "2026-05-20", -1588, 1542.24),
    ("David V. Goeckeler", "Chairman & CEO", "2026-05-20", -2331, 1467.40),
    ("Thomas Caulfield", "Independent Director", "2026-05-11", -9666, None),
    ("Michael Pokorny", "Vice President", "2026-05-11", -2446, 1426.18),
    ("Michael Pokorny", "Vice President", "2026-05-08", -1429, 1562.34),
    ("Necip Sayiner", "Independent Director", "2026-05-07", -579, 1503.11),
]

# ========================================================== MARKET DATA ======
PX_NOW = 1386.00
SHARES = 148.09e6
MKT = [
    ("Last price (2026-08-05, 15:41 ET)", 1386.000, "USD, -2.92%"),
    ("Previous close", 1427.620, "USD"),
    ("Open", 1400.580, "USD"), ("Day High", 1441.758, "USD"), ("Day Low", 1364.030, "USD"),
    ("Day Range %", 5.45, "%"),
    ("Volume", 10.52e6, "shares"), ("Turnover", 14.81*B, "USD"),
    ("Average Price", 1407.825, "USD"), ("Turnover Rate %", 7.20, "%"),
    ("Vol Ratio", 0.56, "x"), ("Bid/Ask %", -36.84, "% (ask-heavy)"),
    ("Market Cap", 205.25*B, "USD"), ("Float Market Cap", 202.36*B, "USD"),
    ("Total Shares", 148.09e6, "shares"), ("Free Float", 146.0e6, "shares"),
    ("P/E (TTM)", 47.37, "x"), ("P/E (LFY)", None, "LOSS"), ("P/B", 14.898, "x"),
    ("52-week High", 2354.390, "USD"), ("52-week Low", 42.820, "USD"),
    ("Historical High", 2354.390, "USD"), ("Historical Low", 27.885, "USD"),
    ("Next earnings", None, "2026-08-05 AFTER CLOSE (today)"),
    ("Money flow — NET OUTFLOW", -218.59*M, "inflow 1,236.36M / outflow 1,454.96M"),
    ("  Extra Large net", (61.04-64.26)*M, "in 61.04M / out 64.26M"),
    ("  Large net", (169.32-219.03)*M, "in 169.32M / out 219.03M"),
    ("  Medium net", (221.69-265.23)*M, "in 221.69M / out 265.23M"),
    ("  Small net", (784.31-906.43)*M, "in 784.31M / out 906.43M"),
]

STREET = [
    ("FQ4 FY2026 report", "2026-08-05, after the close (quarter ended 2026-07-03)", "company"),
    ("FQ4 revenue GUIDANCE", "$7.75B - $8.25B", "company, given at FQ3"),
    ("FQ4 non-GAAP EPS GUIDANCE", "$30.00 - $33.00", "company"),
    ("FQ4 consensus revenue", "$8.30B (+337% y/y) — ABOVE guidance high", "Zacks"),
    ("FQ4 consensus EPS", "$34.24 (Zacks) / $34.80 (MarketWatch) / $35.45 (Visible Alpha) — ALL above guidance", "street"),
    ("Analyst mean PT (moomoo)", "$2,397.27 (12 analysts, Strong Buy)", "screenshot"),
    ("Analyst PT range", "$1,620 low / $3,050 high", "screenshot"),
    ("Rating split", "83.33% Buy / 16.67% Hold / 0% Sell", "screenshot"),
    ("Insider signal", "1 bullish vs 8 bearish insiders in 6 months; CEO, CFO, CLO, EVP and 2 directors all sold May-Jun at $1,400-2,185", "screenshot"),
    ("Driver", "NAND shortage + AI-inference demand for high-capacity eSSD; contract pricing repricing upward every quarter", "company / street"),
]

# ============================================ FQ4-2026 FORECAST ENGINE ========
Q3_REV = 5.95*B
Q3_GM = 4.66/5.95
GUIDE_LO, GUIDE_HI = 7.75*B, 8.25*B
GUIDE_EPS_LO, GUIDE_EPS_HI = 30.00, 33.00
ST_REV, ST_EPS = 8.30*B, 34.24
DILUTED_Q3 = 3.615*B / 23.03      # implied diluted share count from reported EPS

SCEN = {
    "Bear": dict(rev=8.25*B, gm=0.800, opex=520*M, nint=8*M, other=-60*M, tax=0.140, sh=158.5e6),
    "Base": dict(rev=8.50*B, gm=0.820, opex=540*M, nint=10*M, other=-40*M, tax=0.130, sh=158.0e6),
    "Bull": dict(rev=9.00*B, gm=0.840, opex=560*M, nint=12*M, other=-30*M, tax=0.125, sh=157.5e6),
}
SBC_PER_SHARE = 0.60      # non-GAAP add-back estimate (SBC + amortisation)


def forecast_q4(p):
    gp = p["rev"] * p["gm"]
    op = gp - p["opex"]
    pretax = op + p["nint"] + p["other"]
    net = pretax * (1 - p["tax"])
    eps = net / p["sh"]
    return dict(revenue=p["rev"], qoq=p["rev"]/Q3_REV-1, gross_profit=gp, gross_margin=p["gm"],
                opex=p["opex"], op_profit=op, op_margin=op/p["rev"], pretax=pretax,
                net=net, eps_gaap=eps, eps_nongaap=eps+SBC_PER_SHARE)


Q4 = {k: forecast_q4(v) for k, v in SCEN.items()}

# ================================= FY2027 CYCLE MODEL + VALUATION ============
# FY2026 = 2.31 + 3.03 + 5.95 + FQ4
FY26_REV_BASE = (2.31 + 3.03 + 5.95)*B + Q4["Base"]["revenue"]
FY26_EPS_BASE = 0.75 + 5.15 + 23.03 + Q4["Base"]["eps_gaap"]

FY = {
    # FY2027 = fiscal year ending ~July 2027. This is a COMMODITY CYCLE, so the multiple
    # applied to peak earnings is deliberately low — memory peaks trade at trough P/Es.
    "Bear": dict(rev27=28.0*B, net_m=0.45, pe=7.0, shares=160e6,
                 note="NAND pricing plateaus H1-27 then rolls; supply responds"),
    "Base": dict(rev27=40.0*B, net_m=0.55, pe=9.0, shares=159e6,
                 note="shortage persists through FY27, pricing up then flat"),
    "Bull": dict(rev27=48.0*B, net_m=0.60, pe=12.0, shares=158e6,
                 note="AI eSSD demand structural, long-term agreements lock pricing"),
}
PROB = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}


def valuation(p):
    net27 = p["rev27"] * p["net_m"]
    eps27 = net27 / p["shares"]
    target = eps27 * p["pe"]
    return dict(net27=net27, eps27=eps27, target=target, upside=target/PX_NOW-1)


VAL = {k: valuation(v) for k, v in FY.items()}
PW_TARGET = sum(PROB[k]*VAL[k]["target"] for k in PROB)

# ============================================================== WRITE ========
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="2E5C8A")
LEG = PatternFill("solid", fgColor="4A4A4A")
SEC = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")]*4)
NUMFMT = '#,##0.0,,;[Red](#,##0.0,,)'
PCT = '0.0%'


def style_header(ws, r, labels, width0=46, width=13):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=r, column=j, value=lab)
        c.fill = HDR if j == 1 else (LEG if j-2 >= OLD_CO_IDX else SUB)
        c.font = WHITE
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = width0
    for j in range(2, len(labels)+1):
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=r+1, column=2)


def write_stmt(wb, title, data, note):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=9, color="7F7F7F")
    style_header(ws, 2, ["USD, millions"] + COLS)
    r = 3
    for name, vals in data:
        c = ws.cell(row=r, column=1, value=name)
        if not name.startswith(" "):
            c.font = BOLD
        for j, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=j, value=v)
            cell.number_format = NUMFMT
            cell.border = THIN
        r += 1


wb = Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 112
for i, (a, b) in enumerate([
    ("SNDK — SanDisk Corporation model", ""),
    ("Built", "2026-08-05, 15:41 ET, price $1,386.00"),
    ("Primary data source", "moomoo screenshots: Analyst Ratings / Key Indicators / Income Statement / Balance Sheet / Cash Flow / Insiders / quote panel + money flow"),
    ("FISCAL CALENDAR", "SanDisk inherited Western Digital's fiscal year (ends late Jun / early Jul). 2026/Q3 = quarter ended 2026-04-03 (latest reported); 2026/Q4 = quarter ended 2026-07-03, REPORTS TODAY after the close."),
    ("Latest REPORTED quarter", "FQ3 2026: revenue $5.95B (+251% y/y, +96% q/q), gross margin 78.3%, operating profit $4.16B (69.9%), net $3.62B, EPS $23.03"),
    ("LEGACY BREAK 1", "Columns jump from 2024/Q2 back to 2016/Q1 (grey). Western Digital acquired SanDisk in 2016 and re-spun it in Feb 2025 — the 2015-2016 columns are a different company. Never trend across the gap."),
    ("LEGACY BREAK 2", "2025/Q3 (quarter ended 2025-03-28) carries a -$1.83B goodwill impairment, producing the -$13.33 EPS quarter. Spin-off artifact, not operations."),
    ("What this model is NOT", "This is a commodity memory cycle, not a compounder. The valuation sheet deliberately applies LOW multiples to PEAK earnings — the standard memory-cycle convention. A low forward P/E here is a warning, not a bargain."),
], start=1):
    ws.cell(row=i, column=1, value=a).font = BOLD
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

write_stmt(wb, "IS_Quarterly", IS, "moomoo > Financials > Income Statement > Quarterly · All · USD.")
write_stmt(wb, "BS_Quarterly", BS, "moomoo > Financials > Balance Sheet > Quarterly · All · USD. moomoo's balance-sheet tab uses a slightly different column set; blanks are where the two sets do not line up.")
write_stmt(wb, "CF_Quarterly", CF, "moomoo > Financials > Cash Flow > Quarterly · All · USD.")

ws = wb.create_sheet("Key_Indicators_TTM")
ws.cell(row=1, column=1, value="moomoo > Financials > Key Indicators. TTM basis.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["TTM indicator"] + COLS)
r = 3
for name, vals in KI:
    c = ws.cell(row=r, column=1, value=name)
    if name.startswith("—"):
        c.font = BOLD
        for j in range(1, n+2):
            ws.cell(row=r, column=j).fill = SEC
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = '0.00'
        cell.border = THIN
    r += 1

ws = wb.create_sheet("Derived_Quarterly")
ws.cell(row=1, column=1, value="Computed from the entered statements. Ignore the 2015-2016 columns.").font = Font(italic=True, size=9, color="7F7F7F")
style_header(ws, 2, ["Derived metric"] + COLS)
rev = dict(IS)["Total Revenue as Reported"]
cogs = dict(IS)["Cost of Revenue"]
gp = dict(IS)["Gross Profit"]
op = dict(IS)["Operating Profit"]
ocf = dict(CF)["Operating Cash Flow"]
capex = dict(CF)["  Net PPE Purchase and Sale (capex)"]
fin = dict(CF)["Financing Cash Flow"]
ni = dict(CF)["  Net Income from Continuing Operations"]
ar = dict(BS)["    – Accounts Receivable"]
inv = dict(BS)["  Inventory"]


def safe(f, i):
    try:
        return f(i)
    except (TypeError, ZeroDivisionError):
        return None


drv = [
    ("Revenue q/q %", [safe(lambda i: rev[i]/rev[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Gross margin % (quarterly)", [safe(lambda i: gp[i]/rev[i], i) for i in range(n)]),
    ("Cost of revenue (absolute)", cogs),
    ("Cost of revenue q/q %", [safe(lambda i: cogs[i]/cogs[i+1]-1, i) if i+1 < n else None for i in range(n)]),
    ("Operating margin % (quarterly)", [safe(lambda i: op[i]/rev[i], i) for i in range(n)]),
    ("Opex as % of revenue", [safe(lambda i: (gp[i]-op[i])/rev[i], i) for i in range(n)]),
    ("Free cash flow (OCF + capex)", [safe(lambda i: ocf[i]+capex[i], i) for i in range(n)]),
    ("Financing CF (negative = debt paydown / returns)", fin),
    ("Accounts receivable", ar),
    ("AR / quarterly revenue (x)", [safe(lambda i: ar[i]/rev[i], i) for i in range(n)]),
    ("Inventory", inv),
    ("Inventory / quarterly COGS (x)", [safe(lambda i: inv[i]/cogs[i], i) for i in range(n)]),
    ("Net income", ni),
]
r = 3
for name, vals in drv:
    ws.cell(row=r, column=1, value=name)
    for j, v in enumerate(vals, start=2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.number_format = PCT if "%" in name else ('0.00"x"' if "(x)" in name else NUMFMT)
        cell.border = THIN
    r += 1

ws = wb.create_sheet("Analyst_Ratings")
ws.column_dimensions["A"].width = 28
for col, w in zip("BCDEF", (12, 12, 12, 20, 14)):
    ws.column_dimensions[col].width = w
ws.cell(row=1, column=1, value="moomoo > Valuation > Analyst Ratings (updated 2026-08-01).").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for k_, v_ in [("Consensus rating", CONSENSUS_RATING["rating"]),
               ("Analysts covering", CONSENSUS_RATING["analysts"]),
               ("Buy / Hold / Sell %", f"{CONSENSUS_RATING['buy']} / {CONSENSUS_RATING['hold']} / {CONSENSUS_RATING['sell']}"),
               ("Target — High", CONSENSUS_RATING["pt_high"]),
               ("Target — Average", CONSENSUS_RATING["pt_avg"]),
               ("Target — Low", CONSENSUS_RATING["pt_low"]),
               ("Current price", CONSENSUS_RATING["current"]),
               ("Implied upside to mean PT", CONSENSUS_RATING["pt_avg"]/CONSENSUS_RATING["current"]-1),
               ("Price vs 52wk high $2,354.39", CONSENSUS_RATING["current"]/2354.39-1),
               ("Price vs 52wk low $42.82", CONSENSUS_RATING["current"]/42.82-1)]:
    ws.cell(row=r, column=1, value=k_).font = BOLD
    c = ws.cell(row=r, column=2, value=v_)
    if "upside" in k_ or "vs 52wk" in k_:
        c.number_format = '+0.0%;-0.0%'
    r += 1
r += 1
for j, h in enumerate(["Institution", "Rating", "Prior PT", "New PT", "Action", "Date"], start=1):
    c = ws.cell(row=r, column=j, value=h); c.fill = HDR; c.font = WHITE
r += 1
for name, rating, old, new, action, date in RATINGS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=rating)
    ws.cell(row=r, column=3, value=old)
    ws.cell(row=r, column=4, value=new)
    c = ws.cell(row=r, column=5, value=action)
    if old and new and new > old:
        c.font = Font(color="008000")
    elif old and new and new < old:
        c.font = Font(color="C00000")
    ws.cell(row=r, column=6, value=date)
    r += 1
r += 2
ws.cell(row=r, column=1, value="INSIDER ACTIVITY — 1 bullish vs 8 bearish insiders, last 6 months").font = WHITE
for j in range(1, 6):
    ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="8B0000")
r += 1
for j, h in enumerate(["Insider", "Relation", "Date", "Shares", "Price"], start=1):
    c = ws.cell(row=r, column=j, value=h); c.fill = HDR; c.font = WHITE
r += 1
for name, rel, date, sh, px in INSIDERS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=rel)
    ws.cell(row=r, column=3, value=date)
    c = ws.cell(row=r, column=4, value=sh); c.font = Font(color="C00000")
    c = ws.cell(row=r, column=5, value=px)
    if px:
        c.number_format = '$#,##0.00'
    r += 1

ws = wb.create_sheet("Market_Data")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 46
ws.cell(row=1, column=1, value="moomoo quote panel + Trade Overview, 2026-08-05 15:41 ET.").font = Font(italic=True, size=9, color="7F7F7F")
r = 3
for name, val, unit in MKT:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = '#,##0.000' if (val is not None and abs(val) < 10000) else NUMFMT
    ws.cell(row=r, column=3, value=unit)
    r += 1
r += 1
ws.cell(row=r, column=1, value="DERIVED").font = WHITE
ws.cell(row=r, column=1).fill = HDR
r += 1
ttm_rev = (5.95+3.03+2.31+1.90)*B
ttm_ni = (3.62+0.803+0.112-0.023)*B
book = 205.25*B/14.898
for name, val, fmt in [
        ("Book value = Mkt cap / P/B", book, NUMFMT),
        ("Book value per share", book/SHARES, '$#,##0.00'),
        ("TTM revenue", ttm_rev, NUMFMT),
        ("TTM net income", ttm_ni, NUMFMT),
        ("EV (no net debt — company is debt-free)", 205.25*B - 3.74*B, NUMFMT),
        ("EV / TTM revenue", (205.25*B-3.74*B)/ttm_rev, '0.0"x"'),
        ("Implied diluted shares (FQ3 net / EPS)", DILUTED_Q3, '#,##0'),
        ("FY2026E revenue (base)", FY26_REV_BASE, NUMFMT),
        ("FY2026E EPS (base)", FY26_EPS_BASE, '$#,##0.00'),
        ("P/E on FY2026E", PX_NOW/FY26_EPS_BASE, '0.0"x"')]:
    ws.cell(row=r, column=1, value=name)
    c = ws.cell(row=r, column=2, value=val); c.number_format = fmt
    r += 1

ws = wb.create_sheet("Guidance_Street")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 26
ws.cell(row=1, column=1, value="Company guidance + street consensus (web, 2026-08-05). Kept separate from reported data.").font = Font(italic=True, size=9, color="C00000")
r = 3
for a, b, c_ in STREET:
    ws.cell(row=r, column=1, value=a).font = BOLD
    ws.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=3, value=c_).font = Font(size=9, color="7F7F7F")
    r += 1

ws = wb.create_sheet("Forecast_FQ4_2026")
ws.column_dimensions["A"].width = 40
for col in "BCDEF":
    ws.column_dimensions[col].width = 15
ws.cell(row=1, column=1, value=f"FQ3 base: revenue $5.95B, gross margin {Q3_GM:.1%}, opex $498M, EPS $23.03 on {DILUTED_Q3/1e6:.1f}M diluted shares.").font = Font(italic=True, size=9, color="7F7F7F")
ws.cell(row=2, column=1, value="NOTE: consensus ($8.30B / $34.24) sits ABOVE company guidance ($7.75-8.25B / $30-33). The bar is already raised.").font = Font(italic=True, size=9, color="C00000")
r = 4
ws.cell(row=r, column=1, value="FQ4 FY2026 (reports today AMC)").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k_); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=r, column=5, value="Street"); c.fill = SUB; c.font = WHITE
c = ws.cell(row=r, column=6, value="Guide (hi)"); c.fill = SUB; c.font = WHITE
r += 1
fields = [("Revenue", "revenue", NUMFMT, ST_REV, GUIDE_HI),
          ("Revenue q/q", "qoq", PCT, ST_REV/Q3_REV-1, GUIDE_HI/Q3_REV-1),
          ("Gross profit", "gross_profit", NUMFMT, None, None),
          ("Gross margin", "gross_margin", PCT, None, None),
          ("Operating expense", "opex", NUMFMT, None, None),
          ("Operating profit", "op_profit", NUMFMT, None, None),
          ("Operating margin", "op_margin", PCT, None, None),
          ("Pre-tax profit", "pretax", NUMFMT, None, None),
          ("Net income", "net", NUMFMT, None, None),
          ("EPS (GAAP)", "eps_gaap", '0.00', None, None),
          ("EPS (non-GAAP est.)", "eps_nongaap", '0.00', ST_EPS, GUIDE_EPS_HI)]
for label, key, fmt, street, guide in fields:
    ws.cell(row=r, column=1, value=label)
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q4[k_][key]); c.number_format = fmt; c.border = THIN
    if street is not None:
        c = ws.cell(row=r, column=5, value=street); c.number_format = fmt
    if guide is not None:
        c = ws.cell(row=r, column=6, value=guide); c.number_format = fmt
    r += 1
r += 1
for label, key, base in [("vs street revenue $8.30B", "revenue", ST_REV),
                         ("vs guidance high $8.25B", "revenue", GUIDE_HI),
                         ("vs street EPS $34.24", "eps_nongaap", ST_EPS),
                         ("vs guidance high EPS $33.00", "eps_nongaap", GUIDE_EPS_HI)]:
    ws.cell(row=r, column=1, value=label).font = BOLD
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=Q4[k_][key]/base-1); c.number_format = '+0.0%;-0.0%'
    r += 1

ws = wb.create_sheet("Valuation")
ws.column_dimensions["A"].width = 46
for col in "BCD":
    ws.column_dimensions[col].width = 20
ws.cell(row=1, column=1, value="COMMODITY CYCLE VALUATION. Peak earnings get LOW multiples — memory names trade at their cheapest P/E at the top of a pricing cycle. Applying a growth multiple here is the classic mistake.").font = Font(italic=True, size=9, color="C00000")
r = 3
ws.cell(row=r, column=1, value="Scenario").fill = HDR
ws.cell(row=r, column=1).font = WHITE
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=k_); c.fill = HDR; c.font = WHITE; c.alignment = Alignment(horizontal="center")
r += 1
for label, fn, fmt in [
        ("FY2026E revenue", lambda k_: FY26_REV_BASE, NUMFMT),
        ("FY2026E EPS", lambda k_: FY26_EPS_BASE, '$#,##0.00'),
        ("FY2027E revenue", lambda k_: FY[k_]["rev27"], NUMFMT),
        ("FY2027E net margin", lambda k_: FY[k_]["net_m"], PCT),
        ("FY2027E net income", lambda k_: VAL[k_]["net27"], NUMFMT),
        ("FY2027E EPS", lambda k_: VAL[k_]["eps27"], '$#,##0.00'),
        ("P/E applied to PEAK earnings", lambda k_: FY[k_]["pe"], '0.0"x"'),
        ("12-month target", lambda k_: VAL[k_]["target"], '$#,##0.00'),
        ("Upside vs $1,386", lambda k_: VAL[k_]["upside"], '+0.0%;-0.0%')]:
    c0 = ws.cell(row=r, column=1, value=label)
    hi = "target" in label or "Upside" in label
    if hi:
        c0.font = BOLD
    for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
        c = ws.cell(row=r, column=j, value=fn(k_)); c.number_format = fmt; c.border = THIN
        if hi:
            c.font = BOLD
            c.fill = SEC
    r += 1
r += 1
ws.cell(row=r, column=1, value="Cycle thesis").font = BOLD
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=FY[k_]["note"]); c.alignment = Alignment(wrap_text=True, vertical="top")
r += 2
ws.cell(row=r, column=1, value="Scenario probability").font = BOLD
for j, k_ in enumerate(("Bear", "Base", "Bull"), start=2):
    c = ws.cell(row=r, column=j, value=PROB[k_]); c.number_format = PCT
r += 1
ws.cell(row=r, column=1, value="PROBABILITY-WEIGHTED 12-MONTH TARGET").font = Font(bold=True, size=12)
c = ws.cell(row=r, column=2, value=PW_TARGET); c.number_format = '$#,##0.00'; c.font = Font(bold=True, size=12); c.fill = SEC
c = ws.cell(row=r, column=3, value=PW_TARGET/PX_NOW-1); c.number_format = '+0.0%'; c.font = BOLD
r += 2
ws.cell(row=r, column=1, value="SENSITIVITY: FY2027E EPS x multiple").font = BOLD
r += 1
mults = [6, 8, 10, 12, 14]
ws.cell(row=r, column=1, value="FY27 EPS \\ P/E").font = BOLD
for j, m_ in enumerate(mults, start=2):
    c = ws.cell(row=r, column=j, value=f"{m_}x"); c.fill = SUB; c.font = WHITE
r += 1
for eps_ in (80, 110, 140, 170, 200):
    c = ws.cell(row=r, column=1, value=eps_); c.number_format = '$#,##0'; c.fill = SUB; c.font = WHITE
    for j, m_ in enumerate(mults, start=2):
        c = ws.cell(row=r, column=j, value=eps_*m_)
        c.number_format = '$#,##0'
        if abs(eps_*m_ - PX_NOW) < 150:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    r += 1

out = "/Users/antaiwei/Desktop/stock/SNDK_Financial_Model.xlsx"
wb.save(out)
print("saved:", out)

print(f"\nFQ3 actual: rev $5.95B, GM {Q3_GM:.1%}, op $4.16B ({4.16/5.95:.1%}), EPS $23.03, diluted {DILUTED_Q3/1e6:.1f}M")
print(f"Guidance FQ4: ${GUIDE_LO/B:.2f}-{GUIDE_HI/B:.2f}B rev / ${GUIDE_EPS_LO}-{GUIDE_EPS_HI} EPS")
print(f"Street FQ4:   ${ST_REV/B:.2f}B rev / ${ST_EPS} EPS  <-- ABOVE guidance on both")
print("\n== FQ4 FY2026 forecast (reports today AMC) ==")
print(f"{'':32s}{'Bear':>12s}{'Base':>12s}{'Bull':>12s}{'Street':>12s}{'Guide hi':>12s}")
for label, key, _, street, guide in fields:
    vals = "".join(f"{Q4[k_][key]/M:>12,.0f}" if abs(Q4[k_][key]) > 100 else f"{Q4[k_][key]:>12.3f}"
                   for k_ in ("Bear", "Base", "Bull"))
    s = "" if street is None else (f"{street/M:>12,.0f}" if abs(street) > 100 else f"{street:>12.3f}")
    g = "" if guide is None else (f"{guide/M:>12,.0f}" if abs(guide) > 100 else f"{guide:>12.3f}")
    print(f"{label:32s}{vals}{s}{g}")
print(f"\nFY2026E (base): revenue ${FY26_REV_BASE/B:.2f}B, EPS ${FY26_EPS_BASE:.2f} -> P/E {PX_NOW/FY26_EPS_BASE:.1f}x")
print("\n== Cycle valuation ==")
for k_ in ("Bear", "Base", "Bull"):
    v = VAL[k_]
    print(f"{k_:5s} FY27 rev {FY[k_]['rev27']/B:5.1f}B  net {v['net27']/B:5.1f}B  EPS ${v['eps27']:7.2f} "
          f"x {FY[k_]['pe']:4.1f}x -> ${v['target']:8.2f} ({v['upside']:+.1%})")
print(f"\nPROBABILITY-WEIGHTED TARGET: ${PW_TARGET:.2f} ({PW_TARGET/PX_NOW-1:+.1%} vs ${PX_NOW:.2f})")
print(f"Street mean PT $2,397.27 implies FY27 P/E of {2397.27/VAL['Base']['eps27']:.1f}x on base-case EPS")
